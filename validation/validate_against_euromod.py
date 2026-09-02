"""Validate the app against DEVMOD v1.1 run via EUROMOD.

Part A compares run_simulation output person-by-person against EUROMOD's
output microdata (baseline and reform). Part B compares the app's headline
statistics for analysis choices 3 and 4 against the two reference Statistics
Presenter Excels.

Run from the repo root (or anywhere):  venv/bin/python validation/validate_against_euromod.py
"""

import os
import sys

import numpy as np
import openpyxl
import pandas as pd

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, REPO_ROOT)
os.chdir(REPO_ROOT)

import app  # noqa: E402  (imports the Dash app module; no server is started)

REF_BASELINE = 'ref/DEVMOD v1.1/Output/dv_2023_std.txt'
REF_REFORM = 'ref/DEVMOD v1.1/Output/dv_2023_reform_std.txt'
REF_EXCEL_CONS = ('ref/Reform results - consumption based net of indirect taxes/'
                  'DEVMOD 2023 vs 2023_reform run via EUROMOD where reform includes same policy edits - cons based.xlsx')
REF_EXCEL_INC = ('ref/Reform results - income based net of indirect taxes/'
                 'DEVMOD 2023 vs 2023_reform run via EUROMOD where reform includes same policy edits - inc based.xlsx')

# The DV_2023_reform policy edits (verified against DV.xml parameter by parameter)
REFORM_PARAMS = dict(
    app.BASELINE_PARAMS,
    bsa_2_person=300,            # 276 -> 300
    bsa_3_plus_person=450,       # 386 -> 450
    senior_grant_amount=90,      # 76 -> 90
    school_meal_value=110,       # 80 -> 110
    school_meal_age=16,          # <18 -> <16
    pit_bracket2_rate=0.03,      # 0.05 -> 0.03
    pit_bracket3_thresh=800,     # 1000 -> 800
    pit_bracket5_rate=0.30,      # 0.25 -> 0.30
    pit_yse_turnover_threshold=4000,  # 5000 -> 4000 (presumptive band top moves with it)
    presumptive_rate_4=0.04,     # 0.03 -> 0.04
    vat_items_list=app.BASELINE_VAT_STD_RATE_ITEMS + ['x0111', 'x0116'],
)

# Part A columns: (column, rtol). Consumption-chain columns carry uprating
# rounding, so they get a looser relative tolerance.
PART_A_COLS = [
    ('tin_s', 1e-6), ('ttn_s', 1e-6), ('ttb01_s', 1e-6), ('ttb02_s', 1e-6),
    ('ttb_s', 1e-6), ('tscee_s', 1e-6), ('tscer_s', 1e-6),
    ('bsa_s', 1e-6), ('boa_s', 1e-6), ('bed_s', 1e-6),
    ('tva_s', 1e-4), ('il_exp_vat', 1e-4), ('ses', 1e-6),
    ('spl', 1e-6), ('splpf', 1e-6),
    ('ils_origy', 1e-6), ('ils_tax', 1e-6), ('ils_ben', 1e-6), ('ils_benki', 1e-6),
    ('ils_dispy', 1e-6), ('ils_dispyx', 1e-6), ('ils_taxco', 1e-4),
    ('ils_con', 1e-4), ('ils_con_pf', 1e-4), ('ils_dispyx_pf', 1e-4),
    ('xhh_s', 1e-4),
]
# The reference microdata are rounded to 2 decimals, so allow half a cent.
PART_A_ATOL = 0.006

failures = []
checks = 0


def check(name, actual, expected, rtol=1e-4, atol=1e-3):
    global checks
    checks += 1
    actual = float(actual)
    expected = float(expected)
    ok = abs(actual - expected) <= max(atol, rtol * abs(expected))
    status = 'ok  ' if ok else 'FAIL'
    if not ok:
        failures.append((name, actual, expected))
        print(f"  {status} {name}: app={actual:.6f} ref={expected:.6f} diff={actual - expected:+.6f}")
    return ok


# ---------------------------------------------------------------- Part A
def part_a():
    print("=" * 70)
    print("Part A: person-level microdata comparison")
    print("=" * 70)
    input_df = app.ensure_input_dataframe()
    scenarios = [
        ('baseline', app.BASELINE_PARAMS, REF_BASELINE),
        ('reform', REFORM_PARAMS, REF_REFORM),
    ]
    for label, params, ref_path in scenarios:
        sim = app.run_simulation(input_df, params)
        ref = pd.read_csv(ref_path, sep='\t')
        merged = sim.merge(ref, on='idperson', suffixes=('_app', '_ref'), how='inner')
        assert len(merged) == len(ref) == len(sim), (
            f"{label}: row mismatch app={len(sim)} ref={len(ref)} merged={len(merged)}")
        print(f"\n{label}: {len(merged)} persons matched on idperson")
        bad_cols = 0
        for col, rtol in PART_A_COLS:
            a_col, r_col = f'{col}_app', f'{col}_ref'
            if a_col not in merged.columns or r_col not in merged.columns:
                print(f"  SKIP {col}: missing "
                      f"({'app' if a_col not in merged.columns else 'ref'})")
                continue
            a = pd.to_numeric(merged[a_col], errors='coerce').fillna(0.0).to_numpy()
            r = pd.to_numeric(merged[r_col], errors='coerce').fillna(0.0).to_numpy()
            tol = np.maximum(PART_A_ATOL, rtol * np.abs(r))
            bad = np.abs(a - r) > tol
            n_bad = int(bad.sum())
            if n_bad:
                bad_cols += 1
                failures.append((f"{label}/{col}", n_bad, 0))
                worst = np.argsort(-(np.abs(a - r)))[:5]
                print(f"  FAIL {col}: {n_bad} persons out of tolerance; worst:")
                for i in worst[:5]:
                    print(f"        idperson={merged['idperson'].iloc[i]} "
                          f"app={a[i]:.4f} ref={r[i]:.4f}")
        if bad_cols == 0:
            print(f"  ok   all {len(PART_A_COLS)} columns within tolerance")


# ---------------------------------------------------------------- Part B
def sheet_rows(wb, sheet):
    """Yield (label, base_value, reform_value) for rows with a text label."""
    ws = wb[sheet]
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=3):
        label = row[0].value
        if isinstance(label, str):
            yield label.strip(), row[1].value, row[2].value


def get_row(wb, sheet, label, occurrence=0):
    hits = [(b, r) for lab, b, r in sheet_rows(wb, sheet) if lab == label]
    if occurrence >= len(hits):
        raise KeyError(f"{sheet}!'{label}' occurrence {occurrence} not found ({len(hits)} hits)")
    return hits[occurrence]


def get_prefix_rows(wb, sheet, prefix):
    return [(b, r) for lab, b, r in sheet_rows(wb, sheet) if lab.startswith(prefix)]


def part_b():
    print()
    print("=" * 70)
    print("Part B: headline statistics vs reference Excels")
    print("=" * 70)
    input_df = app.ensure_input_dataframe()
    reform_sim = app.run_simulation(input_df, REFORM_PARAMS)

    for choice, excel_path, concept in ((3, REF_EXCEL_CONS, 'cons'), (4, REF_EXCEL_INC, 'inc')):
        print(f"\n--- choice {choice} ({concept} based, net of indirect taxes) ---")
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        base_art = app.get_baseline_artifacts(input_df, choice)
        base_res = base_art['results']
        ref_res, _ = app.run_analysis(reform_sim, choice, base_art['merge_df'])

        # Tax-benefit policy (yearly, millions)
        for row_label, key in [
            ('Sum of government revenue', 'Sum of government revenue'),
            ('- Direct taxes', 'Direct taxes'),
            ('- Social insurance contributions', 'Social insurance contributions'),
            ('- Indirect taxes', 'Indirect taxes'),
            ('Sum of government expenditure', 'Sum of government expenditure'),
            ('- Cash benefits', 'Cash benefits'),
            ('- In-kind benefits', 'In-kind benefits'),
            ('- Indirect subsidies', 'Indirect subsidies'),
        ]:
            b, r = get_row(wb, 'Tax-ben policy', row_label)
            check(f"taxbenpol/{key}/base", base_res['taxbenpol_abs'][key] * 12 / 1e6, b, rtol=1e-5)
            check(f"taxbenpol/{key}/reform", ref_res['taxbenpol_abs'][key] * 12 / 1e6, r, rtol=1e-5)

        # Poverty: All individuals rate (1st occurrence) and gap (2nd)
        b, r = get_row(wb, 'Poverty', 'All individuals', occurrence=0)
        check("poverty/rate/base", base_res['poverty']['All individuals']['Poverty rate (%)'], b)
        check("poverty/rate/reform", ref_res['poverty']['All individuals']['Poverty rate (%)'], r)
        b, r = get_row(wb, 'Poverty', 'All individuals', occurrence=1)
        check("poverty/gap/base", base_res['poverty']['All individuals']['Poverty gap (%)'], b)
        check("poverty/gap/reform", ref_res['poverty']['All individuals']['Poverty gap (%)'], r)
        b, _ = get_row(wb, 'Poverty', 'Absolute national poverty line, yearly')
        check("poverty/line_yearly", base_res['poverty']['povline'] * 12, b)

        # Inequality
        b, r = get_row(wb, 'Inequality', '- Gini coefficient')
        check("inequality/Gini/base", base_res['inequality']['Gini'], b)
        check("inequality/Gini/reform", ref_res['inequality']['Gini'], r)
        b, r = get_row(wb, 'Inequality', '- Atkinson inequality index (ineq. aversion = 0.25)')
        check("inequality/Atkinson/base", base_res['inequality']['Atkinson'], b)
        check("inequality/Atkinson/reform", ref_res['inequality']['Atkinson'], r)
        b, r = get_row(wb, 'Inequality', '- P80/P20 ratio')
        for res, ref_v, tag in ((base_res, b, 'base'), (ref_res, r, 'reform')):
            p = res['inequality']['Percentiles']
            check(f"inequality/P80P20/{tag}", p[80] / p[20] if p[20] else 0, ref_v)
        for pct in (10, 20, 30, 40, 50, 60, 70, 80, 90):
            label = '- 50th (median)' if pct == 50 else f'- {pct}th'
            b, r = get_row(wb, 'Inequality', label)
            check(f"inequality/P{pct}/base", base_res['inequality']['Percentiles'][pct] * 12, b)
            check(f"inequality/P{pct}/reform", ref_res['inequality']['Percentiles'][pct] * 12, r)
        for dec in range(1, 11):
            b, r = get_row(wb, 'Inequality', f'- Decile {dec}')
            for res, ref_v, tag in ((base_res, b, 'base'), (ref_res, r, 'reform')):
                share = (res['inequality'][f'SumEqRank_InBaselineDec{dec}']
                         / res['inequality']['TotalEqRank'] * 100)
                check(f"inequality/decile_share{dec}/{tag}", share, ref_v)

        # Individual and household decile counts
        for dec in range(1, 11):
            b, r = get_row(wb, 'Individuals', f'- Decile {dec}')
            check(f"individuals/decile{dec}/base", base_res['individuals'][f'CountIndDecile{dec}'], b, atol=0.5)
            check(f"individuals/decile{dec}/reform", ref_res['individuals'][f'CountIndDecile{dec}'], r, atol=0.5)
            b, r = get_row(wb, 'Households', f'- Decile {dec}')
            check(f"households/decile{dec}/base", base_res['households'][f'CountHHDecile{dec}'], b, atol=0.5)
            check(f"households/decile{dec}/reform", ref_res['households'][f'CountHHDecile{dec}'], r, atol=0.5)

        # Policy effects: the four "before" rows, in order rate, gap, Gini, P80/P20
        before_rows = get_prefix_rows(wb, 'Policy effects', '- Before taxes and benefits')
        assert len(before_rows) == 4, f"expected 4 before-rows, found {len(before_rows)}"
        check("before/poverty_rate", base_res['PovertyRate_Bef'], before_rows[0][0])
        check("before/poverty_gap", base_res['PovertyGap_Bef'], before_rows[1][0])
        check("before/gini", base_res['Gini_Bef'], before_rows[2][0])
        p80p20_bef = base_res['P80_Bef'] / base_res['P20_Bef'] if base_res['P20_Bef'] else 0
        check("before/p80p20", p80p20_bef, before_rows[3][0])

        # Benefit adequacy: means and shares of the constant baseline denominators
        b, r = get_row(wb, 'Benefits', '- Mean cash benefit amount per beneficiary, yearly')
        check("adequacy/mean_cash/base", base_res['benefits']['Mean_eq_indiv_cash_ben_yearly'], b)
        check("adequacy/mean_cash/reform", ref_res['benefits']['Mean_eq_indiv_cash_ben_yearly'], r)
        b, r = get_row(wb, 'Benefits', '- Mean in-kind benefit amount per beneficiary, yearly')
        check("adequacy/mean_inkind/base", base_res['benefits']['Mean_eq_indiv_inkind_ben_yearly'], b)
        check("adequacy/mean_inkind/reform", ref_res['benefits']['Mean_eq_indiv_inkind_ben_yearly'], r)
        for row_label, mean_key, den_key in [
            ('- Mean cash benefit amount as a share of median consumption, %',
             'Mean_eq_indiv_cash_ben_yearly', 'BaselineMedianEqConsYearly'),
            ('- Mean cash benefit amount as a share of median disposable income, %',
             'Mean_eq_indiv_cash_ben_yearly', 'BaselineMedianEqIncYearly'),
            ('- Mean in-kind benefit amount as a share of median consumption, %',
             'Mean_eq_indiv_inkind_ben_yearly', 'BaselineMedianEqConsYearly'),
            ('- Mean in-kind benefit amount as a share of median disposable income, %',
             'Mean_eq_indiv_inkind_ben_yearly', 'BaselineMedianEqIncYearly'),
        ]:
            b, r = get_row(wb, 'Benefits', row_label)
            for res, ref_v, tag in ((base_res, b, 'base'), (ref_res, r, 'reform')):
                den = res[den_key]
                share = res['benefits'][mean_key] / den * 100 if den else 0
                check(f"adequacy/{row_label[2:40].strip()}/{tag}", share, ref_v)

        # Adequacy denominators constant across scenarios and choices
        check(f"adequacy/denominator_cons/ch{choice}", base_res['BaselineMedianEqConsYearly'], 14317.32, atol=0.1)
        check(f"adequacy/denominator_inc/ch{choice}", base_res['BaselineMedianEqIncYearly'], 3826.20, atol=0.1)
        check(f"adequacy/denominator_cons_reform/ch{choice}", ref_res['BaselineMedianEqConsYearly'], 14317.32, atol=0.1)
        check(f"adequacy/denominator_inc_reform/ch{choice}", ref_res['BaselineMedianEqIncYearly'], 3826.20, atol=0.1)


def main():
    part_a()
    part_b()
    print()
    print("=" * 70)
    if failures:
        print(f"RESULT: {len(failures)} of {checks} checks FAILED")
        sys.exit(1)
    print(f"RESULT: all {checks} checks passed")


if __name__ == '__main__':
    main()
