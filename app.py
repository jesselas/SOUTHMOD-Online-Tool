import dash
import gc
import dash_bootstrap_components as dbc
from dash import dcc, html
from dash.dependencies import Input, Output, State, ALL, MATCH
from dash import callback_context
import pandas as pd
import numpy as np
import textwrap
import re
import html as html_utils
from datetime import datetime
from io import BytesIO
import openpyxl
import plotly.graph_objects as go
import json
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from decimal import Decimal, InvalidOperation
import math

# --- APPLICATION INITIALIZATION ---
# Use a Bootstrap theme for a clean layout
app = dash.Dash(__name__, external_stylesheets=[dbc.themes.BOOTSTRAP], suppress_callback_exceptions=True)
app.title = "SOUTHMOD Online Tool"
server = app.server

# Performance optimizations for deployment
app.scripts.config.serve_locally = True
app.css.config.serve_locally = True

# --- DATA LOADING ---
INPUT_FILE = 'dv_2020_a1.txt'

try:
    INPUT_DF = pd.read_csv(INPUT_FILE, sep=r'\s+', low_memory=False)
except FileNotFoundError:
    INPUT_DF = None
BASELINE_CACHE = {}
DEFAULT_REFORM_NAME = "My reform"
RESULTS_TITLE_TEXT = "Baseline vs. Reform"

# Ready-made demo reform presets. Individually each moves poverty or Gini by
# more than 1 point; combined they are approximately budget-neutral
# (+300M revenue vs +315M expenditure yearly) and progressive (Gini -3.9).
PRESET_REFORMS = {
    'tax': {
        'label': 'Raise taxes',
        'description': (
            "Raises the standard VAT rate by two percentage points and makes personal "
            "income tax more progressive, lifting the marginal rates of the top three "
            "brackets. Small enterprises pay a higher presumptive rate. The package "
            "raises ca. 300 million a year."
        ),
        'params': {
            'tva_rate': 0.18,            # 0.16 -> 0.18
            'pit_bracket3_rate': 0.13,   # 0.10 -> 0.13
            'pit_bracket4_rate': 0.24,   # 0.20 -> 0.24
            'pit_bracket5_rate': 0.31,   # 0.25 -> 0.31
            'presumptive_rate_4': 0.05,  # 0.03 -> 0.05
        },
    },
    'benefits': {
        'label': 'Increase benefits',
        'description': (
            "Raises every social assistance amount, the senior citizens' grant and the "
            "value of school meals by roughly a fifth. On its own the package costs "
            "about 315 million a year, some 21% more government expenditure."
        ),
        'params': {
            'bsa_1_person': 195,         # 165 -> 195
            'bsa_2_person': 330,         # 276 -> 330
            'bsa_3_plus_person': 460,    # 386 -> 460
            'bsa_disabled_topup': 100,   # 80 -> 100
            'senior_grant_amount': 90,   # 76 -> 90
            'school_meal_value': 100,    # 80 -> 100
        },
    },
}

# --- VAT ITEM DEFINITIONS ---
# Dictionary mapping item codes to labels and baseline status
VAT_ITEM_MAP = {
    'x0111': {'label': 'Bread and cereals', 'baseline_vattable': False},
    'x0112': {'label': 'Meat', 'baseline_vattable': True},
    'x0113': {'label': 'Fish and seafood', 'baseline_vattable': True},
    'x0114': {'label': 'Milk, cheese and eggs', 'baseline_vattable': True},
    'x0115': {'label': 'Oils and fats', 'baseline_vattable': True},
    'x0116': {'label': 'Fruit', 'baseline_vattable': False},
    'x0117': {'label': 'Vegetables', 'baseline_vattable': False},
    'x0118': {'label': 'Sugar, jam, honey, chocolate', 'baseline_vattable': True},
    'x0119': {'label': 'Food products n.e.c.', 'baseline_vattable': True},
    'x0121': {'label': 'Coffee, tea and cocoa', 'baseline_vattable': True},
    'x0122': {'label': 'Mineral waters, soft drinks, juices', 'baseline_vattable': True},
    'x0211': {'label': 'Spirits', 'baseline_vattable': True},
    'x0212': {'label': 'Wine', 'baseline_vattable': True},
    'x0213': {'label': 'Beer', 'baseline_vattable': True},
    'x0230': {'label': 'Narcotics', 'baseline_vattable': True},
    'x0311': {'label': 'Clothing materials', 'baseline_vattable': True},
    'x0312': {'label': 'Garments', 'baseline_vattable': True},
    'x0313': {'label': 'Other clothing, clothing accessories', 'baseline_vattable': True},
    'x0314': {'label': 'Cleaning, repair and hire of clothing', 'baseline_vattable': True},
    'x0321': {'label': 'Shoes and other footwear', 'baseline_vattable': True},
    'x0322': {'label': 'Repair and hire of footwear', 'baseline_vattable': True},
    'x0411': {'label': 'Actual rentals paid by tenants', 'baseline_vattable': True},
    'x0412': {'label': 'Other actual rentals', 'baseline_vattable': True},
    'x0431': {'label': 'Dwelling repair and maintenance materials', 'baseline_vattable': True},
    'x0432': {'label': 'Services for dwelling repair and maintenance', 'baseline_vattable': True},
    'x0441': {'label': 'Water supply', 'baseline_vattable': True},
    'x0442': {'label': 'Refuse collection', 'baseline_vattable': True},
    'x0443': {'label': 'Sewerage collection', 'baseline_vattable': True},
    'x0444': {'label': 'Other dwelling-related services', 'baseline_vattable': True},
    'x0451': {'label': 'Electricity', 'baseline_vattable': True},
    'x0452': {'label': 'Gas', 'baseline_vattable': True},
    'x0453': {'label': 'Liquid fuels', 'baseline_vattable': True},
    'x0454': {'label': 'Solid fuels', 'baseline_vattable': True},
    'x0455': {'label': 'Heat energy', 'baseline_vattable': True},
    'x0511': {'label': 'Furniture and furnishings', 'baseline_vattable': True},
    'x0512': {'label': 'Carpets, other floor coverings', 'baseline_vattable': True},
    'x0513': {'label': 'Repair of furniture and floor coverings', 'baseline_vattable': True},
    'x0531': {'label': 'Major household appliances', 'baseline_vattable': True},
    'x0532': {'label': 'Small electric household appliances', 'baseline_vattable': True},
    'x0533': {'label': 'Repair of household appliances', 'baseline_vattable': True},
    'x0551': {'label': 'Major tools and equipment', 'baseline_vattable': True},
    'x0552': {'label': 'Small tools and misc accessories', 'baseline_vattable': True},
    'x0561': {'label': 'Non-durable household goods', 'baseline_vattable': True},
    'x0562': {'label': 'Domestic services, household services', 'baseline_vattable': True},
    'x0611': {'label': 'Pharmaceutical products', 'baseline_vattable': False},
    'x0612': {'label': 'Other medical products', 'baseline_vattable': False},
    'x0613': {'label': 'Therapeutic appliances and equipment', 'baseline_vattable': False},
    'x0621': {'label': 'Medical services', 'baseline_vattable': False},
    'x0622': {'label': 'Dental services', 'baseline_vattable': False},
    'x0623': {'label': 'Paramedical services', 'baseline_vattable': False},
    'x0711': {'label': 'Motor cars', 'baseline_vattable': True},
    'x0712': {'label': 'Motor cycles', 'baseline_vattable': True},
    'x0713': {'label': 'Bicycles', 'baseline_vattable': True},
    'x0714': {'label': 'Animal drawn vehicles', 'baseline_vattable': True},
    'x0721': {'label': 'Accessories for personal transport equipment (PTE)', 'baseline_vattable': True},
    'x0722': {'label': 'Fuels and lubricants for PTE', 'baseline_vattable': True},
    'x0723': {'label': 'Maintenance and repair of PTE', 'baseline_vattable': True},
    'x0724': {'label': 'Other services in respect of PTE', 'baseline_vattable': True},
    'x0731': {'label': 'Passenger transport by railway', 'baseline_vattable': True},
    'x0732': {'label': 'Passenger transport by road', 'baseline_vattable': True},
    'x0733': {'label': 'Passenger transport by air', 'baseline_vattable': True},
    'x0734': {'label': 'Passenger transport by water', 'baseline_vattable': True},
    'x0735': {'label': 'Combined passenger transport', 'baseline_vattable': True},
    'x0810': {'label': 'Postal services', 'baseline_vattable': True},
    'x0911': {'label': 'Equipment relating to sound and picture', 'baseline_vattable': True},
    'x0912': {'label': 'Photographic and cinematographic equipment', 'baseline_vattable': True},
    'x0921': {'label': 'Major durables for outdoor recreation', 'baseline_vattable': True},
    'x0922': {'label': 'Instruments and durables for indoor recreation', 'baseline_vattable': True},
    'x0923': {'label': 'Maintenance of other durables for recreation', 'baseline_vattable': True},
    'x0931': {'label': 'Games, toys and hobbies', 'baseline_vattable': True},
    'x0932': {'label': 'Equipment for sports and open-air recreation', 'baseline_vattable': True},
    'x0941': {'label': 'Recreational and sporting services', 'baseline_vattable': True},
    'x0951': {'label': 'Books', 'baseline_vattable': True},
    'x0960': {'label': 'Package holidays', 'baseline_vattable': True},
    'x1111': {'label': 'Restaurants and cafés', 'baseline_vattable': True},
    'x1112': {'label': 'Canteens', 'baseline_vattable': True},
    'x1211': {'label': 'Hairdressing salons and similar', 'baseline_vattable': True},
    'x1212': {'label': 'Electrical appliances for personal care', 'baseline_vattable': True},
    'x1213': {'label': 'Other products for personal care', 'baseline_vattable': True},
}

# Baseline list of vatable items, derived from the map
TOTAL_VAT_ITEMS = len(VAT_ITEM_MAP)
BASELINE_VAT_STD_RATE_ITEMS = [k for k, v in VAT_ITEM_MAP.items() if v['baseline_vattable']]


# --- BASELINE PARAMETERS (2023) ---
BASELINE_PARAMS = {
    'basic_pov_line': 120, 'upper_pov_line': 150,
    'basic_pov_line_pf': 109, 'upper_pov_line_pf': 136,
    'tscee_rate': 0.05, 'tscer_rate': 0.10, 'tva_rate': 0.16,
    'presumptive_turnover_1': 200, 'presumptive_tax_1': 0,
    'presumptive_turnover_2': 400, 'presumptive_tax_2': 12,
    'presumptive_turnover_3': 1200, 'presumptive_tax_3': 24,
    'presumptive_rate_4': 0.03,
    'pit_yse_turnover_threshold': 5000, 'pit_yag_exemption': 300,
    'pit_bracket1_thresh': 0, 'pit_bracket1_rate': 0.0,
    'pit_bracket2_thresh': 500, 'pit_bracket2_rate': 0.05,
    'pit_bracket3_thresh': 1000, 'pit_bracket3_rate': 0.10,
    'pit_bracket4_thresh': 1500, 'pit_bracket4_rate': 0.20,
    'pit_bracket5_thresh': 2000, 'pit_bracket5_rate': 0.25,
    'bsa_income_threshold': 441, 'bsa_1_person': 165, 'bsa_2_person': 276,
    'bsa_3_plus_person': 386, 'bsa_disabled_topup': 80,
    'senior_grant_age': 55, 'senior_grant_income_threshold': 221, 'senior_grant_amount': 76,
    'school_meal_value': 80, 'school_meal_age': 18,
    'vat_items_list': BASELINE_VAT_STD_RATE_ITEMS
}

# Hard per-field limits ('min'/'max'): rates 0-1; monetary amounts 0-10x their
# baseline value; annual thresholds 0-50,000; ages bounded to sensible ranges.
PARAM_INPUT_META = {
    'tva_rate': {'precision': 3, 'step': 0.01, 'min': 0, 'max': 1},
    'tscee_rate': {'precision': 3, 'step': 0.01, 'min': 0, 'max': 1},
    'tscer_rate': {'precision': 3, 'step': 0.01, 'min': 0, 'max': 1},
    'presumptive_rate_4': {'precision': 3, 'step': 0.01, 'min': 0, 'max': 1},
    'pit_bracket1_rate': {'precision': 3, 'disabled': True, 'step': 0.01, 'min': 0, 'max': 1},
    'pit_bracket2_rate': {'precision': 3, 'step': 0.01, 'min': 0, 'max': 1},
    'pit_bracket3_rate': {'precision': 3, 'step': 0.01, 'min': 0, 'max': 1},
    'pit_bracket4_rate': {'precision': 3, 'step': 0.01, 'min': 0, 'max': 1},
    'pit_bracket5_rate': {'precision': 3, 'step': 0.01, 'min': 0, 'max': 1},
    'pit_bracket1_thresh': {'precision': 0, 'thousands': True, 'disabled': True, 'step': 1, 'min': 0, 'max': 50000},
    'pit_bracket2_thresh': {'precision': 2, 'thousands': True, 'step': 1, 'min': 0, 'max': 50000},
    'pit_bracket3_thresh': {'precision': 2, 'thousands': True, 'step': 1, 'min': 0, 'max': 50000},
    'pit_bracket4_thresh': {'precision': 2, 'thousands': True, 'step': 1, 'min': 0, 'max': 50000},
    'pit_bracket5_thresh': {'precision': 2, 'thousands': True, 'step': 1, 'min': 0, 'max': 50000},
    'pit_yse_turnover_threshold': {'precision': 2, 'thousands': True, 'step': 1, 'min': 0, 'max': 50000},
    'pit_yag_exemption': {'precision': 2, 'thousands': True, 'step': 1, 'min': 0, 'max': 3000},
    'presumptive_turnover_1': {'precision': 2, 'thousands': True, 'step': 1, 'min': 0, 'max': 50000},
    'presumptive_turnover_2': {'precision': 2, 'thousands': True, 'step': 1, 'min': 0, 'max': 50000},
    'presumptive_turnover_3': {'precision': 2, 'thousands': True, 'step': 1, 'min': 0, 'max': 50000},
    'presumptive_tax_2': {'precision': 2, 'thousands': True, 'step': 1, 'min': 0, 'max': 120},
    'presumptive_tax_3': {'precision': 2, 'thousands': True, 'step': 1, 'min': 0, 'max': 240},
    'bsa_income_threshold': {'precision': 2, 'thousands': True, 'step': 1, 'min': 0, 'max': 4410},
    'bsa_1_person': {'precision': 2, 'thousands': True, 'step': 1, 'min': 0, 'max': 1650},
    'bsa_2_person': {'precision': 2, 'thousands': True, 'step': 1, 'min': 0, 'max': 2760},
    'bsa_3_plus_person': {'precision': 2, 'thousands': True, 'step': 1, 'min': 0, 'max': 3860},
    'bsa_disabled_topup': {'precision': 2, 'thousands': True, 'step': 1, 'min': 0, 'max': 800},
    'senior_grant_age': {'precision': 0, 'thousands': False, 'force_int': True, 'step': 1, 'min': 18, 'max': 100},
    'senior_grant_income_threshold': {'precision': 2, 'thousands': True, 'step': 1, 'min': 0, 'max': 2210},
    'senior_grant_amount': {'precision': 2, 'thousands': True, 'step': 1, 'min': 0, 'max': 760},
    'school_meal_age': {'precision': 0, 'force_int': True, 'step': 1, 'min': 0, 'max': 25},
    'school_meal_value': {'precision': 2, 'thousands': True, 'step': 1, 'min': 0, 'max': 800},
}

POLICY_PARAM_SECTIONS = [
    {
        'title': 'Personal income tax',
        'prefix': 'Personal income tax – ',
        'items': [
            ('pit_yse_turnover_threshold', 'Self-employment income threshold (presumptive maximum), annual'),
            ('pit_yag_exemption', 'Exemption on agricultural income, annual'),
            ('pit_bracket2_thresh', 'Bracket 2 lower threshold, annual'),
            ('pit_bracket2_rate', 'Bracket 2 progressive rate, %/100'),
            ('pit_bracket3_thresh', 'Bracket 3 lower threshold, annual'),
            ('pit_bracket3_rate', 'Bracket 3 progressive rate, %/100'),
            ('pit_bracket4_thresh', 'Bracket 4 lower threshold, annual'),
            ('pit_bracket4_rate', 'Bracket 4 progressive rate, %/100'),
            ('pit_bracket5_thresh', 'Bracket 5 lower threshold, annual'),
            ('pit_bracket5_rate', 'Bracket 5 progressive rate, %/100'),
        ],
    },
    {
        'title': 'Social insurance contributions',
        'prefix': 'Social insurance contributions – ',
        'items': [
            ('tscee_rate', 'Employee contribution rate, %/100'),
            ('tscer_rate', 'Employer contribution rate, %/100'),
        ],
    },
    {
        'title': 'Presumptive tax for micro enterprises',
        'prefix': 'Presumptive tax for micro enterprises – ',
        'items': [
            ('presumptive_turnover_1', 'Band 2 lower threshold, annual'),
            ('presumptive_tax_2', 'Band 2 tax amount, annual'),
            ('presumptive_turnover_2', 'Band 3 lower threshold, annual'),
            ('presumptive_tax_3', 'Band 3 tax amount, annual'),
        ],
    },
    {
        'title': 'Presumptive tax for small enterprises',
        'prefix': 'Presumptive tax for small enterprises – ',
        'items': [
            ('presumptive_turnover_3', 'Lower threshold, annual'),
            ('presumptive_rate_4', 'Tax rate, %/100'),
        ],
    },
    {
        'title': 'Value-added tax (VAT)',
        'prefix': 'Value-added tax (VAT) – ',
        'items': [
            ('tva_rate', 'Standard VAT rate, %/100'),
        ],
    },
    {
        'title': 'Social assistance benefit',
        'prefix': 'Social assistance – ',
        'items': [
            ('bsa_income_threshold', 'Income threshold, monthly'),
            ('bsa_1_person', 'Benefit amount (1-person household), monthly'),
            ('bsa_2_person', 'Benefit amount (2-person household), monthly'),
            ('bsa_3_plus_person', 'Benefit amount (3+-person household), monthly'),
            ('bsa_disabled_topup', 'Disability top-up, monthly'),
        ],
    },
    {
        'title': "Senior citizens' grant",
        'prefix': 'Senior grant – ',
        'items': [
            ('senior_grant_age', 'Eligibility age threshold'),
            ('senior_grant_income_threshold', 'Eligibility income threshold, monthly'),
            ('senior_grant_amount', 'Grant amount, monthly'),
        ],
    },
    {
        'title': 'School meals (in-kind)',
        'prefix': 'School meals (in-kind) – ',
        'items': [
            ('school_meal_age', 'School meal age threshold'),
            ('school_meal_value', 'School meal value, monthly'),
        ],
    },
]

POLICY_PARAM_LOOKUP = {
    param_id: {
        'label': label,
        'section': section['title'],
        'prefix': section.get('prefix', '')
    }
    for section in POLICY_PARAM_SECTIONS
    for (param_id, label) in section['items']
}

def policy_values_equal(param_id, baseline_value, reform_value, tol=1e-9):
    if baseline_value is None and reform_value is None:
        return True
    if baseline_value is None or reform_value is None:
        return False
    if isinstance(baseline_value, (int, float)) and isinstance(reform_value, (int, float)):
        # rtol=0: numpy's default relative tolerance would treat 5,000.04 as equal
        # to 5,000, which the client-side "changed" highlight does not
        return np.isclose(float(baseline_value), float(reform_value), rtol=0, atol=tol)
    return str(baseline_value) == str(reform_value)

def format_policy_value(param_id, value):
    if value is None:
        return "n/a"
    if param_id in POLICY_PARAM_LOOKUP:
        formatted = format_param_value(param_id, value)
        return formatted if formatted else "0"
    if isinstance(value, (int, float)):
        formatted = f"{value:,.4f}".rstrip('0').rstrip('.')
        return formatted if formatted else "0"
    return str(value)

def get_param_meta(param_id: str) -> dict:
    base_meta = {
        'precision': 2,
        'thousands': False,
        'allow_negative': False,
        'strip_trailing': True,
        'force_int': False,
        'disabled': False,
        'step': 1,
    }
    base_meta.update(PARAM_INPUT_META.get(param_id, {}))
    return base_meta


def clamp_param_value(param_id: str, value):
    """Clamp a parsed numeric value to the parameter's hard min/max limits."""
    if value is None:
        return None
    meta = get_param_meta(param_id)
    min_v = meta.get('min')
    max_v = meta.get('max')
    if min_v is not None and value < min_v:
        value = min_v
    if max_v is not None and value > max_v:
        value = max_v
    if meta.get('force_int'):
        value = int(round(value))
    return value

def format_param_value(param_id: str, raw_value) -> str:
    if raw_value is None or raw_value == "":
        return ""
    meta = get_param_meta(param_id)
    if isinstance(raw_value, str):
        candidate = raw_value.replace(",", "").strip()
    else:
        candidate = str(raw_value)
    try:
        numeric_value = float(candidate)
    except (TypeError, ValueError):
        return str(raw_value)
    precision = meta['precision']
    thousands = meta['thousands']
    if precision is None:
        formatted = f"{int(round(numeric_value))}"
    else:
        pattern = f"{{:,.{precision}f}}" if thousands else f"{{:.{precision}f}}"
        formatted = pattern.format(numeric_value)
        if meta['strip_trailing'] and '.' in formatted:
            formatted = formatted.rstrip('0').rstrip('.')
    if thousands and '.' not in formatted and precision == 0:
        formatted = f"{int(round(numeric_value)):,}"
    return formatted

def parse_param_input_value(param_id: str, value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if not isinstance(value, str):
        return None
    cleaned = value.replace(",", "").strip()
    if cleaned in {"", ".", "-"}:
        return None
    try:
        parsed = float(cleaned)
    except ValueError:
        return None
    meta = get_param_meta(param_id)
    if meta.get('force_int'):
        return int(round(parsed))
    return parsed

def create_param_input_component(param_id: str, value, disabled: bool = False):
    meta = get_param_meta(param_id)
    formatted_value = format_param_value(param_id, value)
    input_kwargs = {
        'id': {'type': 'param-input', 'index': param_id},
        'value': formatted_value,
        'type': 'text',
        'style': {'width': '100%'},
        'className': "form-control form-control-sm param-input-field",
        'inputmode': 'decimal' if meta.get('precision', 0) and meta['precision'] > 0 else 'numeric',
        'disabled': disabled or meta.get('disabled'),
    }
    input_element = dbc.Input(**input_kwargs)
    if input_kwargs['disabled']:
        return input_element

    dec_button = dbc.Button(
        "−",
        id={'type': 'param-step', 'index': param_id, 'direction': 'dec'},
        color="light",
        size="sm",
        className="param-step-btn"
    )
    inc_button = dbc.Button(
        "+",
        id={'type': 'param-step', 'index': param_id, 'direction': 'inc'},
        color="light",
        size="sm",
        className="param-step-btn"
    )

    return html.Div(
        [
            dec_button,
            input_element,
            inc_button,
        ],
        className="param-input-wrapper"
    )

# --- INFO MODAL CONTENT ---
INFO_MODAL_CONTENT = {
    'taxbenpol': {
        'title': 'About: Tax-benefit policy',
        'body': '''<p style="line-height:1.5">This tab presents an overview of government revenues and expenditures as included in the model, shown in two tables. The first table presents the yearly revenue and expenditure amounts in millions of national currency. The second table shows revenue by source and expenditure by type as shares of total revenue and expenditure (%). In addition to the baseline results, each reform scenario includes the corresponding outcomes and their absolute differences from the baseline values.</p>
<p style="line-height:1.5"><b><span style='color:#0070C0'>Sum of government revenue:</span></b> The total of direct taxes, social insurance contributions, and indirect taxes.</p>
<p style="line-height:1.5"><b><span style='color:#0070C0'>By source:</span></b> Breakdown of revenue components. This categorization is mutually exclusive.</p>
<p style="line-height:1.2; margin-left:20px;"><b>Direct taxes:</b> Taxes levied directly on income or wealth (e.g., personal income tax). Refer to income list <i>ils_tax</i> in the model.</p>
<p style="line-height:1.2; margin-left:20px;"><b>Social insurance contributions:</b> Contributions from employees, employers, and self-employed. Refer to income list <i>ils_sic</i>.</p>
<p style="line-height:1.2; margin-left:20px;"><b>Indirect taxes:</b> Taxes on goods and services (e.g., VAT and excise duties). Refer to income list <i>ils_taxco</i>.</p>
<p style="line-height:1.5"><b><span style='color:#0070C0'>Sum of government expenditure:</span></b> The total of cash benefits, in-kind benefits, and indirect subsidies.</p>
<p style="line-height:1.5"><b><span style='color:#0070C0'>By type:</span></b> Breakdown of expenditure by the type of the transfer. This categorization is mutually exclusive.</p>
<p style="line-height:1.2; margin-left:20px;"><b>Cash benefits:</b> Direct monetary transfers to households/individuals. Refer to income list <i>ils_ben</i>.</p>
<p style="line-height:1.2; margin-left:20px;"><b>In-kind benefits:</b> Benefits provided as goods or services rather than cash. Refer to income list <i>ils_benki</i>. Please note that in-kind benefits are not modelled in all countries.</p>
<p style="line-height:1.2; margin-left:20px;"><b>Indirect subsidies:</b> Subsidies that reduce the price of goods/services (e.g., fuel subsidies). Refer to income list <i>ils_benco</i>. Please note that indirect subsidies are not modelled in all countries.</p>
<p style="line-height:0.1"><b><span style='color:#ffffff'>_</span></b></p>'''
    },
    'poverty': {
        'title': 'About: Poverty',
        'body': '''<p style="line-height:1.5">This tab shows poverty rates and gaps for the total population and for individuals living in different types of households. All calculations are performed at the individual level. In addition to the baseline results, each reform scenario includes the corresponding outcomes and their absolute differences from the baseline values.</p>
      <p style="line-height:1.5"><b><span style='color:#0070C0'>Poverty rate:</span></b> An individual is defined as poor if their household’s disposable income or consumption (an equivalized value calculated for each individual) falls below the poverty line. This measure is also known as the poverty headcount ratio, or the Foster-Greer-Thorbecke (FGT) index for alpha=0, FGT(0).</p>
      <p style="line-height:1.5"><b><span style='color:#0070C0'>Poverty gap:</span></b> This measures the average intensity of poverty. For each poor individual, the distance to the poverty line is calculated as a percentage of that line. The final index is the average of these values over the entire population (both poor and non-poor). This corresponds to the Foster-Greer-Thorbecke (FGT) index for alpha=1, FGT(1).</p>
      <p style="line-height:1.5"><b><span style='color:#0070C0'>Household structure:</span></b> These categories classify households based on their composition of adults (aged 18 and over) and children (under 18):</p>
	  <p style="line-height:1.2; margin-left:20px;"><b>Single person:</b> A household with exactly one person of any age.</p>
      <p style="line-height:1.2; margin-left:20px;"><b>Single parent:</b> A household with one adult (aged 18+) and at least one child (&lt;18).</p>
      <p style="line-height:1.2; margin-left:20px;"><b>2 adults without children:</b> A household with exactly two adults and no children.</p>
      <p style="line-height:1.2; margin-left:20px;"><b>2 adults with children:</b> Households with exactly two adults, categorized by the number of children (1-2, 3-4, or 5+).</p>
      <p style="line-height:1.2; margin-left:20px;"><b>3 or more adults without children:</b> A household with three or more adults and no children.</p>
      <p style="line-height:1.2; margin-left:20px;"><b>3 or more adults with children:</b> A household with three or more adults and at least one child.</p>
      <p style="line-height:1.5"><b><span style='color:#0070C0'>Vulnerable households:</span></b> These categories identify households containing at least one member with a specific characteristic:</p>
	  <p style="line-height:1.2; margin-left:20px;"><b>Young child (aged 0-2):</b> Households with at least one child aged two or younger.</p>
      <p style="line-height:1.2; margin-left:20px;"><b>Elderly member:</b> Households with at least one member aged 65 or older.</p>
      <p style="line-height:1.2; margin-left:20px;"><b>Member with a disability:</b> Households with at least one member reported as having a disability (based on variable <i>ddi</i>).</p>
      <p style="line-height:1.2; margin-left:20px;"><b>No male adults:</b> Households without any male members aged 18 or older.</p>
      <p style="line-height:1.5"><b><span style='color:#0070C0'>Labour market status:</span></b> These categories classify households based on their connection to the labour market: </p>
      <p style="line-height:1.2; margin-left:20px;"><b>No labour market income:</b> No household member has positive labour market income (based on income list <i>ils_earns</i>).</p>
      <p style="line-height:1.2; margin-left:20px;"><b>Informal adult:</b> The household includes at least one adult member identified as an informal worker (based on variable <i>lfo</i>).</p>
      <p style="line-height:1.2; margin-left:20px;"><b>No informal adults:</b> The household does not include any adults identified as informal workers.</p>
      <p style="line-height:1.5"><b><span style='color:#0070C0'>Absolute national poverty line, yearly:</span></b> The annual poverty line used for the calculations, shown in national currency.</p>'''
    },
    'poverty-graphs': {
        'title': 'About: Poverty graphs',
        'body': '''<p style="line-height:1.5">This tab visualises how the reform scenario changes poverty outcomes relative to the baseline. Each bar shows the absolute difference, in percentage points (pp.), between the reform and the baseline scenario for the groups defined on the "Poverty" tab.</p>
      <p style="line-height:1.5"><b><span style='color:#0070C0'>Household structure:</span></b> Uses the same household classifications as the poverty tables (single person, single parent, etc.); see the "Poverty" tab description for full definitions.</p>
      <p style="line-height:1.5"><b><span style='color:#0070C0'>Vulnerability and labour market status:</span></b> Uses the same indicators for vulnerable members (young child, elderly, disability, no male adults) and labour-market engagement (no labour income, informal work) as the poverty tables.</p>
      <p style="line-height:1.5"><b><span style='color:#0070C0'>Interpretation:</span></b> Bars above zero imply that the reform increases the poverty rate or gap for the group relative to the baseline, whereas bars below zero indicate a reduction.</p>'''
    },
    'inequality-graphs': {
        'title': 'About: Inequality graphs',
        'body': '''<p style="line-height:1.5">This tab shows graphs of the absolute difference in the distribution of household resources and key fiscal components between the baseline and each reform scenario. For shares, changes are expressed in percentage points (pp.). For levels, they are in currency units.</p>

      <p style="line-height:1.5">Percentiles and deciles are formed by ranking individuals by their household's equivalised disposable income or consumption. The user's selection of "Distribution statistics" (consumption-based or income-based, and whether net of indirect taxes/benefits) determines the underlying ranking variable used for forming percentiles/deciles and calculating levels. All deciles are based on the distribution in the baseline scenario.</p>

      <p style="line-height:1.5">Only figures are displayed. Corresponding tables (also including the baseline and reform estimates in addition to the differences) can be found on the "Inequality" tab.</p>

      <p style="line-height:1.5"><b><span style='color:#0070C0'>Change in household income/consumption level at percentiles:</span></b> This graph shows the change in yearly equivalised disposable income or consumption at various points (percentiles) of the distribution.</p>
      <p style="line-height:1.5"><b><span style='color:#0070C0'>Change in share of total income/consumption by baseline decile (pp.):</span></b> This graph shows the change in the share of the total equivalised income or consumption possibilities in the economy held by each decile of the population.</p>
      <p style="line-height:1.5"><b><span style='color:#0070C0'>Change in share of total cash benefits by baseline decile (pp.):</span></b> This graph shows the change in the share of total cash benefits (<i>ils_ben</i>) received by each income/consumption decile.</p>
      <p style="line-height:1.5"><b><span style='color:#0070C0'>Change in share of total direct taxes by baseline decile (pp.):</span></b> This graph shows the change in the share of total direct tax payments (<i>ils_tax</i>) contributed by each income/consumption decile.</p>
      <p style="line-height:1.5"><b><span style='color:#0070C0'>Change in share of total indirect taxes by baseline decile (pp.):</span></b> This graph shows the change in the share of total indirect tax payments (<i>ils_taxco</i>) attributed to each income/consumption decile.</p>'''
    },
    'households': {
        'title': 'About: Households',
        'body': '''<p style="line-height:1.5">This tab shows counts of households, grouped by different characteristics. In addition to the baseline counts, each reform scenario includes the corresponding counts and their absolute differences from the baseline values.</p>
          <p style="line-height:1.5"><b><span style='color:#0070C0'>Taxpayer and benefit recipient households:</span></b> This table shows the weighted number of households where at least one member pays a specific tax/contribution or receives a specific benefit. Please note that in-kind benefits and indirect subsidies are not modelled in all countries.</p>
	      <p style="line-height:1.5"><b><span style='color:#0070C0'>Household categories:</span></b> This table shows the total number of households for various demographic and economic subgroups. These categories classify households by their structure of adults (18+), children (&lt;18), and other characteristics; see the "Poverty" tab description for full details. Please note that the sum of the categories under "Household structure" may not equal the total if the dataset contains multi-person households composed exclusively of children (persons under 18), as this rare household type is not separately categorized.</p>
          <p style="line-height:1.5"><b><span style='color:#0070C0'>Household decile distribution:</span></b> This table shows the number of households in each decile. These deciles are calculated at the household level, meaning the population of households (not individuals) has been ranked and divided into ten equal-sized groups based on the selected income/consumption concept. The user's selection of "Distribution statistics" (consumption-based or income-based, and whether net of indirect taxes/benefits) determines the underlying ranking variable used for forming deciles.</p>'''
    },
    'individuals': {
        'title': 'About: Individuals',
        'body': '''<p style="line-height:1.5">This tab shows counts of individuals, grouped by different characteristics. In addition to the baseline counts, each reform scenario includes the corresponding counts and their absolute differences from the baseline values.</p>
          <p style="line-height:1.5"><b><span style='color:#0070C0'>Taxpayer and benefit recipient individuals:</span></b> This table shows the weighted number of individuals who pay a specific tax/contribution or receive a specific benefit. For taxes/benefits that are recorded at the household level, only the household head will be counted as an individual payer/recipient for that specific item. Please note that in-kind benefits and indirect subsidies are not modelled in all countries.</p>
	      <p style="line-height:1.5"><b><span style='color:#0070C0'>Household categories:</span></b> This table shows the total number of individuals living in households with the specified demographic and economic characteristics; see the "Poverty" tab description for full details. Please note that the sum of the categories under "Household structure" may not equal the total if the dataset contains multi-person households composed exclusively of children (persons under 18), as this rare household type is not separately categorized.</p>
          <p style="line-height:1.5"><b><span style='color:#0070C0'>Individual decile distribution:</span></b> This table shows the number of individuals in each decile. Deciles are calculated at the individual level by ranking all persons based on their household's equivalised disposable income or consumption. The user's selection of "Distribution statistics" (consumption-based or income-based, and whether net of indirect taxes/benefits) determines the underlying ranking variable used for forming deciles.</p>'''
    },
    'inequality': {
        'title': 'About: Inequality',
        'body': '''<p style="line-height:1.5">This tab shows measures of inequality and the distribution of household resources. All distributional statistics (e.g., percentiles) are calculated by ranking all individuals according to their household's equivalised disposable income or consumption possibilities (based on the user's selection). In addition to the baseline results, each reform scenario includes the corresponding outcomes and their absolute differences from the baseline values.</p>
	  
          <p style="line-height:1.5"><b><span style='color:#0070C0'>Inequality indices:</span></b> The Gini and Atkinson indices measure inequality on a scale from 0 (perfect equality) to 100 (perfect inequality). The P80/P20 and mean/median ratios compare different points of the distribution.</p>
          
		  <p style="line-height:1.5"><b><span style='color:#0070C0'>Percentiles of distribution and median, yearly:</span></b> This table presents annual amounts in national currency (unless this setting is changed by the user). It shows the level of equivalised disposable income or consumption at various points (percentiles) of the distribution. For example, the 10th percentile shows the level below which the poorest 10% of individuals fall. The 50th percentile represents the median. The user's selection of "Distribution statistics" (consumption-based or income-based, and whether net of indirect taxes/benefits) determines the underlying ranking variable used for forming percentiles and calculating levels.</p>
          
		  <p style="line-height:1.5"><b><span style='color:#0070C0'>Absolute national poverty line, yearly:</span></b> The annual poverty line is also shown in national currency for comparison with the percentile distribution.</p>
          
          <p style="line-height:1.5"><b><span style='color:#0070C0'>Distribution of total income/consumption across deciles, %:</span></b> This shows the percentage of total income/consumption held by each decile (10% group) of the population. For reforms, the distribution is calculated over the deciles as defined in the baseline scenario to ensure comparability. The user's selection of "Distribution statistics" determines the underlying ranking variable used for forming deciles and calculating levels.</p>'''
    },
    'benefits': {
        'title': 'About: Benefits',
        'body': '''<p style="line-height:1.5">This tab shows the distribution of cash benefits (<i>ils_ben</i>) and in-kind benefits (<i>ils_benki</i>). Note that indirect subsidies are not included in this tab's results, and in-kind benefits are not modelled in all countries. In addition to the baseline results, each reform scenario includes the corresponding outcomes and their absolute differences from the baseline values.</p>
          <p style="line-height:1.5"><b><span style='color:#0070C0'>Receipt of benefits by household type, % of households:</span></b> These tables show the percentage of households in a specific category that receive any benefit, cash benefits, or in-kind benefits. For reforms, household categories are fixed to their baseline characteristics.</p>
          <p style="line-height:1.5"><b><span style='color:#0070C0'>Targeting of poor households, % of benefits:</span></b> These rows show what percentage of the total benefit expenditure (of a specific type) is received by households who were defined as poor in the baseline scenario.</p>
          <p style="line-height:1.5"><b><span style='color:#0070C0'>Per-capita adequacy:</span></b> These rows show the mean yearly benefit amount per beneficiary. This amount is calculated at the individual level after equivalising the household benefit amount, making it comparable to the individual-level poverty line. It is also shown as a share of the baseline yearly median individual consumption and disposable income (before indirect taxes and benefits) in the population.</p>
          <p style="line-height:1.5"><b><span style='color:#0070C0'>Distribution across deciles, %:</span></b> These tables show the percentage of total cash or in-kind benefits received by each decile of the population. For reforms, the distribution is calculated over the deciles as defined in the baseline scenario to ensure comparability. The user's selection of "Distribution statistics" (consumption-based or income-based, and whether net of indirect taxes/benefits) determines the underlying ranking variable used for forming deciles.</p>'''
    },
    'taxes': {
        'title': 'About: Taxes',
        'body': '''<p style="line-height:1.5">This tab shows the distribution of direct taxes, indirect taxes, and social contributions across different household groups and income/consumption deciles. In addition to the baseline results, each reform scenario includes the corresponding outcomes and their absolute differences from the baseline values.</p>
          <p style="line-height:1.5"><b><span style='color:#0070C0'>Payment by household type, % of households:</span></b> These tables show what percentage of households with specific characteristics pay any amount of direct tax, indirect tax, or social contributions. "Social contributions" here refers to the sum of employee and self-employed contributions (from income lists <i>ils_sicee</i> and <i>ils_sicse</i>). For reforms, household categories are fixed to their baseline characteristics.</p>
        <p style="line-height:1.5"><b><span style='color:#0070C0'>Effective tax rates, %:</span></b> This table shows various average effective tax rates, calculated as the total amount of taxes (and social insurance contributions, SIC) paid, divided by total original income across the entire population.</p>
          <p style="line-height:1.5"><b><span style='color:#0070C0'>Distribution across deciles, %:</span></b> These tables show the percentage of the total tax or contribution burden that is paid by each decile of the population. For reforms, the distribution is calculated over the deciles as defined in the baseline scenario to ensure comparability. The user's selection of "Distribution statistics" (consumption-based or income-based, and whether net of indirect taxes/benefits) determines the underlying ranking variable used for forming deciles.</p>'''
    },
    'policy-effects': {
        'title': 'About: Policy effects',
        'body': '''<p style="line-height:1.5">This tab illustrates the redistributive impact of the tax-benefit system by comparing the poverty rate and the Gini coefficient "before" and "after" taxes and benefits. In addition to the baseline results, each reform scenario includes the corresponding outcomes and their absolute differences from the baseline values.</p>

          <p style="line-height:1.5"><b><span style='color:#0070C0'>"Before taxes and benefits" measures:</span></b></p>
          <p style="line-height:1.2; margin-left:20px;"><b>Income based distribution statistics:</b> If the user selected "Income based" or "Income based, net of indirect taxes and benefits", the measures are computed based on original market income (income list <i>ils_origy</i>) and imputed home produce (variable <i>xivot</i>, if available).</p>
          <p style="line-height:1.2; margin-left:20px;"><b>Consumption based distribution statistics:</b> If the user selected "Consumption based" or "Consumption based, net of indirect taxes and benefits", a proxy for resources available before direct taxes and cash benefits is used for computing the measures. It is calculated as <i>ils_con - ils_ben + ils_tax + ils_sicee + ils_sicse</i>.</p>
          <p style="line-height:1.2; margin-left:20px;"><b>Poverty line:</b> For the "before" poverty measure, the standard poverty line (variable <i>spl</i>) is used, ensuring consistency with the "before" income/consumption concepts, which are also prior to indirect fiscal effects.</p>
          <p style="line-height:1.2; margin-left:20px;"><b>Reform values:</b> For all "Before taxes and benefits" indicators shown in the reform scenario columns, the values correspond to the baseline system to keep the "Effects" comparable.</p>

          <p style="line-height:1.5"><b><span style='color:#0070C0'>"After taxes and benefits" measures:</span></b></p>
          <p style="line-height:1.2; margin-left:20px;">Indicators are calculated in the same way as on the "Poverty" and "Inequality" tabs, using the user's selected income/consumption concept.</p>

          <p style="line-height:1.5"><b><span style='color:#0070C0'>Effects of tax-benefit system:</span></b></p>
          <p style="line-height:1.2; margin-left:20px;">The absolute difference in percentage points (pp.), After - Before.</p>'''
    },
    'gainers-losers': {
        'title': 'About: Gainers and losers',
        'body': '''<p style="line-height:1.5">This tab shows the share of individuals (as a % of the population in the respective group) who are gainers or losers due to the reform. The change is measured for individuals based on their household's equivalised disposable income or consumption relative to the baseline scenario. All groupings are based on the individual's situation in the baseline.</p>

          <p style="line-height:1.5">Only figures are displayed directly.</p>

          <p style="line-height:1.5"><b><span style='color:#0070C0'>Gainers:</span></b> Individuals whose household resources increase by more than 1% or 5% due to the reform.</p>
          <p style="line-height:1.5"><b><span style='color:#0070C0'>Losers:</span></b> Individuals whose household resources decrease by more than 1% or 5% due to the reform.</p>
          <p style="line-height:1.5"><b><span style='color:#0070C0'>Deciles:</span></b> Calculated at the individual level based on equivalised disposable income or consumption. In income-based runs, interpret carefully if many individuals have zero income, as they will cluster in the bottom decile.</p>
          <p style="line-height:1.5"><b><span style='color:#0070C0'>Household structure, vulnerable households, labour market status:</span></b> Groupings follow the definitions outlined in the "Poverty" tab description.</p>'''
    },
    'default': {
        'title': 'Info',
        'body': 'No information available for this tab.'
    }
}


def normalize_html_text(raw_html: str) -> str:
    if not isinstance(raw_html, str):
        return ""
    cleaned_lines = [line.lstrip() for line in raw_html.splitlines()]
    cleaned = textwrap.dedent("\n".join(cleaned_lines)).strip()
    return cleaned


def html_to_plain_text(raw_html: str) -> str:
    normalized = normalize_html_text(raw_html)
    if not normalized:
        return ""
    text = re.sub(r'<\s*br\s*/?>', '\n', normalized, flags=re.IGNORECASE)
    text = re.sub(r'</p\s*>', '\n\n', text, flags=re.IGNORECASE)
    text = re.sub(r'<p[^>]*>', '', text, flags=re.IGNORECASE)
    text = re.sub(r'</?span[^>]*>', '', text, flags=re.IGNORECASE)
    text = re.sub(r'</?b[^>]*>', '', text, flags=re.IGNORECASE)
    text = re.sub(r'</?i[^>]*>', '', text, flags=re.IGNORECASE)
    text = re.sub(r'</?u[^>]*>', '', text, flags=re.IGNORECASE)
    text = re.sub(r'<a[^>]*href="([^"]*)"[^>]*>(.*?)</a>', r'\2 (\1)', text, flags=re.IGNORECASE)
    text = re.sub(r'<[^>]+>', '', text)
    text = html_utils.unescape(text)
    text = re.sub(r'\n{3,}', '\n\n', text)
    return text.strip()


TABLE_HEADER_LABELS = {
    "By source",
    "By type",
    "Household structure",
    "Vulnerable households",
    "Labour market status",
    "Inequality indices",
    "Percentiles of distribution and median, yearly",
    "Distribution of total income/consumption across baseline deciles, %",
    "Receipt of benefits by household type, % of households",
    "Receipt of cash benefits by household type, % of households",
    "Receipt of in-kind benefits by household type, % of households",
    "Targeting of poor households (baseline poverty status), % of benefits",
    "Per-capita adequacy",
    "Distribution of cash benefits across baseline deciles, %",
    "Distribution of in-kind benefits across baseline deciles, %",
    "Poverty rate",
    "Poverty gap",
    "Gini coefficient",
    "P80/P20 ratio",
    "Cash benefits",
    "In-kind benefits",
}
TABLE_STRONG_LABELS = {
    "Sum of government revenue",
    "Sum of government expenditure",
    "All individuals",
    "Total households",
    "Total individuals",
    "Any taxes or contributions",
    "Any benefits",
    "Absolute national poverty line, yearly",
    "Total",
    "- All households",
}
TABLE_DIVIDER_LABELS = {
    "Sum of government expenditure",
    "By type",
    "By source",
    "Any taxes or contributions",
    "Any benefits",
    "Household structure",
    "Vulnerable households",
    "Labour market status",
    "Percentiles of distribution and median, yearly",
    "Distribution of total income/consumption across baseline deciles, %",
    "Receipt of benefits by household type, % of households",
    "Receipt of cash benefits by household type, % of households",
    "Receipt of in-kind benefits by household type, % of households",
    "Targeting of poor households (baseline poverty status), % of benefits",
    "Per-capita adequacy",
    "Distribution of cash benefits across baseline deciles, %",
    "Distribution of in-kind benefits across baseline deciles, %",
    "Cash benefits",
    "In-kind benefits",
}


def strip_tags(html_text: str) -> str:
    text = re.sub(r'<[^>]+>', '', html_text)
    return html_utils.unescape(text)


def extract_description_lines(info_key: str):
    content = INFO_MODAL_CONTENT.get(info_key)
    if not content:
        return []
    body = normalize_html_text(content.get('body', ''))
    if not body:
        return []
    paragraphs = re.findall(r'<p[^>]*>(.*?)</p>', body, flags=re.IGNORECASE | re.DOTALL)
    lines = []
    for para in paragraphs:
        segment = para.strip()
        if not segment:
            continue
        match = re.match(r'\s*<b>\s*<span[^>]*>(.*?)</span>\s*</b>\s*(.*)', segment, flags=re.IGNORECASE | re.DOTALL)
        if match:
            header_html = match.group(1)
            remainder_html = match.group(2)
            header_text = strip_tags(header_html).strip()
            body_text = strip_tags(remainder_html).strip()
            if header_text:
                lines.append(('header', header_text))
            if body_text:
                lines.append(('body', body_text))
        else:
            plain_text = strip_tags(segment).strip()
            if plain_text:
                lines.append(('body', plain_text))
    return lines


def ensure_input_dataframe():
    """
    Loads the input micro-data once and reuses it across requests to avoid
    repeated disk I/O on resource-constrained hosts (e.g. Render free tier).
    """
    global INPUT_DF
    if INPUT_DF is None:
        df = pd.read_csv(INPUT_FILE, sep=r'\s+', low_memory=False)
        int_cols = df.select_dtypes(include=['int64', 'int32', 'int16', 'int']).columns
        float_cols = df.select_dtypes(include=['float64']).columns

        if len(int_cols) > 0:
            df[int_cols] = df[int_cols].apply(pd.to_numeric, downcast='integer')
        if len(float_cols) > 0:
            df[float_cols] = df[float_cols].apply(pd.to_numeric, downcast='float')

        obj_cols = df.select_dtypes(include=['object']).columns
        for col in obj_cols:
            unique_values = df[col].nunique(dropna=False)
            if 0 < unique_values < 256 and unique_values / len(df) < 0.5:
                df[col] = df[col].astype('category')

        INPUT_DF = df
    return INPUT_DF


def get_baseline_artifacts(df: pd.DataFrame, analysis_choice: int):
    """
    Returns cached baseline simulation results for the selected distribution statistic.
    This prevents recomputing the heavy baseline run on every callback.
    """
    cache_key = analysis_choice
    if cache_key in BASELINE_CACHE:
        return BASELINE_CACHE[cache_key]

    baseline_sim_df = run_simulation(df, BASELINE_PARAMS)
    baseline_results, baseline_analysis_df = run_analysis(baseline_sim_df, analysis_choice)

    cols_to_keep = ['idperson', 'idhh', 'dwt', 'deciles', 'deciles_hh', 'deciles_base', 'eqRank_baseline', 'AllIndividuals_Base_Pass']
    # Raw resource columns needed to recompute the baseline adequacy denominators in reform runs
    cols_to_keep.extend([col for col in ['ils_con', 'ils_dispyx'] if col in baseline_analysis_df.columns])
    cols_to_keep.extend([col for col in baseline_analysis_df.columns if col.startswith('is') and 'HH' in col])
    baseline_merge_df = baseline_analysis_df.loc[:, cols_to_keep].copy()

    artifacts = {
        'results': baseline_results,
        'merge_df': baseline_merge_df,
    }
    BASELINE_CACHE[cache_key] = artifacts

    del baseline_sim_df, baseline_analysis_df
    gc.collect()

    return artifacts


# --- SIMULATION ENGINE ---
def run_simulation(df, params):
    sim_df = df.copy()
    
    # Ensure required identification columns are present
    required = ['idhh', 'dhh', 'idperson']
    missing = [c for c in required if c not in sim_df.columns]
    if missing:
        raise ValueError(f"Missing required id columns: {', '.join(missing)}")

    # 'ses' is no longer in this list, as we will simulate it.
    # Initialize missing columns together to keep the frame compact
    cols_to_add = {}
    for col in ['lfo', 'ddi', 'xivot', 'dec', 'ytn']:
        if col not in sim_df.columns:
            cols_to_add[col] = 0
    if cols_to_add:
        sim_df = sim_df.assign(**cols_to_add)

    # Ensure key monetary variables exist and contain no missing values
    for var in ['yem', 'yse', 'yag', 'yds', 'xhh']:
        if var not in sim_df.columns:
            sim_df[var] = 0
        else:
            sim_df[var] = sim_df[var].fillna(0)

    uprating_factors = {
        'f_CPI_Overall': 1.2092, 'f_CPI_Food': 1.1797, 'f_CPI_Non_Food': 1.2746,
        'f_CPI_Alcohol': 1.2089, 'f_CPI_Tobacco': 1.1842, 'f_CPI_Energy': 1.1275,
        'f_CPI_Earnings': 1.3197
    }

    for var in ['yem', 'yse', 'yag']:
        if var in sim_df.columns:
            sim_df[var] *= uprating_factors['f_CPI_Earnings']

    # Uprate 'yds'
    for var in ['yds','ytn']:
        if var in sim_df.columns:
            sim_df[var] *= uprating_factors['f_CPI_Overall']
            
    # Clip individual labour incomes at zero
    for col in ['yem', 'yse']:
        if col in sim_df.columns:
            sim_df[col] = sim_df[col].clip(lower=0)
    
    if 'yag' in sim_df.columns:
        yag_hh_sum = sim_df.groupby('idhh')['yag'].transform('sum')
        # If household-level YAG <= 0, set all members' yag to 0
        sim_df.loc[yag_hh_sum <= 0, 'yag'] = 0
        # Otherwise, drop only negative individual values
        sim_df.loc[sim_df['yag'] < 0, 'yag'] = 0

    # Apply category-specific uprating factors to expenditure variables
    food_vars = [
        'x0111', 'x0112', 'x0113', 'x0114', 'x0115', 'x0116', 'x0117', 'x0118', 'x0119',
        'x0121', 'x0122', 'x1111', 'x1112'
    ]
    alcohol_vars = ['x0211', 'x0212', 'x0213']
    energy_vars = ['x0451', 'x0452', 'x0453', 'x0454', 'x0455']
    non_food_vars = [
        'x0311', 'x0312', 'x0313', 'x0314', 'x0321', 'x0322', 'x0411', 'x0412',
        'x0431', 'x0432', 'x0441', 'x0442', 'x0443', 'x0444', 'x0511', 'x0512',
        'x0513', 'x0531', 'x0532', 'x0533', 'x0551', 'x0552', 'x0561', 'x0562',
        'x0611', 'x0612', 'x0613', 'x0621', 'x0622', 'x0623', 'x0711', 'x0712',
        'x0713', 'x0714', 'x0721', 'x0722', 'x0723', 'x0724', 'x0731', 'x0732',
        'x0733', 'x0734', 'x0735', 'x0810', 'x0911', 'x0912', 'x0921', 'x0922',
        'x0923', 'x0931', 'x0932', 'x0941', 'x0951', 'x0960', 'x1211', 'x1212',
        'x1213'
    ]
    # List of non-monetary 'x' vars to skip in loop
    x_skip_vars = ['xivot', 'xhh'] 

    for col in sim_df.columns:
        if col.startswith('x') and col not in x_skip_vars:
            factor = uprating_factors['f_CPI_Overall'] # Default
            if col in food_vars: 
                factor = uprating_factors['f_CPI_Food']
            elif col in alcohol_vars: 
                factor = uprating_factors['f_CPI_Alcohol']
            elif col in energy_vars: 
                factor = uprating_factors['f_CPI_Energy']
            elif col in non_food_vars:
                factor = uprating_factors['f_CPI_Non_Food']
            # Note: x0230 (Narcotics) is correctly mapped to f_CPI_Overall by default
            sim_df[col] *= factor

    # Uprate xhh (if it exists) by the overall factor
    if 'xhh' in df.columns:
        sim_df['xhh'] *= uprating_factors['f_CPI_Overall']


    # Uprate poverty lines (monthly values) and align names
    sim_df['spl'] = params['basic_pov_line'] * uprating_factors['f_CPI_Overall']
    sim_df['splpf'] = params['basic_pov_line_pf'] * uprating_factors['f_CPI_Overall']
    sim_df['spl_u'] = params['upper_pov_line'] * uprating_factors['f_CPI_Overall']
    sim_df['splpf_u'] = params['upper_pov_line_pf'] * uprating_factors['f_CPI_Overall']

    # --- Equivalence Scale (ses) Calculation ---
    # Calculate individual 'ses' based on age bins
    bins = [-np.inf, 3, 7, 12, 17, 29, 39, 59, np.inf]
    labels = [0.30, 0.50, 0.70, 0.95, 1.10, 0.95, 0.90, 0.80]
    
    # Ensure 'dag' column exists
    if 'dag' not in sim_df.columns:
        sim_df['dag'] = 0 # Default age to 0 if missing
        
    sim_df['ses_person'] = pd.cut(sim_df['dag'], bins=bins, labels=labels, right=True, ordered=False).astype(float)
    
    # Sum individual 'ses' values to get the household total
    hh_ses_total = sim_df.groupby('idhh')['ses_person'].transform('sum')
    
    # Assign this total 'ses' value *only to the household head*
    sim_df['ses'] = 0.0
    is_head = sim_df['dhh'] == 1
    sim_df.loc[is_head, 'ses'] = hh_ses_total[is_head]
    sim_df = sim_df.drop(columns=['ses_person'])
    # --- End of Equivalence Scale Calculation ---

    # Continue with simulation logic
    
    # Store uprated sums of COICOP items as a fallback consumption measure
    x_cols = sim_df.filter(regex='^x[0-9]').columns
    # Calculate sum of x-cols at individual level first
    xhh_uprated_fallback_indiv = sim_df.loc[:, x_cols].sum(axis=1)
    # Then, create the household-level sum, broadcast to all members
    xhh_uprated_fallback_hh = xhh_uprated_fallback_indiv.groupby(sim_df['idhh']).transform('sum')

    # 'yds' is now the uprated version from the data.
    # This value is used to calculate the change in disposable income.
    yds_hh_uprated = sim_df.groupby('idhh')['yds'].transform('sum')

    # Prefer the dataset xhh column when available, otherwise use the COICOP fallback
    if 'xhh' in df.columns and not sim_df['xhh'].eq(0).all():
        # Use uprated 'xhh' from data (assumed to be HH total repeated for members)
        xhh_base = sim_df['xhh'] 
    else:
        # Use uprated HH-level sum-of-cols fallback
        xhh_base = xhh_uprated_fallback_hh

    # Initialize policy columns in a single assign call
    policy_cols = ['tscee_s', 'tscer_s', 'ttn_s', 'tin_s', 'tva_s', 'bsa_s', 'boa_s', 'bed_s']
    sim_df = sim_df.assign(**{col: 0.0 for col in policy_cols})

    formal_workers = sim_df['lfo'] == 1
    sim_df.loc[formal_workers, 'tscee_s'] = sim_df.loc[formal_workers, 'yem'] * params['tscee_rate']
    sim_df.loc[formal_workers, 'tscer_s'] = sim_df.loc[formal_workers, 'yem'] * params['tscer_rate']

    # Presumptive tax logic (using yearly params from UI, dividing by 12)
    cond1 = (sim_df['ytn'] > params['presumptive_turnover_1'] / 12) & (sim_df['ytn'] <= params['presumptive_turnover_2'] / 12)
    cond2 = (sim_df['ytn'] > params['presumptive_turnover_2'] / 12) & (sim_df['ytn'] <= params['presumptive_turnover_3'] / 12)
    cond3 = (sim_df['ytn'] > params['presumptive_turnover_3'] / 12) & (sim_df['ytn'] <= params['pit_yse_turnover_threshold'] / 12)
    sim_df.loc[cond1, 'ttn_s'] = params['presumptive_tax_2'] / 12
    sim_df.loc[cond2, 'ttn_s'] = params['presumptive_tax_3'] / 12
    sim_df.loc[cond3, 'ttn_s'] = sim_df['ytn'] * params['presumptive_rate_4']
    
    # PIT logic (calculated on an annual basis, then converted to monthly)
    
    # 1. Annualize monthly incomes
    yem_y = sim_df['yem'] * 12
    yse_y = sim_df['yse'] * 12
    yag_y = sim_df['yag'] * 12
    ytn_y = sim_df['ytn'] * 12
    tscee_y = sim_df['tscee_s'] * 12
    
    # 2. Calculate annual tax base (ttb_y)
    ttb01_y = np.where(ytn_y > params['pit_yse_turnover_threshold'], yse_y, 0) + np.maximum(0, yag_y - params['pit_yag_exemption'])
    ttb02_y = np.where(sim_df['lfo'] == 1, yem_y, 0)
    ttb_y = (ttb01_y + ttb02_y - tscee_y).clip(lower=0)

    # Store monthly versions for output
    sim_df['ttb01_s'] = ttb01_y / 12
    sim_df['ttb02_s'] = ttb02_y / 12
    sim_df['ttb_s'] = ttb_y / 12

    # 3. Get annual thresholds and rates from params
    b2_y, b3_y, b4_y, b5_y = params['pit_bracket2_thresh'], params['pit_bracket3_thresh'], params['pit_bracket4_thresh'], params['pit_bracket5_thresh']
    r2, r3, r4, r5 = params['pit_bracket2_rate'], params['pit_bracket3_rate'], params['pit_bracket4_rate'], params['pit_bracket5_rate']
    
    # 4. Calculate annual tax (tax_y)
    tax_y = pd.Series(0.0, index=sim_df.index)
    tax_y += (ttb_y.clip(upper=b3_y) - b2_y).clip(lower=0) * r2
    tax_y += (ttb_y.clip(upper=b4_y) - b3_y).clip(lower=0) * r3
    tax_y += (ttb_y.clip(upper=b5_y) - b4_y).clip(lower=0) * r4
    tax_y += (ttb_y - b5_y).clip(lower=0) * r5
    
    # 5. Convert annual tax back to monthly flow (tin_s)
    sim_df['tin_s'] = tax_y / 12

    # VAT logic (calculated for head only)
    # Get the list of vatable items from the params
    vatable_item_list = params.get('vat_items_list', []) 
    # Filter this list to only include columns that actually exist in the dataframe
    vatable_cols = [col for col in vatable_item_list if col in sim_df.columns]
    
    vat_base = sim_df.loc[is_head, vatable_cols].sum(axis=1)

    sim_df['il_exp_vat'] = 0.0
    sim_df.loc[is_head, 'il_exp_vat'] = vat_base

    # Apply the VAT rate directly to the consumption base
    sim_df.loc[is_head, 'tva_s'] = vat_base * params['tva_rate']

    # Benefit logic (using monthly params from UI)
    sim_df['ils_origy'] = sim_df['yem'] + sim_df['yse'] + sim_df['yag']
    hh_origy = sim_df.groupby('idhh')['ils_origy'].transform('sum')
    hh_size = sim_df.groupby('idhh')['idperson'].transform('count')
    hh_disabled_count = sim_df.groupby('idhh')['ddi'].transform('sum')
    eligible_hh_mask = hh_origy < params['bsa_income_threshold']
    amount = pd.Series(0.0, index=sim_df.index)
    amount.loc[hh_size == 1] = params['bsa_1_person']
    amount.loc[hh_size == 2] = params['bsa_2_person']
    amount.loc[hh_size >= 3] = params['bsa_3_plus_person']
    amount += hh_disabled_count * params['bsa_disabled_topup']
    sim_df.loc[is_head & eligible_hh_mask, 'bsa_s'] = amount[is_head & eligible_hh_mask]

    eligible_seniors = (sim_df['dag'] >= params['senior_grant_age']) & (sim_df['ils_origy'] < params['senior_grant_income_threshold'])
    sim_df.loc[eligible_seniors, 'boa_s'] = params['senior_grant_amount']

    eligible_children = (sim_df['dag'] < params['school_meal_age']) & (sim_df['dec'].isin([2, 3, 4]))
    sim_df.loc[eligible_children, 'bed_s'] = params['school_meal_value'] * (10 / 12) # Averaging 10-month benefit over 12 months

    # Final resource definitions
    ils_tax_indiv = sim_df['tin_s'] + sim_df['ttn_s']
    ils_ben_indiv = sim_df['bsa_s'] + sim_df['boa_s'] # Note: bsa_s is only on head
    ils_dispy_indiv = sim_df['ils_origy'] + ils_ben_indiv - ils_tax_indiv - sim_df['tscee_s']
    ils_benki_indiv = sim_df['bed_s']
    ils_dispyki_indiv = ils_dispy_indiv + ils_benki_indiv
    sim_dispyki_hh = ils_dispyki_indiv.groupby(sim_df['idhh']).transform('sum')
    
    # Calculate the change in disposable income
    delta = sim_dispyki_hh - yds_hh_uprated

    # ils_con = baseline ils_con + 1-to-1 change in ils_dispyki, with EUROMOD's
    # consumption floor: losses cannot push consumption below 25% of baseline
    # (xhhadj_dv, $protect_cons = 0.25)
    xhh_s = np.where(delta < 0, np.maximum(xhh_base + delta, xhh_base * 0.25), xhh_base + delta)

    # Store xhh_s and ils_con, assigning the HH total only to the head
    
    # First, create xhh_s and set to 0 for everyone
    sim_df['xhh_s'] = 0.0
    # Then, assign the calculated xhh_s value ONLY to the household head
    sim_df.loc[is_head, 'xhh_s'] = pd.Series(xhh_s, index=sim_df.index)[is_head].fillna(0) 

    # Do the same for ils_con (which uses the same xhh_s value)
    sim_df['ils_con'] = 0.0
    sim_df.loc[is_head, 'ils_con'] = pd.Series(xhh_s, index=sim_df.index)[is_head].fillna(0) # HH consumption assigned to head

    # Aggregating variables for analysis
    # Use assign to avoid fragmenting the dataframe
    new_analysis_cols = {
        'ils_head': (sim_df['dhh'] == 1).astype(int),
        'ils_earns': sim_df.get('yem', 0) + sim_df.get('yse', 0) + sim_df.get('yag', 0),
        'ils_dispyki': ils_dispyki_indiv,
        'ils_tax': ils_tax_indiv,
        'ils_sicee': sim_df['tscee_s'],
        'ils_sicse': 0.0,
        'ils_sicer': sim_df['tscer_s'],
        'ils_pen': 0.0,
        'ils_benmt': sim_df['bsa_s'] + sim_df['boa_s'],
        'ils_bennt': 0.0,
        'ils_benki': sim_df['bed_s'],
        'ils_benco': 0.0,
        'ils_bch': sim_df['bed_s'],
        'ils_bsu': 0.0,
        'ils_bdi': 0.0,
        'ils_bun': 0.0,
        'ils_bag': 0.0
    }
    sim_df = sim_df.assign(**new_analysis_cols)
    
    sim_df['ils_sic'] = sim_df['ils_sicee'] + sim_df['ils_sicse'] + sim_df['ils_sicer']
    sim_df['ils_ben'] = sim_df['ils_benmt'] + sim_df['ils_bennt'] + sim_df['ils_pen']
    sim_df['ils_bsa'] = sim_df['bsa_s'] + sim_df['boa_s']
    sim_df['ils_dispy'] = ils_dispy_indiv
    
    # Keep VAT totals stored on the household head only
    tva_s_hh = sim_df.groupby('idhh')['tva_s'].transform('sum')
    sim_df['ils_taxco'] = 0.0 # Clear for everyone
    sim_df.loc[is_head, 'ils_taxco'] = tva_s_hh[is_head] # Assign HH total to head
    
    # These lines are now correct because ils_taxco is 0 for non-heads
    sim_df['ils_dispy_pf'] = sim_df['ils_dispy'] - sim_df['ils_taxco'] + sim_df['ils_benco']
    sim_df['ils_con_pf'] = sim_df['ils_con'] - sim_df['ils_taxco'] + sim_df['ils_benco']
    sim_df['ils_dispyx'] = sim_df['ils_dispy'] + sim_df['xivot']
    sim_df['ils_dispyx_pf'] = sim_df['ils_dispyx'] - sim_df['ils_taxco'] + sim_df['ils_benco']

    # EUROMOD's Statistics Presenter reads the output text file, where monetary
    # variables are stored rounded to 2 decimals; round the same way so the
    # statistics computed downstream match EUROMOD's exactly.
    monetary_cols = [
        'xhh_s', 'xhh', 'yem', 'yse', 'yag', 'ytn', 'yds', 'bsa_s', 'boa_s',
        'bed_s', 'tscee_s', 'tscer_s', 'ttn_s', 'ttb01_s', 'ttb02_s', 'ttb_s',
        'tin_s', 'tva_s', 'spl', 'splpf', 'ses', 'ils_earns', 'ils_origy',
        'ils_tax', 'ils_sicee', 'ils_sicse', 'ils_sicer', 'ils_pen',
        'ils_benmt', 'ils_bennt', 'ils_benki', 'ils_ben', 'ils_dispy',
        'ils_dispy_pf', 'ils_dispyki', 'ils_taxco', 'ils_sic', 'ils_benco',
        'ils_bch', 'ils_bsa', 'ils_dispyx', 'ils_dispyx_pf', 'ils_con',
        'ils_con_pf', 'il_exp_vat',
    ]
    for col in monetary_cols:
        if col in sim_df.columns:
            sim_df[col] = pd.to_numeric(sim_df[col], errors='coerce').round(2)

    # Return a compact copy
    return sim_df.copy()

# --- ANALYSIS HELPER FUNCTIONS ---
def weighted_sum(df, column_name, weight_col='dwt'):
    if column_name not in df.columns or weight_col not in df.columns or df.empty: return 0
    return (df[column_name] * df[weight_col]).sum()

def weighted_average(df, column_name, weight_col='dwt'):
    if df.empty: return 0
    if column_name not in df.columns or weight_col not in df.columns: return 0
    weights = df[weight_col]
    if weights.sum() == 0: return 0
    return np.average(df[column_name], weights=weights)


def weighted_gini(values, weights):
    values = np.asarray(values, dtype=float)
    weights = np.asarray(weights, dtype=float)
    mask = (weights > 0) & np.isfinite(values)
    if not mask.any():
        return 0.0
    values = values[mask]
    weights = weights[mask]
    order = np.argsort(values)
    values = values[order]
    weights = weights[order]
    cum_weights = np.cumsum(weights)
    total_weight = cum_weights[-1]
    if total_weight <= 0:
        return 0.0
    weighted_values = values * weights
    cum_weighted_values = np.cumsum(weighted_values)
    total_value = cum_weighted_values[-1]
    if total_value == 0:
        return 0.0
    cum_weights = np.insert(cum_weights, 0, 0)
    cum_weighted_values = np.insert(cum_weighted_values, 0, 0)
    lorenz = cum_weighted_values / total_value
    cum_weights_norm = cum_weights / total_weight
    area_under_lorenz = np.sum((lorenz[1:] + lorenz[:-1]) * np.diff(cum_weights_norm) / 2)
    gini = 1 - 2 * area_under_lorenz
    return max(0.0, min(1.0, gini))


def _em_household_blocks(values, weights, hh_ids):
    """Household blocks as EUROMOD's statistics engine forms them: persons sorted
    ascending by value (stable), households in order of first appearance in that
    sort, block value = the last (max) member value, block weight = sum of member
    weights. Returns (hh_keys, block_values, block_weights) in block order."""
    values = np.asarray(values, dtype=float)
    weights = np.asarray(weights, dtype=float)
    hh_ids = np.asarray(hh_ids)
    order = np.argsort(values, kind='mergesort')
    block_val = {}
    block_wt = {}
    block_keys = []
    for idx in order:
        hh = hh_ids[idx]
        if hh not in block_wt:
            block_keys.append(hh)
            block_wt[hh] = 0.0
        block_wt[hh] += weights[idx]
        block_val[hh] = values[idx]
    return (block_keys,
            np.array([block_val[hh] for hh in block_keys], dtype=float),
            np.array([block_wt[hh] for hh in block_keys], dtype=float))


def em_gini_grouped(values, weights, hh_ids, recode_negatives=True):
    """EUROMOD CalculateGini with GroupingVar=idhh and RecodeNegatives
    (EM_Statistics/EM_TemplateCalculator_Actions.cs): an upper-rectangle
    estimator over household blocks. Returns the Gini on a 0-1 scale."""
    values = np.asarray(values, dtype=float)
    weights = np.asarray(weights, dtype=float)
    hh_ids = np.asarray(hh_ids)
    wt_inc = values * weights
    if recode_negatives:
        wt_inc = np.maximum(wt_inc, 0.0)
    total_inc = wt_inc.sum()
    total_people = weights.sum()
    if total_inc == 0 or total_people == 0:
        return 0.0
    order = np.argsort(values, kind='mergesort')
    block_inc = {}
    block_wt = {}
    block_keys = []
    for idx in order:
        hh = hh_ids[idx]
        if hh not in block_wt:
            block_keys.append(hh)
            block_wt[hh] = 0.0
            block_inc[hh] = 0.0
        block_wt[hh] += weights[idx]
        block_inc[hh] += wt_inc[idx]
    gini = 0.0
    cum_inc_share = 0.0
    cum_people_share = 0.0
    for hh in block_keys:
        v = block_inc[hh]
        if recode_negatives and v <= 0:
            v = 0.0
        w = block_wt[hh]
        cum_inc_share += v / total_inc
        cum_people_share += w / total_people
        gini += (w / total_people) * (cum_people_share - cum_inc_share) * 2
    return gini


def em_create_deciles(values, weights, hh_ids, dec_no=10):
    """EUROMOD CreateDeciles with GroupingVar=idhh: whole households are kept
    together; the decile advances at most once per household block, only when
    the cumulative weight before the block has reached dec/10 AND the block
    value strictly exceeds the previous block's value (ties stay in the lower
    decile). Cutoff k = previous block's value (midpoint when the weight
    boundary is hit exactly). Returns (decile per person as a Series aligned to
    the input index, cutoffs dict {1..dec_no-1})."""
    values_series = pd.Series(values)
    weights_series = pd.to_numeric(pd.Series(weights), errors='coerce').fillna(0.0)
    hh_series = pd.Series(hh_ids)
    cutoffs = {k: float('nan') for k in range(1, int(dec_no))}
    block_keys, block_vals, block_wts = _em_household_blocks(
        pd.to_numeric(values_series, errors='coerce').fillna(0.0).to_numpy(),
        weights_series.to_numpy(),
        hh_series.to_numpy(),
    )
    total_people = block_wts.sum()
    if total_people <= 0:
        return pd.Series(1, index=values_series.index), cutoffs
    dec = 1
    agg_people = 0.0
    prev_var = float('-inf')
    hh_decile = {}
    for hh, block_value, block_weight in zip(block_keys, block_vals, block_wts):
        prev_weight = agg_people
        agg_people += block_weight
        ratio = prev_weight / total_people
        threshold = dec / dec_no
        weight_comp = 0 if ratio == threshold else (1 if ratio > threshold else -1)
        if weight_comp > -1 and prev_var < block_value and dec < dec_no:
            dec += 1
            if weight_comp == 0:
                prev_var = (prev_var + block_value) / 2
            cutoffs[dec - 1] = prev_var
        prev_var = block_value
        hh_decile[hh] = dec
    deciles = hh_series.map(hh_decile).fillna(1).astype(int)
    deciles.index = values_series.index
    return deciles, cutoffs


def weighted_atkinson(values, weights, epsilon=0.25):
    values = np.asarray(values, dtype=float)
    weights = np.asarray(weights, dtype=float)
    mask = (weights > 0) & np.isfinite(values) & (values >= 0)
    if not mask.any():
        return 0.0
    values = values[mask]
    weights = weights[mask]
    total_weight = weights.sum()
    if total_weight <= 0:
        return 0.0
    mean_value = np.average(values, weights=weights)
    if mean_value <= 0:
        return 0.0
    if epsilon == 1:
        positive_mask = values > 0
        if not positive_mask.any():
            return 0.0
        log_vals = np.log(values[positive_mask])
        log_weight = weights[positive_mask]
        weighted_log_mean = np.average(log_vals, weights=log_weight)
        atkinson = 1 - np.exp(weighted_log_mean) / mean_value
    else:
        power = 1 - epsilon
        transformed = np.power(values, power, where=values > 0)
        mean_transformed = np.average(transformed, weights=weights)
        if mean_transformed <= 0:
            return 0.0
        atkinson = 1 - np.power(mean_transformed, 1 / power) / mean_value
    return max(0.0, min(1.0, atkinson))


def weighted_percentiles(values, weights, percentiles):
    values = np.asarray(values, dtype=float)
    weights = np.asarray(weights, dtype=float)
    mask = (weights > 0) & np.isfinite(values)
    if not mask.any():
        return {p: 0 for p in percentiles}
    values = values[mask]
    weights = weights[mask]
    order = np.argsort(values)
    values = values[order]
    weights = weights[order]
    cum_weights = np.cumsum(weights)
    total_weight = cum_weights[-1]
    if total_weight <= 0:
        return {p: 0 for p in percentiles}
    percentile_values = {}
    for p in percentiles:
        threshold = total_weight * (p / 100)
        idx = np.searchsorted(cum_weights, threshold, side='left')
        idx = min(idx, len(values) - 1)
        percentile_values[p] = values[idx]
    return percentile_values

def add_analysis_flags(df):
    
    # Create individual-level flags in one assign call
    flag_cols = {
        'isPers': 1,
        'isChild': (df['dag'] < 18).astype(int),
        'isElderly': (df['dag'] > 64).astype(int),
        'isAdult': (df['dag'] >= 18).astype(int),
        'isMaleAdult': ((df.get('dgn', 0) > 0) & (df['dag'] >= 18)).astype(int),
        'isYoungChild': (df['dag'] <= 2).astype(int),
        'isInformalWorker': (df.get('lfo', 0) == 0).astype(int)
    }
    df = df.assign(**flag_cols)
    
    # Create 'isInformalAdults' based on new columns
    df['isInformalAdults'] = ((df['isInformalWorker'] > 0) & (df['isAdult'] > 0)).astype(int)

    # Group once to create household totals for each flag
    hh_flag_cols = ['isChild', 'isElderly', 'isAdult', 'isMaleAdult', 'isPers', 'isYoungChild', 'ddi', 'isInformalWorker', 'isInformalAdults']
    
    # Ensure all columns exist before grouping
    for col in hh_flag_cols:
        if col not in df.columns:
            df[col] = 0
            
    # Create a DataFrame of the HH sums
    hh_sums_df = df.groupby('idhh')[hh_flag_cols].transform('sum')
    
    # Rename columns to 'n...InHH' format
    hh_sums_df.columns = [f'n{col.replace("is", "")}InHH' for col in hh_flag_cols]
    
    # Join the new count columns back to the main df
    df = df.join(hh_sums_df)
    
    # Create final household flags based on the new count columns
    hh_type_cols = {
        'isHHWithChild': (df['nChildInHH'] > 0).astype(int),
        'isAtLeastOneElderlyHH': (df['nElderlyInHH'] > 0).astype(int),
        'isAtLeastOneDisabledHH': (df['nddiInHH'] > 0).astype(int),
        'isSinglePersonHH': (df['nPersInHH'] == 1).astype(int),
        'is1AdultWithChildrenHH': ((df['nAdultInHH'] == 1) & (df['nChildInHH'] >= 1)).astype(int),
        'is2AdultsNoChildrenHH': ((df['nAdultInHH'] == 2) & (df['nChildInHH'] == 0)).astype(int),
        'is2Adults1_2ChildrenHH': ((df['nAdultInHH'] == 2) & (df['nChildInHH'].between(1, 2))).astype(int),
        'is2Adults3_4ChildrenHH': ((df['nAdultInHH'] == 2) & (df['nChildInHH'].between(3, 4))).astype(int),
        'is2Adults5plusChildrenHH': ((df['nAdultInHH'] == 2) & (df['nChildInHH'] >= 5)).astype(int),
        'is3plusAdultsNoChildrenHH': ((df['nAdultInHH'] >= 3) & (df['nChildInHH'] == 0)).astype(int),
        'is3plusAdultsWithChildrenHH': ((df['nAdultInHH'] >= 3) & (df['nChildInHH'] >= 1)).astype(int),
        'isYoungChildHH': (df['nYoungChildInHH'] > 0).astype(int),
        'isNoMaleAdultHH': (df['nMaleAdultInHH'] == 0).astype(int),
        'isInformalAdultHH': (df['nInformalAdultsInHH'] > 0).astype(int),
        'isNoInformalAdultsHH': (df['nInformalAdultsInHH'] == 0).astype(int),
        'ils_earns': df.get('yem', 0) + df.get('yse', 0) + df.get('yag', 0)
    }
    df = df.assign(**hh_type_cols)
    
    df['TotalHHEarnings'] = df.groupby('idhh')['ils_earns'].transform('sum')
    df['isNoTotalHHEarningsHH'] = (df['TotalHHEarnings'] <= 0).astype(int)
    
    return df

# --- MAIN ANALYSIS ENGINE ---
def run_analysis(sim_df, user_choice, baseline_analysis_df=None):
    results = {}
    analysis_df = sim_df.copy()
    if 'dwt' not in analysis_df.columns:
        analysis_df['dwt'] = 1.0
    analysis_df['isIndividual'] = 1
    
    # 1. DEFINE RESOURCE AND POVERTY LINE
    base_resource_map = {1: 'ils_con', 2: 'ils_dispyx', 3: 'ils_con_pf', 4: 'ils_dispyx_pf'}
    povline_map = {1: 'spl', 2: 'spl', 3: 'splpf', 4: 'splpf'}
    
    base_resource_col = base_resource_map[user_choice]
    analysis_df['ilsRank'] = analysis_df.get(base_resource_col, 0) # Use .get for safety
    analysis_df['povLine_raw'] = analysis_df.get(povline_map[user_choice], 0)
    
    # 2. CALCULATE TOTAL HOUSEHOLD RESOURCE (ilsRankHH)
    analysis_df['ilsRankHH'] = analysis_df.groupby('idhh')['ilsRank'].transform('sum')
    

    # 3. CALCULATE EQUIVALENCE SCALE (eqScale)
    head_ses = analysis_df.loc[analysis_df['dhh'] == 1, ['idhh', 'ses']].set_index('idhh')['ses']
    analysis_df['eqScale'] = analysis_df['idhh'].map(head_ses).fillna(0)

    # 4. CALCULATE EQUIVALIZED RESOURCE (eqRank)
    valid_eq_rank_mask = (analysis_df['ilsRankHH'] >= 0) & (analysis_df['eqScale'] > 0)
    analysis_df['eqRank'] = 0.0
    analysis_df.loc[valid_eq_rank_mask, 'eqRank'] = analysis_df.loc[valid_eq_rank_mask, 'ilsRankHH'] / analysis_df.loc[valid_eq_rank_mask, 'eqScale']
    # 5. DEFINE ANALYSIS GROUPS (DECILES, HH TYPES)
    if baseline_analysis_df is not None:
        # REFORM run
        hh_flag_cols_to_merge = [col for col in baseline_analysis_df if col.startswith('is') and 'HH' in col]
        base_cols_to_merge = ['deciles', 'deciles_hh', 'deciles_base', 'eqRank_baseline', 'AllIndividuals_Base_Pass']
        cols_to_merge = list(dict.fromkeys(base_cols_to_merge + hh_flag_cols_to_merge))
        analysis_df = analysis_df.drop(columns=cols_to_merge, errors='ignore').merge(
            baseline_analysis_df[['idperson'] + cols_to_merge], on='idperson', how='left'
        )
    else:
        # BASELINE run
        analysis_df = add_analysis_flags(analysis_df)
        # EUROMOD CreateDeciles (GroupingVar=idhh) writes the same person-weighted
        # decile to both the individual and household levels
        decile_series, _ = em_create_deciles(
            analysis_df['eqRank'], analysis_df['dwt'], analysis_df['idhh']
        )
        analysis_df['deciles'] = decile_series
        analysis_df['deciles_base'] = analysis_df['deciles']
        # Household deciles: same algorithm with head-only weights
        # (WeightVar=hh_weight in the statistics template)
        hh_decile_series, _ = em_create_deciles(
            analysis_df['eqRank'],
            analysis_df['dwt'] * (analysis_df['dhh'] == 1),
            analysis_df['idhh'],
        )
        analysis_df['deciles_hh'] = hh_decile_series
        analysis_df['eqRank_baseline'] = analysis_df['eqRank']
        baseline_flag_cols = [
            'isSinglePersonHH', 'is1AdultWithChildrenHH', 'is2AdultsNoChildrenHH',
            'is2Adults1_2ChildrenHH', 'is2Adults3_4ChildrenHH', 'is2Adults5plusChildrenHH',
            'is3plusAdultsNoChildrenHH', 'is3plusAdultsWithChildrenHH',
            'isYoungChildHH', 'isAtLeastOneElderlyHH', 'isAtLeastOneDisabledHH',
            'isNoMaleAdultHH', 'isNoTotalHHEarningsHH', 'isInformalAdultHH', 'isNoInformalAdultsHH'
        ]
        for flag_col in baseline_flag_cols:
            if flag_col in analysis_df.columns:
                analysis_df[f"{flag_col}_Base_Pass"] = analysis_df[flag_col]
        analysis_df['AllIndividuals_Base_Pass'] = 1
    
    if 'deciles_base' not in analysis_df.columns:
        analysis_df['deciles_base'] = analysis_df.get('deciles', 1)
    analysis_df['deciles_base'] = pd.to_numeric(analysis_df['deciles_base'], errors='coerce').fillna(1)

    if 'deciles_hh' in analysis_df.columns:
        analysis_df['deciles_hh'] = pd.to_numeric(analysis_df['deciles_hh'], errors='coerce').fillna(1)
    else:
        analysis_df['deciles_hh'] = 1
    analysis_df = assign_current_hh_deciles(analysis_df)
    analysis_df = assign_current_individual_deciles(analysis_df)

    results['gainers_losers'] = None

    if baseline_analysis_df is not None:
        baseline_eq = pd.to_numeric(analysis_df.get('eqRank_baseline'), errors='coerce').fillna(0.0)
        reform_eq = pd.to_numeric(analysis_df.get('eqRank'), errors='coerce').fillna(0.0)
        change = np.zeros(len(analysis_df), dtype=float)
        positive_mask = baseline_eq > 0
        change[positive_mask] = (reform_eq[positive_mask] - baseline_eq[positive_mask]) / baseline_eq[positive_mask]
        zero_mask = baseline_eq == 0
        if zero_mask.any():
            same_zero = zero_mask & (reform_eq == 0)
            change[same_zero] = 0.0
            positive_zero = zero_mask & (reform_eq > 0)
            change[positive_zero] = 9999.0
            negative_zero = zero_mask & (reform_eq < 0)
            change[negative_zero] = -9999.0

        analysis_df['IncomeChangeEqRank'] = change
        analysis_df['gain1_eqRank'] = (change > 0.01).astype(int)
        analysis_df['gain5_eqRank'] = (change > 0.05).astype(int)
        analysis_df['lose1_eqRank'] = (change < -0.01).astype(int)
        analysis_df['lose5_eqRank'] = (change < -0.05).astype(int)

        weight_series = pd.to_numeric(analysis_df.get('dwt'), errors='coerce').fillna(0.0)
        metric_specs = [
            ('gain1', 'gain1_eqRank'),
            ('gain5', 'gain5_eqRank'),
            ('lose1', 'lose1_eqRank'),
            ('lose5', 'lose5_eqRank'),
        ]

        def compute_group_stats(label, mask):
            if isinstance(mask, pd.Series):
                mask_series = mask.reindex(analysis_df.index, fill_value=False)
            else:
                mask_series = pd.Series(mask, index=analysis_df.index)
            mask_series = mask_series.fillna(False)
            weights = weight_series[mask_series]
            weight_sum = weights.sum()
            stats = {'label': label}
            if weight_sum <= 0:
                for metric_key, _ in metric_specs:
                    stats[metric_key] = 0.0
                return stats
            for metric_key, col_name in metric_specs:
                col_values = pd.to_numeric(analysis_df.loc[mask_series, col_name], errors='coerce').fillna(0.0)
                percentage = (col_values * weights).sum() / weight_sum * 100
                stats[metric_key] = round(float(percentage), 1)
            return stats

        def mask_from_flag(flag_col):
            series = pd.to_numeric(analysis_df.get(flag_col), errors='coerce').fillna(0.0)
            return series > 0.5

        decile_series = pd.to_numeric(analysis_df.get('deciles_base'), errors='coerce').fillna(0).astype(int)
        decile_stats = [compute_group_stats("All", pd.Series(True, index=analysis_df.index))]
        for decile in range(1, 11):
            decile_stats.append(compute_group_stats(f"Decile {decile}", decile_series == decile))

        household_groups = [
            ("All", None),
            ("Single person", "isSinglePersonHH_Base_Pass"),
            ("Single parent", "is1AdultWithChildrenHH_Base_Pass"),
            ("2 adults, no children", "is2AdultsNoChildrenHH_Base_Pass"),
            ("2 adults, 1-2 children", "is2Adults1_2ChildrenHH_Base_Pass"),
            ("2 adults, 3-4 children", "is2Adults3_4ChildrenHH_Base_Pass"),
            ("2 adults, 5+ children", "is2Adults5plusChildrenHH_Base_Pass"),
            ("3+ adults, no children", "is3plusAdultsNoChildrenHH_Base_Pass"),
            ("3+ adults, with children", "is3plusAdultsWithChildrenHH_Base_Pass"),
        ]
        household_stats = []
        for label, flag in household_groups:
            if flag is None:
                household_stats.append(compute_group_stats(label, pd.Series(True, index=analysis_df.index)))
            else:
                household_stats.append(compute_group_stats(label, mask_from_flag(flag)))

        vulnerability_groups = [
            ("All", None),
            ("Young child (aged 0-2)", "isYoungChildHH_Base_Pass"),
            ("Elderly member", "isAtLeastOneElderlyHH_Base_Pass"),
            ("Member with a disability", "isAtLeastOneDisabledHH_Base_Pass"),
            ("No male adults", "isNoMaleAdultHH_Base_Pass"),
            ("No labour market income", "isNoTotalHHEarningsHH_Base_Pass"),
            ("Informal adult", "isInformalAdultHH_Base_Pass"),
            ("No informal adults", "isNoInformalAdultsHH_Base_Pass"),
        ]
        vulnerability_stats = []
        for label, flag in vulnerability_groups:
            if flag is None:
                vulnerability_stats.append(compute_group_stats(label, pd.Series(True, index=analysis_df.index)))
            else:
                vulnerability_stats.append(compute_group_stats(label, mask_from_flag(flag)))

        results['gainers_losers'] = {
            'deciles': decile_stats,
            'household': household_stats,
            'vulnerability': vulnerability_stats,
        }

    # 6. CALCULATE POVERTY INDICATORS
    povLine = weighted_average(analysis_df, 'povLine_raw')
    
    analysis_df['isPoor'] = (analysis_df['eqRank'] < povLine).astype(int)
    
    analysis_df['povGap'] = 0.0
    if povLine > 0:
        analysis_df['povGap'] = np.maximum(0, (povLine - analysis_df['eqRank']) / povLine) * analysis_df['isPoor']

    if baseline_analysis_df is None:
        if user_choice in (1, 3):
            ilsRank_Bef = (
                analysis_df['ils_con']
                - analysis_df['ils_ben']
                + analysis_df['ils_tax']
                + analysis_df['ils_sicee']
                + analysis_df['ils_sicse']
            ).fillna(0)
        else:
            ilsRank_Bef = (analysis_df['ils_origy'] + analysis_df.get('xivot', 0)).fillna(0)
        analysis_df['ilsRank_Bef'] = ilsRank_Bef
        analysis_df['ilsRankHH_Bef'] = analysis_df.groupby('idhh')['ilsRank_Bef'].transform('sum')
        analysis_df['eqRank_Bef'] = 0.0
        # EUROMOD zeroes eqRank_Bef when the household's before-income sum is
        # negative (notZeroFlag_Bef)
        valid_eq_rank_bef = (analysis_df['eqScale'] > 0) & (analysis_df['ilsRankHH_Bef'] >= 0)
        analysis_df.loc[valid_eq_rank_bef, 'eqRank_Bef'] = (
            analysis_df.loc[valid_eq_rank_bef, 'ilsRankHH_Bef'] / analysis_df.loc[valid_eq_rank_bef, 'eqScale']
        )
        # The before-t&b poverty line is always spl, never splpf, for every choice
        povLine_for_bef = pd.to_numeric(analysis_df.get('spl', 0), errors='coerce').fillna(0)
        if not isinstance(povLine_for_bef, pd.Series):
            povLine_for_bef = pd.Series(povLine_for_bef, index=analysis_df.index)
        analysis_df['isPoor_Bef'] = (analysis_df['eqRank_Bef'] < povLine_for_bef).astype(int)
        analysis_df['povGap_Bef'] = 0.0
        positive_bef_mask = povLine_for_bef > 0
        gap_vals = (
            (povLine_for_bef[positive_bef_mask] - analysis_df.loc[positive_bef_mask, 'eqRank_Bef'])
            / povLine_for_bef[positive_bef_mask]
        ).clip(lower=0)
        analysis_df.loc[positive_bef_mask, 'povGap_Bef'] = gap_vals * analysis_df.loc[positive_bef_mask, 'isPoor_Bef']
        weights_bef = analysis_df.get('dwt', 0).fillna(0).to_numpy()
        eqrank_bef_values = analysis_df['eqRank_Bef'].fillna(0).to_numpy()
        results['PovertyRate_Bef'] = weighted_average(analysis_df, 'isPoor_Bef') * 100
        results['PovertyGap_Bef'] = weighted_average(analysis_df, 'povGap_Bef') * 100
        results['Gini_Bef'] = em_gini_grouped(eqrank_bef_values, weights_bef, analysis_df['idhh'].to_numpy()) * 100
        bef_percentiles = weighted_percentiles(eqrank_bef_values, weights_bef, [20, 80])
        results['P20_Bef'] = bef_percentiles.get(20, 0)
        results['P80_Bef'] = bef_percentiles.get(80, 0)

    # Track baseline poverty status at the household level so reform runs can reuse it
    if 'isPoor_BaselineDefinitionHH' not in analysis_df.columns:
        hh_poor_flag = analysis_df.groupby('idhh')['isPoor'].transform('max').fillna(0)
        analysis_df['isPoor_BaselineDefinitionHH'] = hh_poor_flag.astype(int)
    
    # 6b. HOUSEHOLD-LEVEL FLAGS AND HEAD-SPECIFIC VARIABLES
    analysis_df['TotalHouseholds'] = 1

    # Ensure key monetary columns exist when computing tax aggregates
    for col_name in ['ils_tax', 'ils_taxco', 'ils_sicee', 'ils_sicse', 'ils_sic', 'ils_origy', 'ils_ben', 'ils_benki']:
        if col_name not in analysis_df.columns:
            analysis_df[col_name] = 0.0
        else:
            analysis_df[col_name] = analysis_df[col_name].fillna(0)
    analysis_df['ssc_ee_se_total'] = analysis_df['ils_sicee'] + analysis_df['ils_sicse']

    hh_component_sources = {
        'HHpaysDirTax': 'ils_tax',
        'HHpaysSSC': 'ils_sic',
        'HHpaysIndirTax': 'tva_s',
        'HHpaysSSC_EE_SE': 'ssc_ee_se_total',
        'HHgetsCashBen': 'ils_ben',
        'HHgetsInKindBen': 'ils_benki',
        'HHgetsIndirSub': 'ils_benco',
        'HHgetsChildBen': 'ils_bch',
        'HHgetsSocAss': 'ils_bsa',
        'HHgetsOrphBen': 'ils_bsu',
        'HHgetsDisBen': 'ils_bdi',
        'HHgetsUeBen': 'ils_bun',
        'HHgetsPenBen': 'ils_pen',
        'HHgetsAgBen': 'ils_bag',
    }
    for source_col in hh_component_sources.values():
        if source_col not in analysis_df.columns:
            analysis_df[source_col] = 0
    hh_indicator_cols = {}
    for flag_name, source_col in hh_component_sources.items():
        hh_totals = analysis_df.groupby('idhh')[source_col].transform('sum')
        hh_indicator_cols[flag_name] = (hh_totals > 0).astype(int)
    analysis_df = analysis_df.assign(**hh_indicator_cols)
    analysis_df['HHpaysAnyTaxOrCont'] = (
        analysis_df[['HHpaysDirTax', 'HHpaysSSC', 'HHpaysIndirTax']].sum(axis=1) > 0
    ).astype(int)
    analysis_df['HHgetsAnyBenefit'] = (
        analysis_df[['HHgetsCashBen', 'HHgetsInKindBen', 'HHgetsIndirSub']].sum(axis=1) > 0
    ).astype(int)
    analysis_df['HHgetsAnyDirectBenefit'] = (
        analysis_df[['HHgetsCashBen', 'HHgetsInKindBen']].sum(axis=1) > 0
    ).astype(int)
    head_flag_sources = [
        'TotalHouseholds',
        'HHpaysAnyTaxOrCont',
        'HHpaysDirTax',
        'HHpaysSSC',
        'HHpaysSSC_EE_SE',
        'HHpaysIndirTax',
        'HHgetsAnyDirectBenefit',
        'HHgetsAnyBenefit',
        'HHgetsCashBen',
        'HHgetsInKindBen',
        'HHgetsIndirSub',
        'HHgetsChildBen',
        'HHgetsSocAss',
        'HHgetsOrphBen',
        'HHgetsDisBen',
        'HHgetsUeBen',
        'HHgetsPenBen',
        'HHgetsAgBen',
        'isSinglePersonHH',
        'is1AdultWithChildrenHH',
        'is2AdultsNoChildrenHH',
        'is2Adults1_2ChildrenHH',
        'is2Adults3_4ChildrenHH',
        'is2Adults5plusChildrenHH',
        'is3plusAdultsNoChildrenHH',
        'is3plusAdultsWithChildrenHH',
        'isHHWithChild',
        'isYoungChildHH',
        'isAtLeastOneElderlyHH',
        'isAtLeastOneDisabledHH',
        'isNoMaleAdultHH',
        'isNoTotalHHEarningsHH',
        'isInformalAdultHH',
        'isNoInformalAdultsHH',
        'isPoor_BaselineDefinitionHH',
    ]
    for flag_name in head_flag_sources:
        if flag_name not in analysis_df.columns:
            analysis_df[flag_name] = 0
        analysis_df[f'Head_{flag_name}'] = analysis_df['dhh'] * analysis_df[flag_name]
    
    is_head_df = analysis_df[analysis_df['dhh']==1]
    analysis_df['isPoor_BaselineDefinition'] = analysis_df.get('isPoor_BaselineDefinitionHH', 0)

    hh_cash_ben = analysis_df.groupby('idhh')['ils_ben'].transform('sum')
    hh_inkind_ben = analysis_df.groupby('idhh')['ils_benki'].transform('sum')
    analysis_df['HH_CashBenefitTotal'] = hh_cash_ben
    analysis_df['HH_InKindBenefitTotal'] = hh_inkind_ben
    analysis_df['AnyDirectBenefitAmount'] = analysis_df['ils_ben'] + analysis_df['ils_benki']
    positive_eq_scale = analysis_df['eqScale'] > 0
    analysis_df['eq_indiv_cash_ben'] = 0.0
    analysis_df['eq_indiv_inkind_ben'] = 0.0
    analysis_df.loc[positive_eq_scale, 'eq_indiv_cash_ben'] = (
        hh_cash_ben[positive_eq_scale] / analysis_df.loc[positive_eq_scale, 'eqScale']
    )
    analysis_df.loc[positive_eq_scale, 'eq_indiv_inkind_ben'] = (
        hh_inkind_ben[positive_eq_scale] / analysis_df.loc[positive_eq_scale, 'eqScale']
    )
    analysis_df['IndivGetsCashBen'] = (analysis_df['ils_ben'] > 0).astype(int)
    analysis_df['IndivGetsInKindBen'] = (analysis_df['ils_benki'] > 0).astype(int)

    def compute_equivalised_resource(column_name):
        if column_name not in analysis_df.columns:
            analysis_df[column_name] = 0.0
        hh_total = analysis_df.groupby('idhh')[column_name].transform('sum')
        eq_series = pd.Series(0.0, index=analysis_df.index)
        eq_series.loc[positive_eq_scale] = hh_total[positive_eq_scale] / analysis_df.loc[positive_eq_scale, 'eqScale']
        return eq_series

    cons_resource_col = 'ils_con_pf' if user_choice in (3, 4) else 'ils_con'
    inc_resource_col = 'ils_dispyx_pf' if user_choice == 4 else 'ils_dispyx'
    eq_cons_resource = compute_equivalised_resource(cons_resource_col)
    eq_inc_resource = compute_equivalised_resource(inc_resource_col)

    # Benefit-adequacy denominators: EUROMOD's median_con_calc / median_dispyx_calc
    # are the decile-5 cutoff from CreateDeciles run on raw person-level ils_con /
    # ils_dispyx (grouped by idhh; never the _pf variants, never equivalised,
    # never choice-dependent), x12 for yearly, fixed to the baseline scenario.
    # The degenerate household ordering on raw person values is EUROMOD's literal
    # behaviour - do not "fix" it.
    def compute_decile5_cutoff_yearly(df, value_col):
        if value_col not in df.columns:
            return 0.0
        vals = pd.to_numeric(df[value_col], errors='coerce').fillna(0.0)
        _, cutoffs = em_create_deciles(vals, df['dwt'], df['idhh'])
        cutoff = cutoffs.get(5)
        if cutoff is None or not np.isfinite(cutoff):
            return 0.0
        return cutoff * 12

    median_source = analysis_df
    if baseline_analysis_df is not None and {'idhh', 'dwt', 'ils_con', 'ils_dispyx'}.issubset(baseline_analysis_df.columns):
        median_source = baseline_analysis_df
    cons_median_yearly = compute_decile5_cutoff_yearly(median_source, 'ils_con')
    inc_median_yearly = compute_decile5_cutoff_yearly(median_source, 'ils_dispyx')
    results['BaselineMedianEqConsYearly'] = cons_median_yearly
    results['BaselineMedianEqIncYearly'] = inc_median_yearly
    
    # 7. AGGREGATE RESULTS
    results['taxbenpol_abs'] = {
        'Direct taxes': weighted_sum(analysis_df, 'ils_tax'),
        'Social insurance contributions': weighted_sum(analysis_df, 'ils_sic'),
        'Indirect taxes': weighted_sum(is_head_df, 'tva_s'), # tva_s is HH total on head
        'Cash benefits': weighted_sum(analysis_df, 'ils_ben'),
        'In-kind benefits': weighted_sum(analysis_df, 'ils_benki'),
        'Indirect subsidies': weighted_sum(analysis_df, 'ils_benco'),
        'Child benefits': weighted_sum(analysis_df, 'ils_bch'),
        'Social assistance': weighted_sum(analysis_df, 'ils_bsa'),
    }
    
    total_rev = sum(results['taxbenpol_abs'][name] for name in ['Direct taxes', 'Social insurance contributions', 'Indirect taxes'])
    total_exp = sum(results['taxbenpol_abs'][name] for name in ['Cash benefits', 'In-kind benefits', 'Indirect subsidies'])
    results['taxbenpol_abs']['Sum of government revenue'] = total_rev
    results['taxbenpol_abs']['Sum of government expenditure'] = total_exp

    results['taxbenpol_share'] = {
        name: (results['taxbenpol_abs'][name] / total_rev * 100) if total_rev > 0 else 0 for name in ['Direct taxes', 'Social insurance contributions', 'Indirect taxes']
    }
    results['taxbenpol_share'].update({
        name: (results['taxbenpol_abs'][name] / total_exp * 100) if total_exp > 0 else 0 for name in ['Cash benefits', 'In-kind benefits', 'Indirect subsidies']
    })
    
    household_results = {
        'TotalHHCount': weighted_sum(analysis_df, 'Head_TotalHouseholds'),
    }
    hh_result_map = {
        'CountHHpaysAnyTaxOrCont': 'Head_HHpaysAnyTaxOrCont',
        'CountHHpaysDirTax': 'Head_HHpaysDirTax',
        'CountHHpaysSSC': 'Head_HHpaysSSC',
        'CountHHpaysIndirTax': 'Head_HHpaysIndirTax',
        'CountHHgetsAnyBenefit': 'Head_HHgetsAnyBenefit',
        'CountHHgetsCashBen': 'Head_HHgetsCashBen',
        'CountHHgetsInKindBen': 'Head_HHgetsInKindBen',
        'CountHHgetsIndirSub': 'Head_HHgetsIndirSub',
        'CountHHgetsChildBen': 'Head_HHgetsChildBen',
        'CountHHgetsSocAss': 'Head_HHgetsSocAss',
        'CountHHgetsOrphBen': 'Head_HHgetsOrphBen',
        'CountHHgetsDisBen': 'Head_HHgetsDisBen',
        'CountHHgetsUeBen': 'Head_HHgetsUeBen',
        'CountHHgetsPenBen': 'Head_HHgetsPenBen',
        'CountHHgetsAgBen': 'Head_HHgetsAgBen',
    }
    for result_key, head_col in hh_result_map.items():
        household_results[result_key] = weighted_sum(analysis_df, head_col)
    hh_category_map = {
        'CountHH_New_SinglePerson': 'Head_isSinglePersonHH',
        'CountHH_New_1AdultWithChildren': 'Head_is1AdultWithChildrenHH',
        'CountHH_New_2AdultsNoChildren': 'Head_is2AdultsNoChildrenHH',
        'CountHH_New_2Adults1_2Children': 'Head_is2Adults1_2ChildrenHH',
        'CountHH_New_2Adults3_4Children': 'Head_is2Adults3_4ChildrenHH',
        'CountHH_New_2Adults5plusChildren': 'Head_is2Adults5plusChildrenHH',
        'CountHH_New_3plusAdultsNoChildren': 'Head_is3plusAdultsNoChildrenHH',
        'CountHH_New_3plusAdultsWithChildren': 'Head_is3plusAdultsWithChildrenHH',
        'CountHH_New_YoungChild': 'Head_isYoungChildHH',
        'CountHH_New_ElderlyMember': 'Head_isAtLeastOneElderlyHH',
        'CountHH_New_DisabledMember': 'Head_isAtLeastOneDisabledHH',
        'CountHH_New_NoMaleAdult': 'Head_isNoMaleAdultHH',
        'CountHH_New_NoLaborIncome': 'Head_isNoTotalHHEarningsHH',
        'CountHH_New_InformalAdult': 'Head_isInformalAdultHH',
        'CountHH_New_NoInformalAdults': 'Head_isNoInformalAdultsHH',
    }
    for result_key, head_col in hh_category_map.items():
        household_results[result_key] = weighted_sum(analysis_df, head_col)
    head_mask = analysis_df['dhh'] == 1
    for decile in range(1, 11):
        decile_mask = head_mask & (analysis_df['deciles_hh_current'] == decile)
        decile_df = analysis_df.loc[decile_mask]
        household_results[f'CountHHDecile{decile}'] = weighted_sum(decile_df, 'Head_TotalHouseholds')
    results['households'] = household_results

    eq_values = analysis_df.get('eqRank', 0).fillna(0).to_numpy()
    ind_weights = analysis_df.get('dwt', 0).fillna(0).to_numpy()
    inequality_percentiles_list = [10, 20, 30, 40, 50, 60, 70, 80, 90]
    percentiles_map = weighted_percentiles(eq_values, ind_weights, inequality_percentiles_list)
    inequality_results = {
        'Gini': em_gini_grouped(eq_values, ind_weights, analysis_df['idhh'].to_numpy()) * 100,
        'Atkinson': weighted_atkinson(eq_values, ind_weights, epsilon=0.25) * 100,
        'Percentiles': percentiles_map,
        'MeanEqRank': weighted_average(analysis_df, 'eqRank'),
        'TotalEqRank': weighted_sum(analysis_df, 'eqRank'),
        'povLine': povLine,
    }
    deciles_base_series = analysis_df.get('deciles_base', 1)
    for decile in range(1, 11):
        decile_mask = deciles_base_series == decile
        inequality_results[f'SumEqRank_InBaselineDec{decile}'] = weighted_sum(analysis_df[decile_mask], 'eqRank')
    results['Gini_Aft'] = inequality_results.get('Gini', 0)
    percentiles_after = inequality_results.get('Percentiles', {})
    results['P20_Aft'] = percentiles_after.get(20, 0)
    results['P80_Aft'] = percentiles_after.get(80, 0)
    results['inequality'] = inequality_results

    weight_series = analysis_df['dwt'].fillna(0)

    def get_series(col_name):
        if col_name in analysis_df.columns:
            return analysis_df[col_name]
        return pd.Series(0, index=analysis_df.index)

    def weighted_condition(mask):
        return float((weight_series * mask.astype(int)).sum())

    individual_results = {
        'TotalIndCount': weighted_sum(analysis_df, 'isIndividual'),
    }
    ind_tax_cols = {
        'CountIndPaysDirTax': get_series('ils_tax') > 0,
        'CountIndPaysSSC': get_series('ils_sic') > 0,
        'CountIndPaysIndirTax': get_series('tva_s') > 0,
    }
    individual_results['CountIndPaysAnyTaxOrCont'] = weighted_condition(
        ind_tax_cols['CountIndPaysDirTax'] | ind_tax_cols['CountIndPaysSSC'] | ind_tax_cols['CountIndPaysIndirTax']
    )
    for result_key, mask in ind_tax_cols.items():
        individual_results[result_key] = weighted_condition(mask)

    ind_benefit_cols = {
        'CountIndGetsCashBen': get_series('ils_ben') > 0,
        'CountIndGetsInKindBen': get_series('ils_benki') > 0,
        'CountIndGetsIndirSub': get_series('ils_benco') > 0,
    }
    individual_results['CountIndGetsAnyBenefit'] = weighted_condition(
        ind_benefit_cols['CountIndGetsCashBen'] | ind_benefit_cols['CountIndGetsInKindBen'] | ind_benefit_cols['CountIndGetsIndirSub']
    )
    for result_key, mask in ind_benefit_cols.items():
        individual_results[result_key] = weighted_condition(mask)

    ind_category_map = {
        'CountInd_New_SinglePerson': 'isSinglePersonHH',
        'CountInd_New_1AdultWithChildren': 'is1AdultWithChildrenHH',
        'CountInd_New_2AdultsNoChildren': 'is2AdultsNoChildrenHH',
        'CountInd_New_2Adults1_2Children': 'is2Adults1_2ChildrenHH',
        'CountInd_New_2Adults3_4Children': 'is2Adults3_4ChildrenHH',
        'CountInd_New_2Adults5plusChildren': 'is2Adults5plusChildrenHH',
        'CountInd_New_3plusAdultsNoChildren': 'is3plusAdultsNoChildrenHH',
        'CountInd_New_3plusAdultsWithChildren': 'is3plusAdultsWithChildrenHH',
        'CountInd_New_YoungChild': 'isYoungChildHH',
        'CountInd_New_ElderlyMember': 'isAtLeastOneElderlyHH',
        'CountInd_New_DisabledMember': 'isAtLeastOneDisabledHH',
        'CountInd_New_NoMaleAdult': 'isNoMaleAdultHH',
        'CountInd_New_NoLaborIncome': 'isNoTotalHHEarningsHH',
        'CountInd_New_InformalAdult': 'isInformalAdultHH',
        'CountInd_New_NoInformalAdults': 'isNoInformalAdultsHH',
    }
    for result_key, flag_col in ind_category_map.items():
        if flag_col in analysis_df.columns:
            individual_results[result_key] = weighted_sum(analysis_df, flag_col)
        else:
            individual_results[result_key] = 0

    for decile in range(1, 11):
        decile_mask = analysis_df.get('deciles_ind_current', 1) == decile
        decile_df = analysis_df.loc[decile_mask]
        individual_results[f'CountIndDecile{decile}'] = weighted_sum(decile_df, 'isIndividual')

    results['individuals'] = individual_results

    weights = analysis_df['dwt'] if 'dwt' in analysis_df.columns else pd.Series(0, index=analysis_df.index)

    def head_weighted_sum(col_name):
        if col_name not in analysis_df.columns:
            return 0.0
        return weighted_sum(analysis_df, col_name)

    def head_weighted_intersection(col_a, col_b):
        if col_a not in analysis_df.columns or col_b not in analysis_df.columns:
            return 0.0
        return float(((analysis_df[col_a] * analysis_df[col_b]) * weights).sum())

    denom_mapping = {
        'TotalHHCount': 'Head_TotalHouseholds',
        'CountPoorHH_BaselineDefinition': 'Head_isPoor_BaselineDefinitionHH',
        'CountHH_AtLeastOneChild': 'Head_isHHWithChild',
        'CountHH_AtLeastOneElderly': 'Head_isAtLeastOneElderlyHH',
        'CountHH_New_NoMaleAdult': 'Head_isNoMaleAdultHH',
        'CountHH_New_InformalAdult': 'Head_isInformalAdultHH',
    }
    denom_values = {}
    for result_key, col_name in denom_mapping.items():
        denom_val = head_weighted_sum(col_name)
        denom_values[result_key] = denom_val

    # Benefits tab aggregates
    benefits_results = {}
    for key, value in denom_values.items():
        benefits_results[key] = value

    benefit_indicator_cols = {
        'AnyBen': 'Head_HHgetsAnyDirectBenefit',
        'CashBen': 'Head_HHgetsCashBen',
        'InKindBen': 'Head_HHgetsInKindBen',
    }
    benefit_category_cols = {
        'BaselinePoorHH': 'Head_isPoor_BaselineDefinitionHH',
        'ChildHH': 'Head_isHHWithChild',
        'ElderlyHH': 'Head_isAtLeastOneElderlyHH',
        'NoMaleHH': 'Head_isNoMaleAdultHH',
        'InformalAdultHH': 'Head_isInformalAdultHH',
    }

    for ben_suffix, ben_col in benefit_indicator_cols.items():
        benefits_results[f'Count_AllHH_Gets{ben_suffix}'] = head_weighted_sum(ben_col)
        for cat_suffix, cat_col in benefit_category_cols.items():
            result_key = f'Count_{cat_suffix}_Gets{ben_suffix}'
            benefits_results[result_key] = head_weighted_intersection(ben_col, cat_col)

    benefits_results['CountBaselinePoorHHGetsAnyDirectBenefit_Reform'] = benefits_results.get('Count_BaselinePoorHH_GetsAnyBen', 0.0)
    benefits_results['TotalDirectCashBenefits'] = weighted_sum(analysis_df, 'ils_ben')
    benefits_results['TotalInKindBenefitAmount'] = weighted_sum(analysis_df, 'ils_benki')
    benefits_results['TotalAnyDirectBenefitAmount'] = benefits_results['TotalDirectCashBenefits'] + benefits_results['TotalInKindBenefitAmount']
    poor_mask = analysis_df['isPoor_BaselineDefinition'] == 1
    benefits_results['SumBen_BaselinePoorHH'] = weighted_sum(analysis_df[poor_mask], 'ils_ben')
    benefits_results['SumInKindBen_BaselinePoorHH'] = weighted_sum(analysis_df[poor_mask], 'ils_benki')
    benefits_results['SumAnyDirectBenefitAmount_BaselinePoorHH'] = weighted_sum(analysis_df[poor_mask], 'AnyDirectBenefitAmount')
    cash_ben_mask = analysis_df['IndivGetsCashBen'] == 1
    inkind_ben_mask = analysis_df['IndivGetsInKindBen'] == 1
    benefits_results['Mean_eq_indiv_cash_ben_yearly'] = weighted_average(analysis_df[cash_ben_mask], 'eq_indiv_cash_ben') * 12
    benefits_results['Mean_eq_indiv_inkind_ben_yearly'] = weighted_average(analysis_df[inkind_ben_mask], 'eq_indiv_inkind_ben') * 12

    for decile in range(1, 11):
        decile_mask = deciles_base_series == decile
        decile_df = analysis_df.loc[decile_mask]
        benefits_results[f'SumBen_InBaselineDec{decile}'] = weighted_sum(decile_df, 'ils_ben')
        benefits_results[f'SumInKindBen_InBaselineDec{decile}'] = weighted_sum(decile_df, 'ils_benki')

    results['benefits'] = benefits_results

    # Taxes tab aggregates
    taxes_results = {}
    for key, value in denom_values.items():
        taxes_results[key] = value

    tax_indicator_cols = {
        'DirTax': 'Head_HHpaysDirTax',
        'IndirTax': 'Head_HHpaysIndirTax',
        'SSC_EE_SE': 'Head_HHpaysSSC_EE_SE',
    }
    tax_category_cols = {
        'BaselinePoorHH': 'Head_isPoor_BaselineDefinitionHH',
        'ChildHH': 'Head_isHHWithChild',
        'ElderlyHH': 'Head_isAtLeastOneElderlyHH',
        'NoMaleHH': 'Head_isNoMaleAdultHH',
        'InformalAdultHH': 'Head_isInformalAdultHH',
    }

    for tax_suffix, tax_col in tax_indicator_cols.items():
        taxes_results[f'Count_AllHH_Pays{tax_suffix}'] = head_weighted_sum(tax_col)
        for cat_suffix, cat_col in tax_category_cols.items():
            result_key = f'Count_{cat_suffix}_Pays{tax_suffix}'
            taxes_results[result_key] = head_weighted_intersection(tax_col, cat_col)

    taxes_results['TotalOriginalIncome'] = weighted_sum(analysis_df, 'ils_origy')
    taxes_results['TotalDirectTaxes'] = weighted_sum(analysis_df, 'ils_tax')
    taxes_results['TotalIndirectTaxes'] = weighted_sum(analysis_df, 'ils_taxco')
    taxes_results['TotalSICEE_SE'] = weighted_sum(analysis_df, 'ils_sicee') + weighted_sum(analysis_df, 'ils_sicse')

    for decile in range(1, 11):
        decile_mask = deciles_base_series == decile
        decile_df = analysis_df.loc[decile_mask]
        taxes_results[f'SumDirTax_InBaselineDec{decile}'] = weighted_sum(decile_df, 'ils_tax')
        taxes_results[f'SumIndirTax_InBaselineDec{decile}'] = weighted_sum(decile_df, 'ils_taxco')
        taxes_results[f'SumSICEE_SE_InBaselineDec{decile}'] = (
            weighted_sum(decile_df, 'ils_sicee') + weighted_sum(decile_df, 'ils_sicse')
        )

    results['taxes'] = taxes_results

    results['poverty'] = {}
    # Include header markers alongside subgroup flags
    subgroup_flags = [
        'All individuals', 
        'header_hh_structure', 'isSinglePersonHH', 'is1AdultWithChildrenHH', 'is2AdultsNoChildrenHH',
        'is2Adults1_2ChildrenHH', 'is2Adults3_4ChildrenHH', 'is2Adults5plusChildrenHH',
        'is3plusAdultsNoChildrenHH', 'is3plusAdultsWithChildrenHH', 
        'header_vulnerable', 'isYoungChildHH', 'isAtLeastOneElderlyHH', 'isAtLeastOneDisabledHH', 
        'isNoMaleAdultHH',
        'header_labor', 'isNoTotalHHEarningsHH', 'isInformalAdultHH', 'isNoInformalAdultsHH'
    ]
    
    for flag in subgroup_flags:
        # Treat header markers separately
        if flag.startswith('header_'):
            results['poverty'][flag] = {'Poverty rate (%)': None, 'Poverty gap (%)': None}
            continue
            
        if flag == 'All individuals':
             sub_df = analysis_df
        elif flag in analysis_df.columns:
            sub_df = analysis_df[analysis_df[flag] == 1]
        else:
            sub_df = pd.DataFrame(columns=analysis_df.columns) # Empty df
            
        results['poverty'][flag] = {
            'Poverty rate (%)': weighted_average(sub_df, 'isPoor') * 100,
            'Poverty gap (%)': weighted_average(sub_df, 'povGap') * 100,
        }
    results['poverty']['povline'] = povLine
    all_individuals_poverty = results['poverty'].get('All individuals', {})
    results['PovertyRate_Aft'] = all_individuals_poverty.get('Poverty rate (%)')
    results['PovertyGap_Aft'] = all_individuals_poverty.get('Poverty gap (%)')

    return results, analysis_df

# --- UI HELPER FUNCTIONS ---
# Helper for building parameter rows with consistent styling
# Units and other qualifications live in a hover tip rather than in the label,
# so the parameter list stays short
RATE_TIP = "A share, not a percentage: 0.05 means 5%"
ANNUAL_TIP = "Annual value"
MONTHLY_TIP = "Monthly value"
AGE_TIP = "Age in years"


def tip_label(text, tip=None):
    """Label text that reveals its units and caveats on hover."""
    if not tip:
        return text
    return html.Span(text, className="param-tip", **{'data-tip': tip})


def make_param_input(label, param_id, value, step=None, label_width=7, input_width=5,
                     disabled=False, tip=None):
    """Creates a neatly formatted row for a parameter input."""
    return dbc.Row([
        dbc.Label(tip_label(label, tip), html_for=param_id, width=label_width,
                  style={'font-size': '0.9rem'}),
        dbc.Col(
            create_param_input_component(param_id, value, disabled=disabled),
            width=input_width
        ),
    ], className="param-input-row align-items-center")

# Helper for rendering the PIT bracket table
def make_pit_table(params):
    """Creates a table for PIT brackets."""
    header = [html.Thead(html.Tr([
        html.Th("Bracket", style={'font-size': '0.9rem'}),
        html.Th(tip_label("Lower limit", "Yearly value"), style={'font-size': '0.9rem'}),
        html.Th(tip_label("Marginal rate", RATE_TIP), style={'font-size': '0.9rem'})
    ]))]
    
    body = html.Tbody([
        html.Tr([
            html.Td("1"),
            html.Td(create_param_input_component('pit_bracket1_thresh', params['pit_bracket1_thresh'], disabled=True)),
            html.Td(create_param_input_component('pit_bracket1_rate', params['pit_bracket1_rate'], disabled=True))
        ]),
        html.Tr([
            html.Td("2"),
            html.Td(create_param_input_component('pit_bracket2_thresh', params['pit_bracket2_thresh'])),
            html.Td(create_param_input_component('pit_bracket2_rate', params['pit_bracket2_rate']))
        ]),
        html.Tr([
            html.Td("3"),
            html.Td(create_param_input_component('pit_bracket3_thresh', params['pit_bracket3_thresh'])),
            html.Td(create_param_input_component('pit_bracket3_rate', params['pit_bracket3_rate']))
        ]),
        html.Tr([
            html.Td("4"),
            html.Td(create_param_input_component('pit_bracket4_thresh', params['pit_bracket4_thresh'])),
            html.Td(create_param_input_component('pit_bracket4_rate', params['pit_bracket4_rate']))
        ]),
        html.Tr([
            html.Td("5"),
            html.Td(create_param_input_component('pit_bracket5_thresh', params['pit_bracket5_thresh'])),
            html.Td(create_param_input_component('pit_bracket5_rate', params['pit_bracket5_rate']))
        ]),
    ])
    
    return dbc.Table(header + [body], bordered=True, size="sm", responsive=True)

def make_control_step(step_number: str, title: str, element_id: str = None) -> html.Div:
    """Creates a highlighted heading for the controls panel."""
    kwargs = {'id': element_id} if element_id else {}
    return html.Div(
        [
            html.Span(step_number, className="control-step-number"),
            html.Span(title, className="control-step-title"),
        ],
        className="control-step-header d-flex align-items-center gap-2 mb-2",
        **kwargs
    )

# Render tables with consistent styling
def ensure_parenthesized(text: str) -> str:
    stripped = text.strip()
    if not stripped:
        return ""
    if stripped.startswith("(") and stripped.endswith(")"):
        return stripped
    return f"({stripped})"


def normalize_title_subtitle(title: str, subtitle: str) -> tuple[str, str]:
    main_title = (title or "").strip()
    sub_title = (subtitle or "").strip()

    if not sub_title:
        if main_title.endswith(")") and "(" in main_title:
            open_idx = main_title.rfind("(")
            candidate = main_title[open_idx + 1 : -1].strip()
            if candidate:
                sub_title = ensure_parenthesized(candidate)
                main_title = main_title[:open_idx].rstrip(", ").rstrip()
        if not sub_title and ", " in main_title:
            main_part, suffix = main_title.split(", ", 1)
            if suffix:
                sub_title = ensure_parenthesized(suffix)
                main_title = main_part.strip()
    else:
        sub_title = ensure_parenthesized(sub_title)

    return main_title, sub_title


def nice_number(value: float, round_up: bool = True) -> float:
    if value == 0:
        return 0.0
    exponent = math.floor(math.log10(abs(value)))
    fraction = abs(value) / (10 ** exponent)
    if round_up:
        if fraction <= 1:
            nice_fraction = 1
        elif fraction <= 2:
            nice_fraction = 2
        elif fraction <= 5:
            nice_fraction = 5
        else:
            nice_fraction = 10
    else:
        if fraction < 1:
            nice_fraction = 0.5
        elif fraction < 2:
            nice_fraction = 1
        elif fraction < 5:
            nice_fraction = 2
        else:
            nice_fraction = 5
    return math.copysign(nice_fraction * (10 ** exponent), value)


def compute_axis_settings(values) -> dict:
    numeric = [
        float(v)
        for v in values
        if isinstance(v, (int, float)) and not pd.isna(v)
    ]
    if not numeric:
        return {
            "decimals": 2,
            "dtick": None,
            "range": None,
            "tickformat": ".2f",
        }

    min_val = min(numeric)
    max_val = max(numeric)
    span = max_val - min_val
    max_abs = max(abs(min_val), abs(max_val))

    if span == 0:
        baseline_interval = max_abs / 5 if max_abs != 0 else 0.2
    else:
        baseline_interval = span / 5
    if baseline_interval == 0:
        baseline_interval = 0.2

    tick_interval = nice_number(baseline_interval, round_up=True)
    if tick_interval == 0:
        tick_interval = baseline_interval or 0.2

    if span == 0:
        if max_val > 0:
            axis_min = max(0.0, max_val - tick_interval)
            if axis_min >= max_val:
                axis_min = max(0.0, axis_min - tick_interval)
            axis_max = max_val + tick_interval
        elif max_val < 0:
            axis_min = min_val - tick_interval
            axis_max = min(0.0, max_val + tick_interval)
            if axis_max <= max_val:
                axis_max += tick_interval
        else:
            axis_min = 0.0
            axis_max = tick_interval
    else:
        positive_only = min_val >= 0
        negative_only = max_val <= 0

        if positive_only:
            # Always start from 0 for positive values
            axis_min = 0.0
            axis_max = math.ceil(max_val / tick_interval) * tick_interval
            if axis_max <= max_val:
                axis_max += tick_interval
        elif negative_only:
            # Always end at 0 for negative values
            axis_max = 0.0
            axis_min = math.floor(min_val / tick_interval) * tick_interval
            if axis_min >= min_val:
                axis_min -= tick_interval
        else:
            # Mixed positive and negative - ensure 0 is included
            axis_min = math.floor(min_val / tick_interval) * tick_interval
            if axis_min >= min_val:
                axis_min -= tick_interval
            axis_max = math.ceil(max_val / tick_interval) * tick_interval
            if axis_max <= max_val:
                axis_max += tick_interval

    if axis_max <= axis_min:
        axis_max = axis_min + tick_interval

    # Extend the range slightly beyond the last tick to ensure its grid line is visible
    # Add a small buffer (5% of tick interval) to show the grid line for the outermost tick
    buffer = tick_interval * 0.05
    axis_min -= buffer
    axis_max += buffer

    decimals = 0
    if tick_interval < 1:
        decimals = min(6, max(0, -int(math.floor(math.log10(tick_interval)))))

    tickformat = ",.0f" if decimals == 0 else f".{decimals}f"

    return {
        "decimals": decimals,
        "dtick": tick_interval,
        "range": [axis_min, axis_max],
        "tickformat": tickformat,
    }


def format_axis_value(value, decimals: int):
    if value is None or pd.isna(value):
        return "n/a"
    fmt = f"{{:,.{decimals}f}}" if decimals > 0 else "{:,.0f}"
    formatted = fmt.format(value)
    if decimals > 0 and "." in formatted:
        formatted = formatted.rstrip("0").rstrip(".")
    return formatted


def create_styled_table(data_dict, title, subtitle):
    title, subtitle = normalize_title_subtitle(title, subtitle)
    if data_dict:
        original_keys = list(data_dict.keys())
        replace_first = original_keys[0] in ('Component', 'Household category')
        table_dict = {}
        for idx, key in enumerate(original_keys):
            new_key = '' if replace_first and idx == 0 else key
            table_dict[new_key] = data_dict[key]
    else:
        table_dict = data_dict

    max_len = 0
    if table_dict:
         max_len = max(len(v) for v in table_dict.values() if v is not None)
         
    for col in table_dict:
        if table_dict[col] is None: table_dict[col] = []
        if len(table_dict[col]) < max_len:
            table_dict[col].extend([None] * (max_len - len(table_dict[col]))) # Use None, not ''

    column_names = list(table_dict.keys()) if table_dict else []
    num_columns = len(column_names)

    header_cells = []
    for idx, col in enumerate(column_names):
        th_classes = "table-header-cell"
        if idx > 0:
            th_classes += " text-end"
        header_cells.append(html.Th(col, className=th_classes))
    header = [html.Thead(html.Tr(header_cells))]
    
    body_rows = []
    if not table_dict:
        return [html.P("No data for table.")]
        
    row_titles_col_name = next(iter(table_dict))
    
    for i, row_title in enumerate(table_dict[row_titles_col_name]):
        if row_title is None:
            continue # Should not happen, but good to check

        row_title_str = str(row_title)
        display_title = row_title_str[1:].strip() if row_title_str.startswith('-') else row_title_str
        is_header_row = row_title_str in TABLE_HEADER_LABELS
        is_sub_row = row_title_str.startswith('-') and row_title_str not in TABLE_STRONG_LABELS
        is_strong_row = row_title_str in TABLE_STRONG_LABELS

        row_classes = []
        if row_title_str in TABLE_DIVIDER_LABELS:
            row_classes.append("table-section-divider")
        if is_header_row:
            row_classes.append("table-section-header-row")
        if is_strong_row:
            row_classes.append("table-row-strong")
        if is_sub_row:
            row_classes.append("table-sub-row")

        first_cell_classes = ["table-first-column"]
        if is_header_row:
            first_cell_classes.append("table-first-column-header")
        if is_sub_row:
            first_cell_classes.append("table-first-column-child")
        if is_strong_row:
            first_cell_classes.append("table-row-strong")

        if is_header_row:
            merged_classes = first_cell_classes + ["table-header-merged", "table-data-header"]
            row_cells = [
                html.Td(
                    display_title,
                    className=" ".join(merged_classes),
                    colSpan=max(1, num_columns)
                )
            ]
        else:
            row_cells = [html.Td(display_title, className=" ".join(first_cell_classes))]

            for col_name in column_names[1:]:
                val = table_dict[col_name][i]
                formatted_val = f"{val:,.2f}" if pd.notna(val) and isinstance(val, (int, float)) else ("" if pd.isna(val) else val)
                cell_classes = ["text-end", "table-data-cell"]
                if is_strong_row:
                    cell_classes.append("table-row-strong")
                if is_sub_row:
                    cell_classes.append("table-data-sub")
                row_cells.append(html.Td(formatted_val, className=" ".join(cell_classes)))

        row_classes_for_render = list(row_classes)
        if not is_header_row and not is_sub_row:
            row_classes_for_render.append("table-main-row")
        body_rows.append(html.Tr(row_cells, className=" ".join(row_classes_for_render)))

    body = [html.Tbody(body_rows)]
    
    return [
        html.Div(
            [
                html.H5(title, className="mb-0"),
                html.Span(subtitle, className="text-muted small ms-2 table-subtitle"),
            ],
            className="table-title-row d-flex align-items-center gap-2 mt-4"
        ),
        dbc.Table(header + body, bordered=True, hover=True, responsive=True, className="table-sm")
    ]


def _flatten_component_text(value):
    """Recursively extract a readable string from dash component children."""
    if value is None:
        return ""
    if isinstance(value, (str, int, float)):
        return str(value)
    if isinstance(value, (list, tuple)):
        parts = [_flatten_component_text(v) for v in value]
        return " ".join([part for part in parts if part])
    return ""


def _extract_section_heading(section):
    """Derive title/subtitle text from the standard table or graph header components."""
    components = list(section) if isinstance(section, (list, tuple)) else [section]
    title_text = None
    subtitle_text = None

    def parse_header_div(div_component):
        nonlocal title_text, subtitle_text
        children = div_component.children
        child_list = children if isinstance(children, (list, tuple)) else [children]
        for child in child_list:
            if title_text is None and isinstance(child, html.H5):
                title_text = _flatten_component_text(child.children)
            elif subtitle_text is None and isinstance(child, html.Span):
                subtitle_text = _flatten_component_text(child.children)

    if components:
        first = components[0]
        if isinstance(first, html.Div):
            class_name = getattr(first, 'className', "") or ""
            if 'table-title-row' in class_name:
                parse_header_div(first)
            elif 'graph-section' in class_name:
                graph_children = first.children
                graph_child_list = graph_children if isinstance(graph_children, (list, tuple)) else [graph_children]
                if graph_child_list and isinstance(graph_child_list[0], html.Div):
                    header_div = graph_child_list[0]
                    header_class = getattr(header_div, 'className', "") or ""
                    if 'table-title-row' in header_class:
                        parse_header_div(header_div)

    if not title_text:
        title_text = "Section"

    return title_text, subtitle_text, components


def build_results_accordion(section_components, accordion_id):
    """Wrap a collection of sections in a Bootstrap accordion with the first item expanded."""
    items = []
    first_item_id = None

    for idx, section in enumerate(section_components):
        if not section:
            continue
        title_text, subtitle_text, body_components = _extract_section_heading(section)
        body_children = list(body_components) if isinstance(body_components, (list, tuple)) else [body_components]
        def strip_table_titles(component):
            if component is None:
                return None
            if isinstance(component, html.Div):
                class_name = getattr(component, "className", "") or ""
                if 'table-title-row' in class_name:
                    return None
            if hasattr(component, "children"):
                children = getattr(component, "children")
                if isinstance(children, (list, tuple)):
                    new_children = []
                    for child_component in children:
                        stripped_child = strip_table_titles(child_component)
                        if stripped_child is None:
                            continue
                        new_children.append(stripped_child)
                    component.children = new_children
                else:
                    stripped_single = strip_table_titles(children)
                    component.children = stripped_single
            return component

        filtered_children = []
        for child in body_children:
            stripped = strip_table_titles(child)
            if stripped is None:
                continue
            filtered_children.append(stripped)
        if not filtered_children:
            continue
        item_id = f"{accordion_id}-item-{idx}"

        header_children = [html.Span(title_text, className="accordion-item-title-text")]
        if subtitle_text:
            header_children.append(html.Span(subtitle_text, className="accordion-item-subtitle ms-2"))
        header_node = html.Span(
            header_children,
            className="accordion-item-title d-flex flex-column flex-md-row align-items-start align-items-md-center gap-1"
        )

        items.append(
            dbc.AccordionItem(
                filtered_children,
                title=header_node,
                item_id=item_id,
                className="results-accordion-item"
            )
        )
        if first_item_id is None:
            first_item_id = item_id

    if not items:
        return []

    return [
        dbc.Accordion(
            items,
            id=f"{accordion_id}-accordion",
            start_collapsed=False,
            active_item=first_item_id,
            className="results-accordion"
        )
    ]


def count_max_decimals(series) -> int:
    """Return the maximum number of decimal places present in a numeric series."""
    max_decimals = 0
    for val in series:
        if pd.isna(val):
            continue
        try:
            dec = Decimal(str(val)).normalize()
        except (InvalidOperation, ValueError):
            continue
        exponent = dec.as_tuple().exponent
        decimals = 0 if exponent >= 0 else -exponent
        if decimals > max_decimals:
            max_decimals = decimals
    return max_decimals


def derive_excel_number_format(template: str, decimals: int, signed: bool = False) -> str:
    """Build an Excel number format string while respecting the desired decimal precision."""
    base_template = template or "#,##0.00"
    segments = base_template.split(';')
    positive_template = segments[0] if segments else base_template
    include_thousands = ',' in positive_template
    include_plus = ('+' in positive_template) or signed
    if '.' in positive_template:
        decimals_in_template = len(positive_template.split('.', 1)[1])
    else:
        decimals_in_template = 0
    decimals_final = decimals_in_template
    if template is None and decimals_in_template == 0 and decimals:
        decimals_final = min(max(decimals, 0), 6)
    base = "#,##0" if include_thousands else "0"
    if decimals_final > 0:
        base += "." + ("0" * decimals_final)
    positive = f"+{base}" if include_plus else base
    negative = f"-{base}"
    zero = base
    return f"{positive};{negative};{zero}"


def format_signed_value(val):
    if pd.isna(val):
        return val
    if val > 0:
        return f"+{val:,.2f}"
    if val < 0:
        return f"{val:,.2f}"
    return "0.00"


def format_int_value(val):
    if pd.isna(val):
        return ""
    return f"{val:,.0f}"


def format_int_difference(val):
    if pd.isna(val):
        return ""
    if val > 0:
        return f"+{val:,.0f}"
    if val < 0:
        return f"{val:,.0f}"
    return "0"


def format_two_decimal_value(val):
    if pd.isna(val):
        return ""
    return f"{val:,.2f}"


def format_one_decimal_value(val):
    if pd.isna(val):
        return ""
    return f"{val:,.1f}"


def format_one_decimal_difference(val):
    if pd.isna(val):
        return ""
    if val > 0:
        return f"+{val:,.1f}"
    if val < 0:
        return f"{val:,.1f}"
    return "0.0"


def assign_current_hh_deciles(df: pd.DataFrame) -> pd.DataFrame:
    """Assign household deciles for the current scenario via EUROMOD CreateDeciles
    (grouped by idhh, head-only weights: WeightVar=hh_weight in the template)."""
    deciles, _ = em_create_deciles(df['eqRank'], df['dwt'] * (df['dhh'] == 1), df['idhh'])
    df['deciles_hh_current'] = deciles
    return df


def assign_current_individual_deciles(df: pd.DataFrame) -> pd.DataFrame:
    """Assign individual deciles for the current scenario via EUROMOD CreateDeciles
    (grouped by idhh, so household members share a decile)."""
    deciles, _ = em_create_deciles(df['eqRank'], df['dwt'], df['idhh'])
    df['deciles_ind_current'] = deciles
    return df

# Helper for building baseline parameter modal sections
def create_baseline_param_section(title, params_dict):
    """Creates an accordion item for the baseline parameters modal."""
    def format_value(value):
        if isinstance(value, (int, float)):
            if isinstance(value, float) and not value.is_integer():
                formatted = f"{value:,.2f}".rstrip('0').rstrip('.')
                return formatted if formatted else "0"
            return f"{int(round(value)):,}"
        return value

    rows = [
        dbc.Row([
            dbc.Col(html.Span(format_value(value), className="baseline-param-value"), width="auto"),
            dbc.Col(html.Strong(label, className="baseline-param-label"), width=True)
        ], className="align-items-center mb-1 baseline-param-row") for label, value in params_dict.items()
    ]

    return dbc.AccordionItem(
        rows,
        title=title,
        item_id=title,
        className="baseline-param-accordion-item",
    )

def create_policy_changes_section(title, rows, notes=None):
    """Builds policy change section content for the modal."""
    if notes is None:
        notes = []
    section_children = []
    for row in rows:
        value_classes = ["policy-change-value"]
        if row.get('changed'):
            value_classes.append("policy-change-value-changed")
        else:
            value_classes.append("policy-change-value-unchanged")
        display_value = row.get('display_value', '')
        label_text = row.get('label', '')
        section_children.append(
            dbc.Row([
                dbc.Col(
                    html.Span(display_value, className=" ".join(value_classes)),
                    width="auto",
                ),
                dbc.Col(html.Strong(label_text, className="policy-change-label"), width=True)
            ], className="align-items-center mb-1 policy-change-row")
        )

    for note in notes:
        note_classes = ["policy-change-note"]
        if note.get('changed'):
            note_classes.append("policy-change-note-changed")
        else:
            note_classes.append("policy-change-note-unchanged")
        note_text = note.get('text', '')
        section_children.append(
            dbc.Row([
                dbc.Col(html.Span(note_text, className=" ".join(note_classes)), width=12)
            ], className="mb-1 policy-change-note-row")
        )
    accordion_item = dbc.AccordionItem(
        section_children,
        title=title,
        item_id=title,
        className="policy-change-accordion-item",
    )
    return accordion_item

def collect_reform_params(param_ids, param_values, vat_checklist_value):
    """Builds the full reform parameter dict from the current UI inputs.
    Returns (reform_params, selected_vat_items, added_exemptions, removed_exemptions)."""
    reform_params = BASELINE_PARAMS.copy()
    for pid, raw_val in zip(param_ids or [], param_values or []):
        param_key = pid.get('index') if isinstance(pid, dict) else None
        if not param_key:
            continue
        parsed_val = parse_param_input_value(param_key, raw_val)
        if parsed_val is None:
            continue
        reform_params[param_key] = clamp_param_value(param_key, parsed_val)
    selected_vat_items = vat_checklist_value if vat_checklist_value is not None else BASELINE_VAT_STD_RATE_ITEMS
    selected_vat_items = list(selected_vat_items)
    reform_params['vat_items_list'] = selected_vat_items
    baseline_vat_set = set(BASELINE_VAT_STD_RATE_ITEMS)
    reform_vat_set = set(selected_vat_items)
    added_exemptions = sorted(baseline_vat_set - reform_vat_set)
    removed_exemptions = sorted(reform_vat_set - baseline_vat_set)
    return reform_params, selected_vat_items, added_exemptions, removed_exemptions


def reform_differs_from_baseline(reform_params, added_exemptions, removed_exemptions):
    """True when at least one policy parameter or VAT item differs from 2023."""
    if added_exemptions or removed_exemptions:
        return True
    return any(
        not policy_values_equal(param_id, BASELINE_PARAMS.get(param_id),
                                reform_params.get(param_id, BASELINE_PARAMS.get(param_id)))
        for param_id in POLICY_PARAM_LOOKUP
    )


def build_policy_changes_data(reform_params, reform_name, added_exemptions, removed_exemptions):
    """Builds the baseline-vs-reform parameter comparison payload for the
    policy changes modal."""
    policy_sections_payload = []
    for section in POLICY_PARAM_SECTIONS:
        section_rows = []
        for param_id, label in section['items']:
            baseline_val = BASELINE_PARAMS.get(param_id)
            reform_val = reform_params.get(param_id, baseline_val)
            baseline_display = format_policy_value(param_id, baseline_val)
            reform_display = format_policy_value(param_id, reform_val)
            changed = not policy_values_equal(param_id, baseline_val, reform_val)
            display_value = f"{baseline_display} -> {reform_display}" if changed else baseline_display
            section_rows.append({
                'param_id': param_id,
                'label': label,
                'baseline_display': baseline_display,
                'reform_display': reform_display,
                'display_value': display_value,
                'changed': changed,
            })
        section_payload = {'title': section['title'], 'rows': section_rows}
        if section['title'] == 'Value-added tax (VAT)':
            notes = []
            if not added_exemptions and not removed_exemptions:
                notes.append({'text': 'No changes in standard-rated/exempt items', 'changed': False})
            else:
                if added_exemptions:
                    count = len(added_exemptions)
                    noun = "exemption" if count == 1 else "exemptions"
                    notes.append({'text': f"{count} {noun} added", 'changed': True})
                if removed_exemptions:
                    count = len(removed_exemptions)
                    noun = "exemption" if count == 1 else "exemptions"
                    notes.append({'text': f"{count} {noun} removed", 'changed': True})
            section_payload['notes'] = notes
        policy_sections_payload.append(section_payload)

    policy_changes_data = {
        'scenario_name': reform_name or "Reform scenario",
        'sections': policy_sections_payload,
    }
    policy_changes_data['no_changes'] = not any(
        row['changed'] for section in policy_sections_payload for row in section.get('rows', [])
    ) and not any(
        note.get('changed') for section in policy_sections_payload for note in section.get('notes', [])
    )
    return policy_changes_data


def build_policy_changes_modal_body(policy_data):
    """Returns the body components for the policy changes modal."""
    sections = policy_data.get('sections', []) if policy_data else []
    no_changes = policy_data.get('no_changes') if policy_data else False

    accordion_items = []
    active_items = []
    if no_changes:
        return [html.H5("No policy changes", className="policy-no-changes-header mt-3 mb-2")]

    for section in sections:
        section_title = section.get('title', '')
        rows = section.get('rows', [])
        notes = section.get('notes', [])
        has_changes = any(row.get('changed') for row in rows) or any(note.get('changed') for note in notes)
        accordion_item = create_policy_changes_section(section_title, rows, notes)
        accordion_items.append(accordion_item)
        if has_changes:
            active_items.append(section_title)

    accordion = dbc.Accordion(
        accordion_items,
        always_open=True,
        active_item=active_items if active_items else [],
        className="policy-changes-accordion mt-2",
    )
    return [accordion]

# Result tabs: (label, key). Tab content lives in always-mounted panes below the
# nav (toggled client-side), so the info/policy-change buttons sit between the
# tab names and the content and never remount on tab switches.
RESULT_TAB_DEFS = [
    ("Tax-benefit policy", "taxbenpol"),
    ("Households", "households"),
    ("Individuals", "individuals"),
    ("Poverty", "poverty"),
    ("Poverty graphs", "poverty-graphs"),
    ("Inequality", "inequality"),
    ("Inequality graphs", "inequality-graphs"),
    ("Benefits", "benefits"),
    ("Taxes", "taxes"),
    ("Policy effects", "policy-effects"),
    ("Gainers & losers", "gainers-losers"),
]

# Tabs shown by default in the pilot interface. The rest stay mounted but their
# nav items are hidden until "Show additional tabs" is clicked.
CORE_TAB_KEYS = ['taxbenpol', 'poverty-graphs', 'inequality-graphs',
                 'policy-effects', 'gainers-losers']
EXTRA_TAB_KEYS = [key for _, key in RESULT_TAB_DEFS if key not in CORE_TAB_KEYS]
# Shorter labels while the graph tabs are the only poverty/inequality tabs shown
SIMPLE_TAB_LABELS = {'poverty-graphs': 'Poverty', 'inequality-graphs': 'Inequality'}

# Inactive panes stay laid out at full width (see the pane switcher below).
# The active pane leaves visibility unset so that it still inherits the hidden
# state dcc.Loading applies while a simulation is running.
PANE_STYLE_VISIBLE = {'display': 'block', 'position': 'relative',
                      'pointerEvents': 'auto'}
PANE_STYLE_HIDDEN = {'display': 'block', 'position': 'absolute', 'top': 0,
                     'left': 0, 'right': 0, 'visibility': 'hidden',
                     'pointerEvents': 'none', 'zIndex': -1}

RESULTS_TITLE_PLACEHOLDER = html.Span(
    "Your results will appear here", className="results-title-placeholder"
)


# Background text shown by the DEVMOD info hover card
DEVMOD_INFO_BODY = [
            html.P([
                "DEVMOD is a synthetic tax-benefit microsimulation model developed under UNU-WIDER’s ",
                html.A("SOUTHMOD project", href="https://www.wider.unu.edu/project/southmod-simulating-tax-and-benefit-policies-development-phase-3", target="_blank"),
                ". The model runs on the EUROMOD platform and mirrors real SOUTHMOD country models. It uses artificial data, so you can learn and experiment without handling sensitive micro data. The model is taught and used in the ",
                html.A("SOUTHMOD online course", href="https://www.wider.unu.edu/about/southmod-online-course", target="_blank"),
                ", delivered through UNU-WIDER’s learning platform at ",
                html.A("learning.wider.unu.edu", href="https://learning.wider.unu.edu/group/2", target="_blank"),
                "."
            ], style={"lineHeight": 1.5}),
        
            html.P("This simulator allows you to run DEVMOD on the web. The outputs correspond to what DEVMOD produces when run and analysed in EUROMOD. Based on the model’s synthetic input dataset, you can run a baseline policy system for 2023, change parameters to create reform scenarios, and compare baseline and reform indicators for various distributional and budgetary outcomes – similar to the SOUTHMOD Statistics Presenter in EUROMOD.", style={"lineHeight": 1.5}),
            html.P([
                "DEVMOD follows standard SOUTHMOD conventions for identifiers, income variables, and policy functions, and supports simulations of direct and indirect taxes, social contributions, and cash benefits. It is maintained by UNU-WIDER as an accompanion to the SOUTHMOD bundle. Refer to the ",
                html.A("SOUTHMOD User Manual", href="https://www.wider.unu.edu/sites/default/files/Projects/PDF/SOUTHMOD_UserManual_20250718.pdf", target="_blank"),
                " for details."
            ], style={"lineHeight": 1.5}),

            html.P([
                "To run DEVMOD in the standard EUROMOD environment, download the model from ",
                html.A("here (zip file)", href="https://www.wider.unu.edu/sites/default/files/About/DEVMOD%20v1.0.zip", target="_blank"),
                ", optionally review its data requirement document from ",
                html.A("here (Excel file)", href="https://www.wider.unu.edu/sites/default/files/About/DRD%20DEVMOD%20for%20dataset%20dv_2020_a1.xlsx", target="_blank"),
                ", and install EUROMOD software from  ",
                html.A("here (zip file)", href="https://euromod-web.jrc.ec.europa.eu/sites/default/files/EUROMOD_installer_64bit_latest_version.zip", target="_blank"),
                ". Finally open the DEVMOD model folder in EUROMOD, click on the DV flag, and edit or run the model as needed. For background on the modelling platform itself, see ",
                html.A("What is EUROMOD?", href="https://euromod-web.jrc.ec.europa.eu/overview/what-is-euromod", target="_blank"),
                 " by the Joint Research Centre, the European Commission’s science and knowledge service that develops and maintains EUROMOD.",
            ], style={"lineHeight": 1.5})

]


def build_hover_card(key, title, body_children, body_id=None):
    """Shell for the cards that appear in the results area on hover. All of the
    explanatory text in the tool uses this: the predefined reforms, the two
    numbered steps, the run button, DEVMOD background and the tab descriptions."""
    body_kwargs = {'id': body_id} if body_id else {}
    return html.Div(
        [
            html.Div(title, className="hovercard-title", id=f"hovercard-{key}-title"),
            html.Div(body_children, className="hovercard-body", **body_kwargs),
        ],
        className="hovercard",
        id=f"hovercard-{key}",
        hidden=True,
    )


def build_tab_info_bodies():
    """One pre-rendered description per results tab. They are all in the layout
    and switched client-side: filling this from a callback would put the whole
    results card into its loading state on every tab click."""
    bodies = []
    for _, key in RESULT_TAB_DEFS:
        content = INFO_MODAL_CONTENT.get(key, INFO_MODAL_CONTENT['default'])
        bodies.append(html.Div(
            dcc.Markdown(normalize_html_text(content.get('body', '')),
                         dangerously_allow_html=True, className="hovercard-markdown"),
            id=f'tabinfo-{key}',
            style={'display': 'block' if key == RESULT_TAB_DEFS[0][1] else 'none'},
        ))
    return bodies


TAB_INFO_TITLES = {
    key: INFO_MODAL_CONTENT.get(key, INFO_MODAL_CONTENT['default']).get('title', 'About this tab')
    for _, key in RESULT_TAB_DEFS
}


# The explanatory copy behind the numbered steps and the run button
STEP_1_HELP = [
    html.P("Build a reform by changing the policy parameters below. Open a section "
           "and edit any value: income tax thresholds and rates, social insurance "
           "contributions, the VAT rate and which goods are standard-rated, or the "
           "amounts and eligibility rules of each benefit.", className="hovercard-desc"),
    html.P("Every value that differs from the 2023 baseline is highlighted, so you can "
           "always see what your reform consists of. If you would rather start from a "
           "ready-made package, use one of the predefined reforms below the parameter "
           "sections, on their own or on top of your own edits.", className="hovercard-desc"),
]

STEP_2_HELP = [
    html.P("Chooses the measure of living standards used to rank people and to compute "
           "the poverty and inequality results.", className="hovercard-desc"),
    html.Div([
        html.Div([html.Span("Consumption based", className="hovercard-term"),
                  html.Span("what the household actually consumes, which is the more "
                            "reliable measure where much income is informal or "
                            "irregular.")], className="hovercard-deflist-row"),
        html.Div([html.Span("Income based", className="hovercard-term"),
                  html.Span("the household's disposable income after direct taxes and "
                            "cash benefits, plus the value of what it produces for its "
                            "own use.")], className="hovercard-deflist-row"),
        html.Div([html.Span("Net of indirect taxes", className="hovercard-term"),
                  html.Span("the same two measures with the VAT the household pays "
                            "taken off. These options also use the post-fiscal poverty "
                            "line, and are the ones to choose when your reform touches "
                            "VAT.")],
                 className="hovercard-deflist-row"),
    ], className="hovercard-deflist"),
    html.P("The choice shapes the poverty, inequality and distributional results; "
           "revenue and spending totals are the same either way. Household rankings "
           "always come from the baseline, so the two scenarios stay comparable.",
           className="hovercard-desc"),
]

RUN_HELP = [
    html.P("Runs the model twice on the same synthetic dataset: once with the 2023 "
           "baseline policies and once with your reform. Nothing you change affects "
           "the baseline, so the comparison is always like for like.",
           className="hovercard-desc"),
    html.P("The results tabs then show the baseline, the reform, and the difference "
           "between them for revenues and spending, poverty, inequality, and who gains "
           "and loses. A run takes a few seconds.", className="hovercard-desc"),
]


def build_preset_preview_card(preset_key):
    """Hover preview for a predefined reform: what it does and the parameters
    it changes (baseline -> reform). Built once, shown/hidden client-side."""
    preset = PRESET_REFORMS[preset_key]
    params = preset['params']
    body = []
    for section in POLICY_PARAM_SECTIONS:
        rows = []
        for param_id, label in section['items']:
            if param_id not in params:
                continue
            rows.append(html.Div([
                html.Span(label, className="preset-preview-label"),
                html.Span([
                    html.Span(format_policy_value(param_id, BASELINE_PARAMS.get(param_id)),
                              className="preset-preview-from"),
                    html.Span(" → ", className="preset-preview-arrow"),
                    html.Strong(format_policy_value(param_id, params[param_id])),
                ], className="preset-preview-values"),
            ], className="preset-preview-row"))
        if rows:
            body.append(html.Div(section['title'], className="preset-preview-section"))
            body.extend(rows)

    return build_hover_card(
        preset_key,
        preset['label'],
        [html.P(preset['description'], className="hovercard-desc"),
         html.Div(body, className="preset-preview-rows")],
    )

# --- APP LAYOUT ---
app.layout = dbc.Container([
    dcc.Download(id='download-simulation-output'),
    dcc.Store(id='policy-changes-data'),
    # Per-parameter formatting/limit metadata for the client-side input handler
    dcc.Store(id='param-meta-store', data={
        pid: dict(get_param_meta(pid), baseline=BASELINE_PARAMS.get(pid, 0))
        for pid in PARAM_INPUT_META
    }),
    # Baseline standard-rated VAT items, for the client-side change highlight
    dcc.Store(id='vat-baseline-store', data=BASELINE_VAT_STD_RATE_ITEMS),
    # Parameters each predefined reform writes, for the client-side preset toggle
    dcc.Store(id='preset-defs-store',
              data={key: val['params'] for key, val in PRESET_REFORMS.items()}),
    
    # Baseline Parameters Modal
    dbc.Modal([
        dbc.ModalHeader(dbc.ModalTitle("Baseline parameters (2023)")),
        dbc.ModalBody(
            dbc.Accordion(
                [
                    create_baseline_param_section("Personal income tax", {
                        'Self-employment income threshold (presumptive maximum), annual': BASELINE_PARAMS['pit_yse_turnover_threshold'],
                        'Exemption on agricultural income, annual': BASELINE_PARAMS['pit_yag_exemption'],
                        'Bracket 2 (non-zero tax) lower threshold, annual': BASELINE_PARAMS['pit_bracket2_thresh'],
                        'Bracket 2 progressive rate, %/100': BASELINE_PARAMS['pit_bracket2_rate'],
                        'Bracket 3 lower threshold, annual': BASELINE_PARAMS['pit_bracket3_thresh'],
                        'Bracket 3 progressive rate, %/100': BASELINE_PARAMS['pit_bracket3_rate'],
                        'Bracket 4 lower threshold, annual': BASELINE_PARAMS['pit_bracket4_thresh'],
                        'Bracket 4 progressive rate, %/100': BASELINE_PARAMS['pit_bracket4_rate'],
                        'Bracket 5 lower threshold, annual': BASELINE_PARAMS['pit_bracket5_thresh'],
                        'Bracket 5 progressive rate, %/100': BASELINE_PARAMS['pit_bracket5_rate'],
                    }),
                    create_baseline_param_section("Social insurance contributions", {
                        'Employee contribution rate, %/100': BASELINE_PARAMS['tscee_rate'],
                        'Employer contribution rate, %/100': BASELINE_PARAMS['tscer_rate'],
                    }),
                    create_baseline_param_section("Presumptive tax for micro enterprises", {
                        'Band 2 (non-zero tax) lower threshold, annual': BASELINE_PARAMS['presumptive_turnover_1'],
                        'Band 2 tax amount, annual': BASELINE_PARAMS['presumptive_tax_2'],
                        'Band 3 lower threshold, annual': BASELINE_PARAMS['presumptive_turnover_2'],
                        'Band 3 tax amount, annual': BASELINE_PARAMS['presumptive_tax_3'],
                    }),
                    create_baseline_param_section("Presumptive tax for small enterprises", {
                        'Lower threshold, annual': BASELINE_PARAMS['presumptive_turnover_3'],
                        'Tax rate, %/100': BASELINE_PARAMS['presumptive_rate_4'],
                    }),
                    create_baseline_param_section("Value-added tax (VAT)", {
                        'Standard VAT rate, %/100': BASELINE_PARAMS['tva_rate'],
                    }),
                    create_baseline_param_section("Social assistance benefit", {
                        'Income threshold, monthly': BASELINE_PARAMS['bsa_income_threshold'],
                        'Benefit amount (1-person household), monthly': BASELINE_PARAMS['bsa_1_person'],
                        'Benefit amount (2-person household), monthly': BASELINE_PARAMS['bsa_2_person'],
                        'Benefit amount (3+-person household), monthly': BASELINE_PARAMS['bsa_3_plus_person'],
                        'Disability top-up, monthly': BASELINE_PARAMS['bsa_disabled_topup'],
                    }),
                    create_baseline_param_section("Senior citizens' grant", {
                        'Eligibility age threshold': BASELINE_PARAMS['senior_grant_age'],
                        'Eligibility income threshold, monthly': BASELINE_PARAMS['senior_grant_income_threshold'],
                        'Grant amount, monthly': BASELINE_PARAMS['senior_grant_amount'],
                    }),
                    create_baseline_param_section("School meals (in-kind)", {
                        'School meal age threshold': BASELINE_PARAMS['school_meal_age'],
                        'School meal value, monthly': BASELINE_PARAMS['school_meal_value'],
                    }),
                    create_baseline_param_section("Poverty lines", {
                        'Basic poverty line, monthly': BASELINE_PARAMS['basic_pov_line'],
                        'Basic post-fiscal poverty line, monthly': BASELINE_PARAMS['basic_pov_line_pf'],
                    }),
                ],
                always_open=True,
                active_item=[],
                className="baseline-param-accordion",
            )
        ),
        dbc.ModalFooter(dbc.Button("Close", id="close-baseline-modal", className="ms-auto", n_clicks=0)),
    ], id="baseline-modal", is_open=False, size="lg", scrollable=True, className="baseline-modal"), # Made modal scrollable
    dbc.Modal([
        dbc.ModalHeader(dbc.ModalTitle(id="policy-changes-modal-title")),
        dbc.ModalBody(html.Div(id="policy-changes-modal-body"), className="policy-changes-modal-body"),
        dbc.ModalFooter(dbc.Button("Close", id="close-policy-changes-modal", className="ms-auto"))
    ], id="policy-changes-modal", is_open=False, size="lg", scrollable=True, className="baseline-modal"),

    # Primary layout: controls column and results column
    dbc.Row([
        dbc.Col(
            html.Div([
                html.A([
                    html.Img(src=app.get_asset_url("unu-wider-logo.png"),
                             alt="UNU-WIDER", className="app-title-logo"),
                    html.Div([
                        html.H1("SOUTHMOD Online Tool", className="app-title-heading"),
                        html.P("Pilot interface", className="app-title-subheading"),
                    ], className="app-title-text"),
                ], className="app-title", href="https://www.wider.unu.edu/southmod",
                    target="_blank", title="UNU-WIDER SOUTHMOD"),

                dbc.Card([
                    dbc.CardBody([
                                # DEVMOD mode parameter configuration
                                html.Div(id='devmod-params-section', children=[
                        make_control_step("1", "Configure reform parameters", element_id="step-1-header"),
                        dbc.Accordion(id='devmod-accordion', children=[
                            dbc.AccordionItem([
                                html.P("Personal income tax", className="accordion-section-title"),
                                make_param_input("Self-employment income threshold", 'pit_yse_turnover_threshold', BASELINE_PARAMS['pit_yse_turnover_threshold'],
                                                 tip="Also the turnover ceiling for presumptive tax. Annual value"), 
                                make_param_input("Exemption on agricultural income", 'pit_yag_exemption', BASELINE_PARAMS['pit_yag_exemption'], tip=ANNUAL_TIP), 
                                html.Hr(),
                                make_pit_table(BASELINE_PARAMS),
                                html.Hr(), 
                                html.P("Social insurance contributions", className="accordion-section-title"),
                                make_param_input("Employee SIC rate", 'tscee_rate', BASELINE_PARAMS['tscee_rate'], 0.01, tip=RATE_TIP), 
                                make_param_input("Employer SIC rate", 'tscer_rate', BASELINE_PARAMS['tscer_rate'], 0.01, tip=RATE_TIP),
                                html.Hr(),
                                html.P("Presumptive tax for micro enterprises", className="accordion-section-title"),
                                make_param_input("Band 2 lower threshold", 'presumptive_turnover_1', BASELINE_PARAMS['presumptive_turnover_1'],
                                                 tip="The first band with a non-zero tax. Annual value"), 
                                make_param_input("Band 2 tax amount", 'presumptive_tax_2', BASELINE_PARAMS['presumptive_tax_2'], tip=ANNUAL_TIP), 
                                make_param_input("Band 3 lower threshold", 'presumptive_turnover_2', BASELINE_PARAMS['presumptive_turnover_2'], tip=ANNUAL_TIP), 
                                make_param_input("Band 3 tax amount", 'presumptive_tax_3', BASELINE_PARAMS['presumptive_tax_3'], tip=ANNUAL_TIP), 
                                html.Hr(),
                                html.P("Presumptive tax for small enterprises", className="accordion-section-title"),
                                make_param_input("Lower threshold", 'presumptive_turnover_3', BASELINE_PARAMS['presumptive_turnover_3'], tip=ANNUAL_TIP), 
                        make_param_input("Tax rate", 'presumptive_rate_4', BASELINE_PARAMS['presumptive_rate_4'], 0.01, tip=RATE_TIP)
                            ], title="Direct taxes"),
                            dbc.AccordionItem([
                        html.P("Value-added tax (VAT)", className="accordion-section-title"),
                        make_param_input("Standard VAT rate", 'tva_rate', BASELINE_PARAMS['tva_rate'], 0.01, tip=RATE_TIP),
                        html.Hr(),
                                dbc.Label("Select standard-rated goods", className="fw-bold", style={"font-size": "0.92rem"}),
                                dbc.Row([
                                    dbc.Col(dbc.Button("Select all", id="vat-select-all", color="link", size="sm", className="p-0 me-2"), width="auto"),
                                    dbc.Col(dbc.Button("Exempt all", id="vat-deselect-all", color="link", size="sm", className="p-0 me-2"), width="auto"),
                                    dbc.Col(dbc.Button("Back to baseline", id="vat-baseline", color="link", size="sm", className="p-0"), width="auto"),
                                ], className="mb-2 vat-button-row align-items-center"),
                                html.Div(
                                    dcc.Checklist(
                                        id='vat-checklist',
                                        options=[{'label': v['label'], 'value': k} for k, v in VAT_ITEM_MAP.items()],
                                        value=BASELINE_VAT_STD_RATE_ITEMS,
                                        className="dbc_checklist",
                                        style={'height': '180px', 'overflowY': 'auto', 'padding': '8px'},
                                        labelStyle={'display': 'block', 'font-size': '0.9rem', 'marginLeft': '0.4rem'}
                                    ),
                                    id='vat-checklist-wrapper',
                                    className="vat-checklist-wrapper",
                                )
                            ], title="Indirect taxes"),
                            dbc.AccordionItem([
                                html.P("Social assistance", className="accordion-section-title"),
                                make_param_input("Eligibility income threshold", 'bsa_income_threshold', BASELINE_PARAMS['bsa_income_threshold'], tip=MONTHLY_TIP), 
                                make_param_input("1-person benefit amount", 'bsa_1_person', BASELINE_PARAMS['bsa_1_person'], tip=MONTHLY_TIP), 
                                make_param_input("2-person benefit amount", 'bsa_2_person', BASELINE_PARAMS['bsa_2_person'], tip=MONTHLY_TIP), 
                                make_param_input("3+ person benefit amount", 'bsa_3_plus_person', BASELINE_PARAMS['bsa_3_plus_person'], tip=MONTHLY_TIP), 
                                make_param_input("Disability top-up", 'bsa_disabled_topup', BASELINE_PARAMS['bsa_disabled_topup'], tip=MONTHLY_TIP), 
                                html.Hr(), 
                                html.P("Senior citizens' grant", className="accordion-section-title"),
                                make_param_input("Eligibility age threshold", 'senior_grant_age', BASELINE_PARAMS['senior_grant_age'], tip=AGE_TIP), 
                                make_param_input("Eligibility income threshold", 'senior_grant_income_threshold', BASELINE_PARAMS['senior_grant_income_threshold'], tip=MONTHLY_TIP), 
                                make_param_input("Senior grant amount", 'senior_grant_amount', BASELINE_PARAMS['senior_grant_amount'], tip=MONTHLY_TIP), 
                                html.Hr(), 
                                html.P("School meals (in-kind)", className="accordion-section-title"),
                                make_param_input("School meal age threshold", 'school_meal_age', BASELINE_PARAMS['school_meal_age'], tip=AGE_TIP),
                                make_param_input("School meal value", 'school_meal_value', BASELINE_PARAMS['school_meal_value'], tip=MONTHLY_TIP)
                            ], title="Benefit policies"),
                        ], start_collapsed=True, className="mb-3"),
                        html.P("Predefined reforms", className="preset-row-heading"),
                        html.Div([
                            dbc.Button(PRESET_REFORMS['tax']['label'], id='preset-tax-button', color="secondary", outline=True, size="sm", className="preset-btn"),
                            dbc.Button(PRESET_REFORMS['benefits']['label'], id='preset-benefits-button', color="secondary", outline=True, size="sm", className="preset-btn"),
                            dbc.Button("Reset all", id='preset-reset-button', color="secondary", outline=True, size="sm", className="preset-btn preset-btn-reset"),
                        ], className="preset-button-row d-flex gap-2"),
                        dcc.Store(id='preset-state-store', data={'tax': False, 'benefits': False}),
                        ])
                    ], className="modern-card-body")
                ], className="modern-card shadow-sm border-0 control-card"),

                # Shared controls
                dbc.Card([
                    dbc.CardBody([
                        html.Div([
                            make_control_step("2", "Select distribution statistic", element_id="step-2-header"),
                            dbc.Select(
                                id='analysis-choice',
                                options=[
                                    {'label': 'Consumption based', 'value': '1'},
                                    {'label': 'Income based', 'value': '2'},
                                    {'label': 'Consumption based, net of indirect taxes', 'value': '3'},
                                    {'label': 'Income based, net of indirect taxes', 'value': '4'}
                                ],
                                value='3',
                                className="modern-select"
                            ),
                        ], className="mb-2"),

                        dbc.Button("Run simulation", id='run-button', color="primary", size="lg", className="w-100 btn-run-simulation"),
                        html.Div(
                            dcc.Loading(id="loading-icon", children=[html.Div(id="loading-output", className="text-center")], type="default"),
                            className="run-status-zone"
                        ),
                        dbc.Row([
                            dbc.Col(dbc.Button("DEVMOD info", id="view-devmod-button", color="secondary", outline=True, size="sm", className="w-100 btn-baseline btn-baseline-sub"), width=6),
                            dbc.Col(dbc.Button("Baseline parameters", id="view-baseline-button", color="secondary", outline=True, size="sm", className="w-100 btn-baseline btn-baseline-sub"), width=6),
                        ], className="g-2 info-button-row"),
                    ], className="modern-card-body")
                ], className="modern-card shadow-sm border-0 run-card"),
            ], className="left-column d-flex flex-column gap-3"),
            xs=12,
            lg=4,
            className="layout-left-col"
        ),

        dbc.Col(
            dbc.Card(
                dbc.CardBody([
                    dcc.Loading(
                        id="loading-main",
                        children=[
                            html.H4(id='results-title', children=RESULTS_TITLE_PLACEHOLDER, className="mt-1 mb-3 text-center results-title"),
                            html.Div(
                                [
                                    # Floated, so it reserves room on the first
                                    # row of tabs only
                                    dbc.Button(
                                        "Show additional indicators",
                                        id='show-extra-tabs-button',
                                        color="link",
                                        size="sm",
                                        className="btn-extra-tabs",
                                    ),
                                    dbc.Tabs(
                                        id='results-tabs',
                                        active_tab='taxbenpol',
                                        children=[
                                            dbc.Tab(
                                                id=f'tabnav-{key}',
                                                label=SIMPLE_TAB_LABELS.get(key, label),
                                                tab_id=key,
                                                className="modern-tab",
                                                tab_style=None if key in CORE_TAB_KEYS else {'display': 'none'},
                                            )
                                            for label, key in RESULT_TAB_DEFS
                                        ],
                                        className="modern-tabs"
                                    ),
                                ],
                                className="results-tabs-row"
                            ),
                            html.Div(
                                [
                                    dbc.Button(
                                        "Description of tab's indicators",
                                        id='tab-info-button',
                                        color="secondary",
                                        outline=True,
                                        size="sm",
                                        className="info-button btn-description",
                                    ),
                                    dbc.Button(
                                        "Policy changes",
                                        id='tab-policy-changes-button',
                                        color="secondary",
                                        outline=True,
                                        size="sm",
                                        className="info-button btn-description btn-policy-changes",
                                    ),
                                    dbc.Button(
                                        "Clear results & reset parameters",
                                        id='clear-results-button',
                                        color="secondary",
                                        outline=True,
                                        size="sm",
                                        className="info-button btn-description btn-clear-results",
                                        style={'display': 'none'},
                                    ),
                                ],
                                className="tab-info-wrapper d-flex align-items-center gap-2"
                            ),
                            html.Div(
                                [
                                    html.Div(
                                        [
                                            html.Div(
                                                html.Div(id=f'tab-{key}'),
                                                id=f'tabpane-{key}',
                                                style=(PANE_STYLE_VISIBLE if key == 'taxbenpol'
                                                       else PANE_STYLE_HIDDEN),
                                            )
                                            for _, key in RESULT_TAB_DEFS
                                        ],
                                        className="tab-panes"
                                    ),
                                    html.Div(
                                        [build_preset_preview_card(key) for key in PRESET_REFORMS] + [
                                            build_hover_card("step1", "Configure reform parameters", STEP_1_HELP),
                                            build_hover_card("step2", "Select distribution statistic", STEP_2_HELP),
                                            build_hover_card("run", "Run simulation", RUN_HELP),
                                            build_hover_card("devmod", "About DEVMOD", DEVMOD_INFO_BODY),
                                            build_hover_card("tabinfo", "About this tab",
                                                             build_tab_info_bodies()),
                                        ],
                                        className="hovercard-container",
                                    ),
                                ],
                                id='results-content-wrapper',
                                className="results-content-wrapper results-empty",
                            ),
                        ],
                        type="default",
                        className="results-loading"
                    )
                ], className="modern-card-body"),
                className="modern-card shadow-sm border-0 results-card"
            ),
            xs=12,
            lg=8,
            className="layout-right-col modern-results-panel"
        )
    ], className="layout-row g-4 align-items-start"),
    
], fluid=True, className="app-shell py-4")

# --- CALLBACKS ---

def warm_baseline_cache():
    try:
        df = ensure_input_dataframe()
    except Exception:
        return

    for choice in (1, 2, 3, 4):
        try:
            get_baseline_artifacts(df, choice)
        except Exception:
            continue


warm_baseline_cache()

# Baseline parameters modal callback
@app.callback(
    Output("baseline-modal", "is_open"),
    [Input("view-baseline-button", "n_clicks"), Input("close-baseline-modal", "n_clicks")],
    [State("baseline-modal", "is_open")],
)
def toggle_baseline_modal(n_view, n_close, is_open):
    if n_view or n_close:
        return not is_open
    return is_open

# Callback for VAT checklist buttons
@app.callback(
    Output('vat-checklist', 'value'),
    Input('vat-select-all', 'n_clicks'),
    Input('vat-deselect-all', 'n_clicks'),
    Input('vat-baseline', 'n_clicks'),
    prevent_initial_call=True
)
def update_vat_checklist(select_all, deselect_all, baseline_click):
    ctx = dash.callback_context
    if not ctx.triggered:
        return dash.no_update
    
    trigger_id = ctx.triggered[0]['prop_id'].split('.')[0]
    
    if trigger_id == 'vat-select-all':
        return list(VAT_ITEM_MAP.keys())
    elif trigger_id == 'vat-deselect-all':
        return []
    elif trigger_id == 'vat-baseline':
        return BASELINE_VAT_STD_RATE_ITEMS
    
    return dash.no_update

# Client-side parameter input handling (formatting, stepping, hard limits) so
# +/- clicks and typing respond instantly without a server round trip.
app.clientside_callback(
    """
    function(currentValue, decClicks, incClicks, componentId, metaStore) {
        const nu = window.dash_clientside.no_update;
        const paramId = componentId && componentId.index;
        if (!paramId) return nu;
        const meta = Object.assign(
            {precision: 2, thousands: false, allow_negative: false, strip_trailing: true,
             force_int: false, step: 1, min: null, max: null, baseline: 0},
            (metaStore || {})[paramId] || {}
        );
        const ctx = window.dash_clientside.callback_context;
        if (!ctx || !ctx.triggered || !ctx.triggered.length) return nu;
        const trigProp = ctx.triggered[0].prop_id;

        const clamp = (v) => {
            if (v === null || v === undefined || isNaN(v)) return v;
            if (meta.min !== null && meta.min !== undefined && v < meta.min) v = meta.min;
            if (meta.max !== null && meta.max !== undefined && v > meta.max) v = meta.max;
            if (meta.force_int) v = Math.round(v);
            return v;
        };
        const parse = (val) => {
            if (val === null || val === undefined) return null;
            if (typeof val === 'number') return isNaN(val) ? null : val;
            if (typeof val !== 'string') return null;
            const cleaned = val.replace(/,/g, '').trim();
            if (cleaned === '' || cleaned === '.' || cleaned === '-') return null;
            const p = Number(cleaned);
            if (isNaN(p)) return null;
            return meta.force_int ? Math.round(p) : p;
        };
        const format = (v) => {
            if (v === null || v === undefined || v === '') return '';
            const num = (typeof v === 'number') ? v : Number(String(v).replace(/,/g, '').trim());
            if (isNaN(num)) return String(v);
            const precision = (meta.precision === null || meta.precision === undefined) ? 0 : meta.precision;
            let formatted = num.toLocaleString('en-US', {
                minimumFractionDigits: precision,
                maximumFractionDigits: precision,
                useGrouping: !!meta.thousands,
            });
            if (meta.strip_trailing && formatted.indexOf('.') !== -1) {
                formatted = formatted.replace(/0+$/, '').replace(/\.$/, '');
            }
            return formatted;
        };

        // Stepper buttons
        if (trigProp.indexOf('param-step') !== -1) {
            let direction = null;
            if (trigProp.indexOf('"direction":"inc"') !== -1) direction = 'inc';
            else if (trigProp.indexOf('"direction":"dec"') !== -1) direction = 'dec';
            if (!direction) return nu;
            let parsed = parse(currentValue);
            if (parsed === null) parsed = meta.baseline || 0;
            let step = meta.step || 1;
            const magnitude = Math.abs(parsed);
            if (magnitude >= 1000 && step >= 1) step = 100;
            else if (magnitude >= 100 && step >= 1) step = 10;
            let newValue = (direction === 'inc') ? parsed + step : parsed - step;
            if (!meta.allow_negative && newValue < 0) newValue = 0;
            newValue = clamp(newValue);
            return format(newValue);
        }

        // Typed input
        if (typeof currentValue === 'string') {
            const stripped = currentValue.trim();
            if (stripped === '') return '';
            const normalized = stripped.replace(/,/g, '');
            const fullNumeric = /^-?\d*(\.\d*)?$/.test(normalized);
            const typed = Number(normalized);
            // Hard limits while typing: negatives and over-max clamp immediately;
            // positive values below a nonzero minimum are left for run time
            if (fullNumeric && normalized !== '' && !isNaN(typed)) {
                if (typed < 0 && meta.min !== null && meta.min !== undefined) return format(meta.min);
                if (meta.max !== null && meta.max !== undefined && typed > meta.max) return format(meta.max);
            }
            if (stripped.endsWith('.')) return stripped;
            if (fullNumeric) {
                if (['-', '-0', '+', '+0'].indexOf(normalized) !== -1) return stripped;
                if (normalized.indexOf('.') !== -1) {
                    const decimals = normalized.split('.')[1];
                    const precision = (meta.precision === null || meta.precision === undefined) ? 2 : meta.precision;
                    if (decimals.length <= precision) return stripped;
                }
            }
        }
        if (currentValue === null || currentValue === '') return '';
        let parsed = parse(currentValue);
        if (parsed === null) {
            const cleaned = String(currentValue).split('').filter(
                (ch) => '0123456789.,-'.indexOf(ch) !== -1
            ).join('');
            if (cleaned === String(currentValue)) return nu;
            return cleaned;
        }
        // Typed values: clamp negatives and over-max only; positive values
        // below a nonzero minimum stay as typed (clamped when the simulation
        // runs) so multi-digit entry is not disturbed
        if (parsed < 0) {
            parsed = (meta.min !== null && meta.min !== undefined) ? meta.min : 0;
        }
        if (meta.max !== null && meta.max !== undefined && parsed > meta.max) parsed = meta.max;
        if (meta.force_int) parsed = Math.round(parsed);
        const formatted = format(parsed);
        if (formatted === String(currentValue)) return nu;
        return formatted;
    }
    """,
    Output({'type': 'param-input', 'index': MATCH}, 'value'),
    Input({'type': 'param-input', 'index': MATCH}, 'value'),
    Input({'type': 'param-step', 'index': MATCH, 'direction': 'dec'}, 'n_clicks'),
    Input({'type': 'param-step', 'index': MATCH, 'direction': 'inc'}, 'n_clicks'),
    State({'type': 'param-input', 'index': MATCH}, 'id'),
    State('param-meta-store', 'data'),
    prevent_initial_call=True,
)

# Client-side highlight of parameters that differ from their baseline value
app.clientside_callback(
    """
    function(paramValues, paramIds, metaStore) {
        const base = 'form-control form-control-sm param-input-field';
        const meta = metaStore || {};
        return (paramIds || []).map(function (cid, i) {
            const pid = cid && cid.index;
            if (!pid || !(pid in meta)) return base;
            const raw = paramValues[i];
            let num = null;
            if (typeof raw === 'number') num = raw;
            else if (typeof raw === 'string') {
                const cleaned = raw.replace(/,/g, '').trim();
                if (cleaned !== '' && cleaned !== '.' && cleaned !== '-') {
                    const parsed = Number(cleaned);
                    if (!isNaN(parsed)) num = parsed;
                }
            }
            if (num === null) return base;
            return Math.abs(num - meta[pid].baseline) > 1e-9 ? base + ' param-changed' : base;
        });
    }
    """,
    Output({'type': 'param-input', 'index': ALL}, 'className'),
    Input({'type': 'param-input', 'index': ALL}, 'value'),
    State({'type': 'param-input', 'index': ALL}, 'id'),
    State('param-meta-store', 'data'),
)

# Same highlight for the VAT item list when the standard-rated set has changed
app.clientside_callback(
    """
    function(selected, baseline) {
        const cur = new Set(selected || []);
        const base = new Set(baseline || []);
        let same = cur.size === base.size;
        if (same) { for (const item of cur) { if (!base.has(item)) { same = false; break; } } }
        return same ? 'vat-checklist-wrapper' : 'vat-checklist-wrapper vat-changed';
    }
    """,
    Output('vat-checklist-wrapper', 'className'),
    Input('vat-checklist', 'value'),
    State('vat-baseline-store', 'data'),
)

# Client-side tab pane switching. Inactive panes are taken out of the flow but
# keep their width, so charts are laid out at their final size on mount instead
# of resizing the first time their tab is shown.
app.clientside_callback(
    """
    function(active_tab) {
        const keys = ["taxbenpol", "households", "individuals", "poverty", "poverty-graphs", "inequality", "inequality-graphs", "benefits", "taxes", "policy-effects", "gainers-losers"];
        return keys.map(function (k) {
            return k === active_tab ? PANE_VISIBLE : PANE_HIDDEN;
        });
    }
    """.replace('PANE_VISIBLE', json.dumps(PANE_STYLE_VISIBLE))
       .replace('PANE_HIDDEN', json.dumps(PANE_STYLE_HIDDEN)),
    [Output(f'tabpane-{key}', 'style') for _, key in RESULT_TAB_DEFS],
    Input('results-tabs', 'active_tab'),
)

# Client-side toggle between the five demo tabs and the full set. Expanding also
# restores the '... graphs' labels so the two poverty/inequality tabs stay apart.
app.clientside_callback(
    """
    function(nClicks, activeTab) {
        const extras = __EXTRA_KEYS__;
        const expanded = !!nClicks && (nClicks % 2 === 1);
        const styles = extras.map(function () { return expanded ? {} : {display: 'none'}; });
        let active = window.dash_clientside.no_update;
        if (!expanded && extras.indexOf(activeTab) !== -1) active = 'taxbenpol';
        return styles.concat([
            expanded ? 'Poverty graphs' : 'Poverty',
            expanded ? 'Inequality graphs' : 'Inequality',
            expanded ? 'Show fewer indicators' : 'Show additional indicators',
            active,
        ]);
    }
    """.replace('__EXTRA_KEYS__', json.dumps(EXTRA_TAB_KEYS)),
    [Output(f'tabnav-{key}', 'tab_style') for key in EXTRA_TAB_KEYS] + [
        Output('tabnav-poverty-graphs', 'label'),
        Output('tabnav-inequality-graphs', 'label'),
        Output('show-extra-tabs-button', 'children'),
        Output('results-tabs', 'active_tab', allow_duplicate=True),
    ],
    Input('show-extra-tabs-button', 'n_clicks'),
    State('results-tabs', 'active_tab'),
    prevent_initial_call=True,
)

# The tab description hover card follows the active tab, client-side: a server
# callback here would show the results card's loading spinner on every tab click
app.clientside_callback(
    """
    function(activeTab) {
        const keys = __KEYS__;
        const titles = __TITLES__;
        return keys.map(function (k) {
            return {display: k === activeTab ? 'block' : 'none'};
        }).concat([titles[activeTab] || 'About this tab']);
    }
    """.replace('__KEYS__', json.dumps([key for _, key in RESULT_TAB_DEFS]))
       .replace('__TITLES__', json.dumps(TAB_INFO_TITLES)),
    [Output(f'tabinfo-{key}', 'style') for _, key in RESULT_TAB_DEFS] +
    [Output('hovercard-tabinfo-title', 'children')],
    Input('results-tabs', 'active_tab'),
)


# Main simulation and results callback
@app.callback(
    [Output(f'tab-{tab_name}', 'children') for tab_name in 
     ['taxbenpol', 'households', 'individuals', 'poverty', 'poverty-graphs', 
      'inequality', 'inequality-graphs', 'benefits', 'taxes', 
      'policy-effects', 'gainers-losers']],
    Output('loading-output', 'children'),
    Output('results-title', 'children'),  # Title for the results card
    Output('policy-changes-data', 'data'),
    Output('download-simulation-output', 'data'),  # Excel download payload
    Input('run-button', 'n_clicks'),
    State('analysis-choice', 'value'),
    State({'type': 'param-input', 'index': ALL}, 'id'),
    State({'type': 'param-input', 'index': ALL}, 'value'),
    State('vat-checklist', 'value')  # Selected VAT checklist values
)
def run_and_display_results(n_clicks, analysis_choice,
                            param_ids, param_values, vat_checklist_value):
    # The pilot interface has no scenario-name field and no Excel switch; the
    # export path below stays in place for the full build.
    reform_name = None
    generate_excel = False

    dev_placeholder = html.Div(dbc.Alert("Output for this tab is under development.", color="info"), className="p-4")
    run_placeholder = html.Div(dbc.Alert("Run a simulation to see results.", color="info"), className="p-4")
    if not n_clicks:
        initial_tabs = [
            run_placeholder,  # Tax-benefit policy
            run_placeholder,  # Households
            run_placeholder,  # Individuals
            run_placeholder,  # Poverty
            run_placeholder,  # Poverty graphs
            run_placeholder,  # Inequality
            run_placeholder,  # Inequality graphs
            run_placeholder,  # Benefits
            run_placeholder,  # Taxes
            run_placeholder,  # Policy effects
            run_placeholder,  # Gainers & losers
        ]
        return initial_tabs + ["", RESULTS_TITLE_PLACEHOLDER, None, dash.no_update]

    # A run with no parameter changes would compare the baseline against itself,
    # so say what is missing instead of spending ten seconds on it
    reform_params, _, added_exemptions, removed_exemptions = collect_reform_params(
        param_ids, param_values, vat_checklist_value
    )
    if not reform_differs_from_baseline(reform_params, added_exemptions, removed_exemptions):
        no_change_notice = html.Div(
            dbc.Alert(
                "Change at least one parameter to model a reform, then run a "
                "simulation to see results.",
                color="warning", className="no-change-alert",
            ),
            className="p-4",
        )
        return [no_change_notice] * 11 + ["", RESULTS_TITLE_PLACEHOLDER, None, dash.no_update]

    try:
        analysis_choice = int(analysis_choice)
    except (TypeError, ValueError):
        analysis_choice = 3

    distribution_labels = {
        1: "Consumption based",
        2: "Income based",
        3: "Consumption based, net of indirect taxes",
        4: "Income based, net of indirect taxes",
    }
    distribution_label = distribution_labels.get(analysis_choice, "Consumption based")

    try:
        try:
            df = ensure_input_dataframe()
        except FileNotFoundError:
            error_msg = dbc.Alert(
                f"Error: Input file '{INPUT_FILE}' not found in the application folder.",
                color="danger"
            )
            return [error_msg] * 11 + ["Error", dash.no_update, None, dash.no_update]
        except Exception as e:
            error_msg = dbc.Alert(f"Error loading '{INPUT_FILE}': {e}", color="danger")
            return [error_msg] * 11 + ["Error", dash.no_update, None, dash.no_update]

        baseline_artifacts = get_baseline_artifacts(df, analysis_choice)
        baseline_results = baseline_artifacts['results']
        baseline_analysis_df = baseline_artifacts['merge_df']

        # Build reform parameters
        reform_params, selected_vat_items, added_exemptions, removed_exemptions = collect_reform_params(
            param_ids, param_values, vat_checklist_value
        )

        reform_sim_df = run_simulation(df, reform_params)
        reform_results, reform_analysis_df = run_analysis(reform_sim_df, analysis_choice, baseline_analysis_df)

    except Exception as e:
        import traceback
        print(f"Error during simulation: {e}")
        traceback.print_exc()
        error_msg = dbc.Alert([
            html.H5("An error occurred during simulation:", className="alert-heading"),
            html.P(f"Error: {e}"),
            html.P("Please check your input data. Common issues include missing 'dag' column or 0 weights.")
        ], color="danger")
        return [error_msg] * 11 + ["Simulation failed.", dash.no_update, None, dash.no_update]

    policy_changes_data = build_policy_changes_data(
        reform_params, reform_name, added_exemptions, removed_exemptions
    )

    # --- Generate TaxBenPol Tab Content ---
    tbp_rows = ['Sum of government revenue', 'By source', '- Direct taxes', '- Social insurance contributions', '- Indirect taxes',
                'Sum of government expenditure', 'By type', '- Cash benefits', '- In-kind benefits', '- Indirect subsidies']
    
    abs_data = {'Component': tbp_rows, 'Baseline': [], 'Reform': []}
    for row in tbp_rows:
        key = row.replace('- ','')
        if key in baseline_results['taxbenpol_abs']:
            # Multiply monthly totals by 12 to get yearly totals (in millions)
            abs_data['Baseline'].append(baseline_results['taxbenpol_abs'][key] * 12 / 1e6)
            abs_data['Reform'].append(reform_results['taxbenpol_abs'][key] * 12 / 1e6)
        else:
            abs_data['Baseline'].append(None) # Use None for blank rows
            abs_data['Reform'].append(None)

    abs_df = pd.DataFrame(abs_data)
    if 'Component' in abs_df.columns:
        abs_df = abs_df.rename(columns={'Component': ''})
    abs_df['Difference'] = abs_df.apply(lambda row: row['Reform'] - row['Baseline'] if pd.notna(row['Reform']) and pd.notna(row['Baseline']) else None, axis=1)
    abs_df_excel = abs_df.copy()
    abs_df['Difference'] = abs_df['Difference'].apply(format_signed_value)
    tab1_part1 = create_styled_table(abs_df.to_dict('list'), "Total revenue and expenditure", "(yearly, millions of national currency)")

    share_rows = ['By source', '- Direct taxes', '- Social insurance contributions', '- Indirect taxes',
                  'By type', '- Cash benefits', '- In-kind benefits', '- Indirect subsidies']
    share_data = {'Component': share_rows, 'Baseline (%)': [], 'Reform (%)': []}
    for row in share_rows:
        key = row.replace('- ','')
        if key in baseline_results['taxbenpol_share']:
            share_data['Baseline (%)'].append(baseline_results['taxbenpol_share'][key])
            share_data['Reform (%)'].append(reform_results['taxbenpol_share'][key])
        else:
            share_data['Baseline (%)'].append(None)
            share_data['Reform (%)'].append(None)
            
    share_df = pd.DataFrame(share_data)
    if 'Component' in share_df.columns:
        share_df = share_df.rename(columns={'Component': ''})
    share_df['Difference (pp.)'] = share_df.apply(lambda row: row['Reform (%)'] - row['Baseline (%)'] if pd.notna(row['Reform (%)']) and pd.notna(row['Baseline (%)']) else None, axis=1)
    share_df_excel = share_df.copy()
    share_df['Difference (pp.)'] = share_df['Difference (pp.)'].apply(format_signed_value)
    tab1_part2 = create_styled_table(share_df.to_dict('list'), "Shares of total revenue and expenditure", "(%)")
    tab1_content = build_results_accordion([tab1_part1, tab1_part2], "taxbenpol")

    # --- Generate Poverty Tab Content ---
    pov_row_map = {
        'All individuals': 'All individuals',
        'header_hh_structure': 'Household structure',
        'isSinglePersonHH': '- Single person',
        'is1AdultWithChildrenHH': '- Single parent',
        'is2AdultsNoChildrenHH': '- 2 adults without children',
        'is2Adults1_2ChildrenHH': '- 2 adults with 1-2 children',
        'is2Adults3_4ChildrenHH': '- 2 adults with 3-4 children',
        'is2Adults5plusChildrenHH': '- 2 adults with 5+ children',
        'is3plusAdultsNoChildrenHH': '- 3+ adults without children',
        'is3plusAdultsWithChildrenHH': '- 3+ adults with children',
        'header_vulnerable': 'Vulnerable households',
        'isYoungChildHH': '- HH with young child (0-2)',
        'isAtLeastOneElderlyHH': '- HH with elderly member',
        'isAtLeastOneDisabledHH': '- HH with disabled member',
        'isNoMaleAdultHH': '- HH with no male adults',
        'header_labor': 'Labour market status',
        'isNoTotalHHEarningsHH': '- HH with no labour income',
        'isInformalAdultHH': '- HH with informal adult(s)',
        'isNoInformalAdultsHH': '- HH with no informal adults'
    }
    
    pov_rate_data = {'Household category': [], 'Baseline (%)': [], 'Reform (%)': []}
    pov_gap_data = {'Household category': [], 'Baseline (%)': [], 'Reform (%)': []}
    baseline_poverty = baseline_results['poverty']
    reform_poverty = reform_results['poverty']

    def get_yearly_povline(source):
        if not isinstance(source, dict):
            return None
        value = source.get('povline')
        if value is None or pd.isna(value):
            return None
        try:
            return float(value) * 12
        except (TypeError, ValueError):
            return None
    
    for k, v in pov_row_map.items():
        pov_rate_data['Household category'].append(v)
        pov_gap_data['Household category'].append(v)
        
        # This is a data row
        if not k.startswith('header_'):
            pov_rate_data['Baseline (%)'].append(baseline_poverty[k]['Poverty rate (%)'])
            pov_rate_data['Reform (%)'].append(reform_poverty[k]['Poverty rate (%)'])
            pov_gap_data['Baseline (%)'].append(baseline_poverty[k]['Poverty gap (%)'])
            pov_gap_data['Reform (%)'].append(reform_poverty[k]['Poverty gap (%)'])
        # This is a header row
        else:
            for data_dict in [pov_rate_data, pov_gap_data]:
                data_dict['Baseline (%)'].append(None) # Use None for blanks
                data_dict['Reform (%)'].append(None)

    povline_label = "Absolute national poverty line, yearly"
    baseline_povline_yearly = get_yearly_povline(baseline_poverty)
    reform_povline_yearly = get_yearly_povline(reform_poverty)
    if reform_povline_yearly is None:
        reform_povline_yearly = baseline_povline_yearly
    for data_dict in [pov_rate_data, pov_gap_data]:
        data_dict['Household category'].append(povline_label)
        data_dict['Baseline (%)'].append(baseline_povline_yearly)
        data_dict['Reform (%)'].append(reform_povline_yearly)
    
    pov_rate_df = pd.DataFrame(pov_rate_data)
    pov_gap_df = pd.DataFrame(pov_gap_data)
    if 'Household category' in pov_rate_df.columns:
        pov_rate_df = pov_rate_df.rename(columns={'Household category': ''})
    if 'Household category' in pov_gap_df.columns:
        pov_gap_df = pov_gap_df.rename(columns={'Household category': ''})

    def calc_diff(row):
        if pd.notna(row['Reform (%)']) and pd.notna(row['Baseline (%)']):
            return row['Reform (%)'] - row['Baseline (%)']
        return None
        
    pov_rate_df['Difference (pp.)'] = pov_rate_df.apply(calc_diff, axis=1)
    pov_gap_df['Difference (pp.)'] = pov_gap_df.apply(calc_diff, axis=1)
    pov_rate_df_excel = pov_rate_df.copy()
    pov_gap_df_excel = pov_gap_df.copy()
    for df_excel in (pov_rate_df_excel, pov_gap_df_excel):
        for col in ['Baseline (%)', 'Reform (%)', 'Difference (pp.)']:
            if col in df_excel.columns:
                df_excel[col] = df_excel[col].apply(lambda x: round(x, 2) if pd.notna(x) else np.nan)
                df_excel[col] = df_excel[col].astype(float)

    for df in (pov_rate_df, pov_gap_df):
        first_col = df.columns[0] if len(df.columns) > 0 else None
        diff_col = 'Difference (pp.)'
        if first_col and diff_col in df.columns:
            mask = df[first_col] == povline_label
            if mask.any():
                df.loc[mask, diff_col] = None

    pov_rate_df['Difference (pp.)'] = pov_rate_df['Difference (pp.)'].apply(format_signed_value)
    pov_gap_df['Difference (pp.)'] = pov_gap_df['Difference (pp.)'].apply(format_signed_value)

    poverty_rate_section = create_styled_table(pov_rate_df.to_dict('list'), "Poverty rate", "(share of poor population, %)")
    poverty_gap_section = create_styled_table(pov_gap_df.to_dict('list'), "Poverty gap", "(average normalised poverty gap, %)")
    tab2_content = build_results_accordion([poverty_rate_section, poverty_gap_section], "poverty")

    poverty_graphs_content = [run_placeholder]
    poverty_graphs_excel_figures = []

    household_structure_groups = [
        ('All individuals', 'All individuals'),
        ('Single person', 'isSinglePersonHH'),
        ('Single parent', 'is1AdultWithChildrenHH'),
        ('2 adults, no children', 'is2AdultsNoChildrenHH'),
        ('2 adults, 1-2 children', 'is2Adults1_2ChildrenHH'),
        ('2 adults, 3-4 children', 'is2Adults3_4ChildrenHH'),
        ('2 adults, 5+ children', 'is2Adults5plusChildrenHH'),
        ('3+ adults, no children', 'is3plusAdultsNoChildrenHH'),
        ('3+ adults, with children', 'is3plusAdultsWithChildrenHH'),
    ]

    vulnerability_groups = [
        ('All individuals', 'All individuals'),
        ('Young child (aged 0-2)', 'isYoungChildHH'),
        ('Elderly member', 'isAtLeastOneElderlyHH'),
        ('Member with a disability', 'isAtLeastOneDisabledHH'),
        ('No male adults', 'isNoMaleAdultHH'),
        ('No labour market income', 'isNoTotalHHEarningsHH'),
        ('Informal adult', 'isInformalAdultHH'),
        ('No informal adults', 'isNoInformalAdultsHH'),
    ]

    structure_palette = ['#1d4ed8', '#2563eb', '#38bdf8', '#0ea5e9', '#14b8a6', '#10b981', '#f59e0b', '#ef4444', '#9333ea']
    vulnerability_palette = ['#1d4ed8', '#0ea5e9', '#9333ea', '#f97316', '#ef4444', '#2563eb', '#22c55e', '#14b8a6']
    structure_color_map = {
        label: structure_palette[idx % len(structure_palette)]
        for idx, (label, _) in enumerate(household_structure_groups)
    }
    vulnerability_color_map = {
        label: vulnerability_palette[idx % len(vulnerability_palette)]
        for idx, (label, _) in enumerate(vulnerability_groups)
    }

    def safe_metric(subgroup_results, metric_name):
        if not isinstance(subgroup_results, dict):
            return None
        value = subgroup_results.get(metric_name)
        if value is None or pd.isna(value):
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    def compute_records(group_specs, metric_name):
        records = []
        for display_label, subgroup_key in group_specs:
            baseline_entry = baseline_poverty.get(subgroup_key, {})
            reform_entry = reform_poverty.get(subgroup_key, {})
            baseline_val = safe_metric(baseline_entry, metric_name)
            reform_val = safe_metric(reform_entry, metric_name)
            difference = None
            if baseline_val is not None and reform_val is not None:
                difference = reform_val - baseline_val
            records.append({
                'label': display_label,
                'baseline': baseline_val,
                'reform': reform_val,
                'difference': difference,
            })
        return records

    def format_hover(val, decimals=2):
        if val is None or pd.isna(val):
            return "n/a"
        if decimals is None:
            formatted = format(val, ",.15g")
            return formatted
        fmt = f"{{:.{decimals}f}}" if decimals > 0 else "{:.0f}"
        formatted = fmt.format(val)
        if decimals > 0 and "." in formatted:
            formatted = formatted.rstrip("0").rstrip(".")
        return formatted

    def build_difference_figure(records, yaxis_title, color_map):
        filtered = [
            rec for rec in records
            if rec.get('difference') is not None and not pd.isna(rec.get('difference'))
        ]
        if not filtered:
            return None
        labels = [rec['label'] for rec in filtered]
        raw_differences = [float(rec['difference']) for rec in filtered]
        axis_settings = compute_axis_settings(raw_differences)
        decimals = axis_settings['decimals']
        tickformat = axis_settings['tickformat']
        value_decimals = max(decimals, 2)
        differences = []
        hovertext = []
        for rec, diff_val in zip(filtered, raw_differences):
            diff_display = format_hover(diff_val, 2)
            rec['difference'] = diff_val
            differences.append(diff_val)
            hovertext.append(
                f"<b>{rec['label']}</b><br>"
                f"Baseline: {format_hover(rec['baseline'], value_decimals)}%<br>"
                f"Reform: {format_hover(rec['reform'], value_decimals)}%<br>"
                f"Difference: {diff_display} pp"
            )
        colors = [color_map[label] for label in labels]
        fig = go.Figure(
            data=[
                go.Bar(
                    x=labels,
                    y=differences,
                    marker_color=colors,
                    marker_opacity=0.85,
                    hovertext=hovertext,
                    hovertemplate="%{hovertext}<extra></extra>",
                )
            ]
        )
        fig.add_hline(y=0, line_color="#94a3b8", line_width=1)
        fig.update_traces(marker_line_color="#ffffff", marker_line_width=0.7)
        fig.update_layout(
            margin=dict(l=60, r=20, t=20, b=100),
            plot_bgcolor="#ffffff",
            paper_bgcolor="#ffffff",
            bargap=0.45,
            showlegend=False,
            font=dict(size=12, color="#1f2a44"),
            hoverlabel=dict(bgcolor="#f8fafc"),
        )
        axis_kwargs = {
            "title": yaxis_title,
            "zeroline": False,
            "gridcolor": "#e5e7eb",
            "tickformat": tickformat,
        }
        if axis_settings.get("dtick"):
            axis_kwargs["dtick"] = axis_settings["dtick"]
        if axis_settings.get("range"):
            axis_kwargs["range"] = axis_settings["range"]
        fig.update_yaxes(**axis_kwargs)
        
        # Check if all differences are negative or zero
        all_negative_or_zero = all(d <= 0 for d in differences)
        # If all negative/zero, rotate labels to read top-left to bottom-right and move axis to top
        tick_angle = 35 if all_negative_or_zero else -35
        
        xaxis_kwargs = {
            "tickangle": tick_angle,
            "tickfont": dict(size=11),
            "ticks": 'outside',
            "ticklen": 12,
            "tickson": 'labels',
            "categoryorder": 'array',
            "categoryarray": labels,
        }
        if all_negative_or_zero:
            xaxis_kwargs["side"] = "top"
        
        fig.update_xaxes(**xaxis_kwargs)
        return fig

    def build_graph_section(title, subtitle, figure):
        title, subtitle = normalize_title_subtitle(title, subtitle)
        header_children = [html.H5(title, className="mb-0")]
        if subtitle:
            header_children.append(html.Span(subtitle, className="text-muted small ms-2"))
        return html.Div(
            [
                html.Div(
                    header_children,
                    className="table-title-row d-flex align-items-center gap-2 mt-4"
                ),
                dcc.Graph(
                    figure=figure,
                    config={'displayModeBar': False, 'responsive': True},
                    style={'height': '380px', 'width': '100%'},
                    className="graph-section-figure flex-grow-1 w-100"
                )
            ],
            className="graph-section h-100 d-flex flex-column w-100"
        )

    structure_rate_records = compute_records(household_structure_groups, 'Poverty rate (%)')
    vulnerability_rate_records = compute_records(vulnerability_groups, 'Poverty rate (%)')
    structure_gap_records = compute_records(household_structure_groups, 'Poverty gap (%)')
    vulnerability_gap_records = compute_records(vulnerability_groups, 'Poverty gap (%)')

    graph_definitions = [
        {
            'title': "Change in poverty rate by household structure (pp.)",
            'yaxis_title': "Difference in poverty rate (pp.)",
            'records': structure_rate_records,
            'color': '#2563eb',
        },
        {
            'title': "Change in poverty rate by vulnerability and labour market status (pp.)",
            'yaxis_title': "Difference in poverty rate (pp.)",
            'records': vulnerability_rate_records,
            'color': '#2563eb',
        },
        {
            'title': "Change in poverty gap by household structure (pp.)",
            'yaxis_title': "Difference in poverty gap (pp.)",
            'records': structure_gap_records,
            'color': '#2563eb',
        },
        {
            'title': "Change in poverty gap by vulnerability and labour market status (pp.)",
            'yaxis_title': "Difference in poverty gap (pp.)",
            'records': vulnerability_gap_records,
            'color': '#2563eb',
        },
    ]

    graph_sections_for_excel = []
    for graph in graph_definitions:
        color_value = graph.get('color', '#2563eb')
        color_map = {rec['label']: color_value for rec in graph['records']}
        fig = build_difference_figure(graph['records'], graph['yaxis_title'], color_map)
        if fig is None:
            continue
        fig.update_xaxes(ticks='outside', ticklen=12, tickson='labels')
        poverty_graphs_excel_figures.append(
            {
                'title': graph['title'],
                'subtitle': graph['yaxis_title'],
                'figure': fig,
            }
        )
        graph_sections_for_excel.append(build_graph_section(graph['title'], None, fig))

    if graph_sections_for_excel:
        poverty_graphs_content = build_results_accordion(graph_sections_for_excel, "poverty-graphs")
    else:
        poverty_graphs_content = [dbc.Alert("No poverty graph data available for this simulation.", color="warning")]

    # --- Generate Households Tab Content ---
    household_baseline = baseline_results.get('households', {})
    household_reform = reform_results.get('households', {})

    def build_population_table(row_specs, baseline_dict, reform_dict):
        data = {'Component': [], 'Baseline': [], 'Reform': []}
        for label, result_key in row_specs:
            data['Component'].append(label)
            if result_key:
                base_val = baseline_dict.get(result_key)
                reform_val = reform_dict.get(result_key)
            else:
                base_val = None
                reform_val = None
            data['Baseline'].append(base_val)
            data['Reform'].append(reform_val)
        df = pd.DataFrame(data)
        if 'Component' in df.columns:
            df = df.rename(columns={'Component': ''})
        df['Difference'] = df.apply(
            lambda row: row['Reform'] - row['Baseline'] if pd.notna(row['Reform']) and pd.notna(row['Baseline']) else None,
            axis=1
        )
        df_excel = df.copy()
        for col in ['Baseline', 'Reform', 'Difference']:
            if col in df_excel.columns:
                df_excel[col] = df_excel[col].apply(lambda x: round(x) if pd.notna(x) else x)
        df['Baseline'] = df['Baseline'].apply(format_int_value)
        df['Reform'] = df['Reform'].apply(format_int_value)
        df['Difference'] = df['Difference'].apply(format_int_difference)
        return df, df_excel

    table1_specs = [
        ('Total households', 'TotalHHCount'),
        ('Any taxes or contributions', 'CountHHpaysAnyTaxOrCont'),
        ('By source', None),
        ('- Direct taxes', 'CountHHpaysDirTax'),
        ('- Social insurance contributions', 'CountHHpaysSSC'),
        ('- Indirect taxes', 'CountHHpaysIndirTax'),
        ('Any benefits', 'CountHHgetsAnyBenefit'),
        ('By type', None),
        ('- Cash benefits', 'CountHHgetsCashBen'),
        ('- In-kind benefits', 'CountHHgetsInKindBen'),
        ('- Indirect subsidies', 'CountHHgetsIndirSub'),
    ]
    households_table1_df, households_table1_df_excel = build_population_table(table1_specs, household_baseline, household_reform)
    table1_display = create_styled_table(
        households_table1_df.to_dict('list'),
        "Taxpayer and benefit recipient households",
        "(number of households)"
    )

    table2_specs = [
        ('Total households', 'TotalHHCount'),
        ('Household structure', None),
        ('- Single person', 'CountHH_New_SinglePerson'),
        ('- Single parent', 'CountHH_New_1AdultWithChildren'),
        ('- 2 adults without children', 'CountHH_New_2AdultsNoChildren'),
        ('- 2 adults with 1-2 children', 'CountHH_New_2Adults1_2Children'),
        ('- 2 adults with 3-4 children', 'CountHH_New_2Adults3_4Children'),
        ('- 2 adults with 5+ children', 'CountHH_New_2Adults5plusChildren'),
        ('- 3+ adults without children', 'CountHH_New_3plusAdultsNoChildren'),
        ('- 3+ adults with children', 'CountHH_New_3plusAdultsWithChildren'),
        ('Vulnerable households', None),
        ('- Young child (aged 0-2)', 'CountHH_New_YoungChild'),
        ('- Elderly member', 'CountHH_New_ElderlyMember'),
        ('- Member with a disability', 'CountHH_New_DisabledMember'),
        ('- No male adults', 'CountHH_New_NoMaleAdult'),
        ('Labour market status', None),
        ('- No labour market income', 'CountHH_New_NoLaborIncome'),
        ('- Informal adult', 'CountHH_New_InformalAdult'),
        ('- No informal adults', 'CountHH_New_NoInformalAdults'),
    ]
    households_table2_df, households_table2_df_excel = build_population_table(table2_specs, household_baseline, household_reform)
    table2_display = create_styled_table(
        households_table2_df.to_dict('list'),
        "Household categories",
        "(number of households)"
    )

    table3_specs = [('Total households', 'TotalHHCount')] + [
        (f'- Decile {i}', f'CountHHDecile{i}') for i in range(1, 11)
    ]
    households_table3_df, households_table3_df_excel = build_population_table(table3_specs, household_baseline, household_reform)
    table3_display = create_styled_table(
        households_table3_df.to_dict('list'),
        "Household decile distribution",
        "(number of households)"
    )
    tab3_content = build_results_accordion([table1_display, table2_display, table3_display], "households")

    # --- Generate Individuals Tab Content ---
    individual_baseline = baseline_results.get('individuals', {})
    individual_reform = reform_results.get('individuals', {})

    individual_table1_specs = [
        ('Total individuals', 'TotalIndCount'),
        ('Any taxes or contributions', 'CountIndPaysAnyTaxOrCont'),
        ('By source', None),
        ('- Direct taxes', 'CountIndPaysDirTax'),
        ('- Social insurance contributions', 'CountIndPaysSSC'),
        ('- Indirect taxes', 'CountIndPaysIndirTax'),
        ('Any benefits', 'CountIndGetsAnyBenefit'),
        ('By type', None),
        ('- Cash benefits', 'CountIndGetsCashBen'),
        ('- In-kind benefits', 'CountIndGetsInKindBen'),
        ('- Indirect subsidies', 'CountIndGetsIndirSub'),
    ]
    individuals_table1_df, individuals_table1_df_excel = build_population_table(individual_table1_specs, individual_baseline, individual_reform)
    ind_table1_display = create_styled_table(
        individuals_table1_df.to_dict('list'),
        "Taxpayer and benefit recipient individuals",
        "(number of individuals)"
    )

    individual_table2_specs = [
        ('Total individuals', 'TotalIndCount'),
        ('Household structure', None),
        ('- Single person', 'CountInd_New_SinglePerson'),
        ('- Single parent', 'CountInd_New_1AdultWithChildren'),
        ('- 2 adults without children', 'CountInd_New_2AdultsNoChildren'),
        ('- 2 adults with 1-2 children', 'CountInd_New_2Adults1_2Children'),
        ('- 2 adults with 3-4 children', 'CountInd_New_2Adults3_4Children'),
        ('- 2 adults with 5+ children', 'CountInd_New_2Adults5plusChildren'),
        ('- 3+ adults without children', 'CountInd_New_3plusAdultsNoChildren'),
        ('- 3+ adults with children', 'CountInd_New_3plusAdultsWithChildren'),
        ('Vulnerable households', None),
        ('- Young child (aged 0-2)', 'CountInd_New_YoungChild'),
        ('- Elderly member', 'CountInd_New_ElderlyMember'),
        ('- Member with a disability', 'CountInd_New_DisabledMember'),
        ('- No male adults', 'CountInd_New_NoMaleAdult'),
        ('Labour market status', None),
        ('- No labour market income', 'CountInd_New_NoLaborIncome'),
        ('- Informal adult', 'CountInd_New_InformalAdult'),
        ('- No informal adults', 'CountInd_New_NoInformalAdults'),
    ]
    individuals_table2_df, individuals_table2_df_excel = build_population_table(individual_table2_specs, individual_baseline, individual_reform)
    ind_table2_display = create_styled_table(
        individuals_table2_df.to_dict('list'),
        "Individual categories",
        "(number of individuals)"
    )

    individual_table3_specs = [('Total individuals', 'TotalIndCount')] + [
        (f'- Decile {i}', f'CountIndDecile{i}') for i in range(1, 11)
    ]
    individuals_table3_df, individuals_table3_df_excel = build_population_table(individual_table3_specs, individual_baseline, individual_reform)
    ind_table3_display = create_styled_table(
        individuals_table3_df.to_dict('list'),
        "Individual decile distribution",
        "(number of individuals)"
    )
    tab4_content = build_results_accordion([ind_table1_display, ind_table2_display, ind_table3_display], "individuals")

    inequality_baseline = baseline_results.get('inequality', {})
    inequality_reform = reform_results.get('inequality', {})

    def get_percentile_value(result_dict, percentile):
        if not result_dict:
            return 0
        return (result_dict.get('Percentiles') or {}).get(percentile, 0)

    def safe_ratio(numerator, denominator):
        if denominator and denominator != 0:
            return numerator / denominator
        return 0

    def compute_p80_p20(result_dict):
        return safe_ratio(get_percentile_value(result_dict, 80), get_percentile_value(result_dict, 20))

    def compute_mean_median_ratio(result_dict):
        median = get_percentile_value(result_dict, 50)
        mean_val = result_dict.get('MeanEqRank', 0) if result_dict else 0
        return safe_ratio(mean_val, median)

    def compute_decile_share(result_dict, decile):
        if not result_dict:
            return 0
        total = result_dict.get('TotalEqRank', 0)
        if total == 0:
            return 0
        return result_dict.get(f'SumEqRank_InBaselineDec{decile}', 0) / total * 100

    inequality_sections = []

    def begin_section(title):
        inequality_sections.append({'title': title, 'rows': [], 'formats': []})

    def add_data_row(label, baseline_val, reform_val, fmt):
        diff_val = None
        if fmt != 'header':
            try:
                if baseline_val is not None and reform_val is not None and not (pd.isna(baseline_val) or pd.isna(reform_val)):
                    diff_val = float(reform_val) - float(baseline_val)
            except (TypeError, ValueError):
                diff_val = None
        inequality_sections[-1]['rows'].append({
            'Metric': label,
            'Baseline': baseline_val,
            'Reform': reform_val,
            'Difference': diff_val,
        })
        inequality_sections[-1]['formats'].append(fmt)

    def add_header_row(label):
        inequality_sections[-1]['rows'].append({'Metric': label, 'Baseline': None, 'Reform': None, 'Difference': None})
        inequality_sections[-1]['formats'].append('header')

    begin_section("Inequality indices")
    add_data_row("- Gini coefficient", inequality_baseline.get('Gini', 0), inequality_reform.get('Gini', 0), 'two_dec')
    add_data_row("- Atkinson inequality index (ineq. aversion = 0.25)", inequality_baseline.get('Atkinson', 0), inequality_reform.get('Atkinson', 0), 'two_dec')
    add_data_row("- P80/P20 ratio", compute_p80_p20(inequality_baseline), compute_p80_p20(inequality_reform), 'two_dec')
    add_data_row("- Mean/median ratio", compute_mean_median_ratio(inequality_baseline), compute_mean_median_ratio(inequality_reform), 'two_dec')

    begin_section("Percentiles of distribution and median, yearly")
    for perc in [10, 20, 30, 40, 50, 60, 70, 80, 90]:
        label = f"- {perc}th"
        if perc == 50:
            label = "- 50th (median)"
        baseline_val = get_percentile_value(inequality_baseline, perc) * 12
        reform_val = get_percentile_value(inequality_reform, perc) * 12
        add_data_row(label, baseline_val, reform_val, 'int')
    add_data_row("Absolute national poverty line, yearly", inequality_baseline.get('povLine', 0) * 12, inequality_reform.get('povLine', 0) * 12, 'int')

    begin_section("Distribution of total income/consumption across baseline deciles, %")
    for decile in range(1, 11):
        label = f"- Decile {decile}"
        baseline_share = compute_decile_share(inequality_baseline, decile)
        reform_share = compute_decile_share(inequality_reform, decile)
        add_data_row(label, baseline_share, reform_share, 'one_dec')
    add_data_row("Total", 100.0, 100.0, 'one_dec')

    inequality_table_sections = []
    inequality_excel_tables = []

    for section in inequality_sections:
        df = pd.DataFrame(section['rows'])
        if 'Metric' in df.columns:
            df = df.rename(columns={'Metric': ''})
        
        # Rename Difference column for distribution section
        if section['title'] == "Distribution of total income/consumption across baseline deciles, %":
            df = df.rename(columns={'Difference': 'Difference (pp.)'})
        
        df_excel = df.copy()
        df_display = df.copy().astype('object')
        for idx, fmt in enumerate(section['formats']):
            if fmt == 'header':
                for col in ['Baseline', 'Reform', 'Difference', 'Difference (pp.)']:
                    if col in df_display.columns:
                        df_display.at[idx, col] = None
                        df_excel.at[idx, col] = np.nan
                continue
            if fmt == 'int':
                for col in ['Baseline', 'Reform']:
                    if col not in df_display.columns:
                        continue
                    raw_val = df_excel.at[idx, col]
                    df_display.at[idx, col] = format_int_value(raw_val)
                diff_col = 'Difference (pp.)' if 'Difference (pp.)' in df_display.columns else 'Difference'
                if diff_col in df_display.columns:
                    diff_val = section['rows'][idx].get('Difference')
                    df_display.at[idx, diff_col] = format_int_difference(diff_val)
            elif fmt == 'one_dec':
                for col in ['Baseline', 'Reform']:
                    if col not in df_display.columns:
                        continue
                    raw_val = df_excel.at[idx, col]
                    df_display.at[idx, col] = format_one_decimal_value(raw_val)
                diff_col = 'Difference (pp.)' if 'Difference (pp.)' in df_display.columns else 'Difference'
                if diff_col in df_display.columns:
                    diff_val = section['rows'][idx].get('Difference')
                    df_display.at[idx, diff_col] = format_one_decimal_difference(diff_val)
            elif fmt == 'two_dec':
                for col in ['Baseline', 'Reform']:
                    if col not in df_display.columns:
                        continue
                    raw_val = df_excel.at[idx, col]
                    df_display.at[idx, col] = format_two_decimal_value(raw_val)
                diff_col = 'Difference (pp.)' if 'Difference (pp.)' in df_display.columns else 'Difference'
                if diff_col in df_display.columns:
                    diff_val = section['rows'][idx].get('Difference')
                    df_display.at[idx, diff_col] = format_signed_value(diff_val) if diff_val is not None else ""

        inequality_table_sections.append(create_styled_table(df_display.to_dict('list'), section['title'], ""))
        
        # Determine which column name to use for Excel
        diff_col_excel = 'Difference (pp.)' if section['title'] == "Distribution of total income/consumption across baseline deciles, %" else 'Difference'
        
        # Use thousands separator for sections with integer values (like percentiles)
        if section['title'] == "Percentiles of distribution and median, yearly":
            inequality_excel_tables.append((df_excel, section['title'], [diff_col_excel], "+#,##0;-#,##0;0", "#,##0"))
        else:
            inequality_excel_tables.append((df_excel, section['title'], [diff_col_excel], "+0.00;-0.00;0.00", "#,##0.00"))

    inequality_table_content = build_results_accordion(inequality_table_sections, "inequality")

    inequality_graphs_content = [run_placeholder]
    inequality_graphs_excel_figures = []

    percentile_palette = ['#1d4ed8', '#2563eb', '#38bdf8', '#0ea5e9', '#14b8a6', '#10b981', '#f59e0b', '#ef4444', '#9333ea']
    decile_palette = ['#1d4ed8', '#2563eb', '#38bdf8', '#0ea5e9', '#14b8a6', '#10b981', '#facc15', '#f97316', '#ef4444', '#9333ea']

    def safe_difference(reform_val, baseline_val, multiplier=1.0, precision=6):
        if reform_val is None or baseline_val is None:
            return None
        if pd.isna(reform_val) or pd.isna(baseline_val):
            return None
        try:
            diff = (float(reform_val) - float(baseline_val)) * multiplier
            if precision is not None:
                return round(diff, precision)
            return diff
        except (TypeError, ValueError):
            return None

    def safe_share(numerator, denominator):
        if numerator is None or denominator is None:
            return None
        if pd.isna(numerator) or pd.isna(denominator):
            return None
        denominator = float(denominator)
        if denominator == 0:
            return None
        try:
            value = float(numerator) / denominator * 100
            return round(value, 6)
        except (TypeError, ValueError):
            return None

    percentile_records = []
    if inequality_baseline and inequality_reform:
        for idx, perc in enumerate(range(10, 100, 10)):
            base_val = get_percentile_value(inequality_baseline, perc)
            reform_val = get_percentile_value(inequality_reform, perc)
            diff_val = safe_difference(reform_val, base_val, multiplier=12)
            percentile_records.append({
                'label': f"{perc}th",
                'baseline': base_val * 12 if base_val is not None else None,
                'reform': reform_val * 12 if reform_val is not None else None,
                'difference': diff_val,
                'color': percentile_palette[idx % len(percentile_palette)],
            })

    inequality_decile_records = []
    benefits_decile_records = []
    dir_tax_decile_records = []
    indir_tax_decile_records = []

    if inequality_baseline and inequality_reform:
        inequality_decile_records = []
        total_baseline = inequality_baseline.get('TotalEqRank')
        total_reform = inequality_reform.get('TotalEqRank')
        for decile in range(1, 11):
            label = f"Decile {decile}"
            baseline_share = safe_share(inequality_baseline.get(f"SumEqRank_InBaselineDec{decile}"), total_baseline)
            reform_share = safe_share(inequality_reform.get(f"SumEqRank_InBaselineDec{decile}"), total_reform)
            diff = safe_difference(reform_share, baseline_share)
            inequality_decile_records.append({
                'label': label,
                'baseline': baseline_share,
                'reform': reform_share,
                'difference': diff,
                'color': decile_palette[(decile - 1) % len(decile_palette)],
            })

    benefits_baseline = baseline_results.get('benefits', {})
    benefits_reform = reform_results.get('benefits', {})
    if benefits_baseline and benefits_reform:
        benefits_decile_records = []
        total_benefits_base = benefits_baseline.get('TotalDirectCashBenefits')
        total_benefits_reform = benefits_reform.get('TotalDirectCashBenefits')
        for decile in range(1, 11):
            label = f"Decile {decile}"
            baseline_share = safe_share(benefits_baseline.get(f"SumBen_InBaselineDec{decile}"), total_benefits_base)
            reform_share = safe_share(benefits_reform.get(f"SumBen_InBaselineDec{decile}"), total_benefits_reform)
            diff = safe_difference(reform_share, baseline_share)
            benefits_decile_records.append({
                'label': label,
                'baseline': baseline_share,
                'reform': reform_share,
                'difference': diff,
                'color': decile_palette[(decile - 1) % len(decile_palette)],
            })

    taxes_baseline = baseline_results.get('taxes', {})
    taxes_reform = reform_results.get('taxes', {})
    if taxes_baseline and taxes_reform:
        dir_tax_decile_records = []
        indir_tax_decile_records = []
        total_dir_tax_base = taxes_baseline.get('TotalDirectTaxes')
        total_dir_tax_reform = taxes_reform.get('TotalDirectTaxes')
        total_indir_tax_base = taxes_baseline.get('TotalIndirectTaxes')
        total_indir_tax_reform = taxes_reform.get('TotalIndirectTaxes')
        for decile in range(1, 11):
            label = f"Decile {decile}"
            baseline_dir = safe_share(taxes_baseline.get(f"SumDirTax_InBaselineDec{decile}"), total_dir_tax_base)
            reform_dir = safe_share(taxes_reform.get(f"SumDirTax_InBaselineDec{decile}"), total_dir_tax_reform)
            dir_diff = safe_difference(reform_dir, baseline_dir)
            dir_tax_decile_records.append({
                'label': label,
                'baseline': baseline_dir,
                'reform': reform_dir,
                'difference': dir_diff,
                'color': decile_palette[(decile - 1) % len(decile_palette)],
            })

            baseline_indir = safe_share(taxes_baseline.get(f"SumIndirTax_InBaselineDec{decile}"), total_indir_tax_base)
            reform_indir = safe_share(taxes_reform.get(f"SumIndirTax_InBaselineDec{decile}"), total_indir_tax_reform)
            indir_diff = safe_difference(reform_indir, baseline_indir)
            indir_tax_decile_records.append({
                'label': label,
                'baseline': baseline_indir,
                'reform': reform_indir,
                'difference': indir_diff,
                'color': decile_palette[(decile - 1) % len(decile_palette)],
            })

    def build_graph_records(records, yaxis_title, bar_color='#0891b2', difference_suffix=" pp"):
        filtered = [
            rec for rec in records
            if rec.get('difference') is not None and not pd.isna(rec.get('difference'))
        ]
        if not filtered:
            return None, None
        labels = [rec['label'] for rec in filtered]
        colors = [bar_color for _ in filtered]
        raw_differences = [float(rec['difference']) for rec in filtered]
        axis_settings = compute_axis_settings(raw_differences)
        decimals = axis_settings['decimals']
        value_decimals = max(decimals, 2)

        diffs = []
        hovertext = []
        diff_suffix = difference_suffix or ""
        for rec, diff_val in zip(filtered, raw_differences):
            diff_display = format_hover(diff_val, 2)
            rec['difference'] = diff_val
            diffs.append(diff_val)
            hovertext.append(
                f"<b>{rec['label']}</b><br>"
                f"Baseline: {format_axis_value(rec['baseline'], value_decimals)}<br>"
                f"Reform: {format_axis_value(rec['reform'], value_decimals)}<br>"
                f"Difference: {diff_display}{diff_suffix}"
            )

        fig = go.Figure(
            data=[
                go.Bar(
                    x=labels,
                    y=diffs,
                    marker_color=colors,
                    marker_opacity=0.85,
                    hovertext=hovertext,
                    hovertemplate="%{hovertext}<extra></extra>",
                )
            ]
        )
        fig.add_hline(y=0, line_color="#94a3b8", line_width=1)
        fig.update_traces(marker_line_color="#ffffff", marker_line_width=0.7)
        fig.update_layout(
            margin=dict(l=60, r=20, t=20, b=100),
            plot_bgcolor="#ffffff",
            paper_bgcolor="#ffffff",
            bargap=0.45,
            showlegend=False,
            font=dict(size=12, color="#1f2a44"),
            hoverlabel=dict(bgcolor="#f8fafc"),
        )
        tickformat = axis_settings['tickformat']
        axis_kwargs = {
            "title": yaxis_title,
            "zeroline": False,
            "gridcolor": "#e5e7eb",
            "tickformat": tickformat,
        }
        if axis_settings.get("dtick"):
            axis_kwargs["dtick"] = axis_settings["dtick"]
        if axis_settings.get("range"):
            axis_kwargs["range"] = axis_settings["range"]
        fig.update_yaxes(**axis_kwargs)
        
        # Check if all differences are negative or zero
        all_negative_or_zero = all(d <= 0 for d in diffs)
        # If all negative/zero, rotate labels to read top-left to bottom-right and move axis to top
        tick_angle = 35 if all_negative_or_zero else -35
        
        xaxis_kwargs = {
            "tickangle": tick_angle,
            "tickfont": dict(size=11),
            "ticks": 'outside',
            "ticklen": 12,
            "tickson": 'labels',
        }
        if all_negative_or_zero:
            xaxis_kwargs["side"] = "top"
        
        fig.update_xaxes(**xaxis_kwargs)
        return fig, filtered

    inequality_graph_definitions = [
        {
            'title': "Change in household income/consumption level at percentiles",
            'yaxis': "Difference in yearly level",
            'records': percentile_records,
            'color': '#0ea5e9',
            'difference_suffix': '',
        },
        {
            'title': "Change in share of total income/consumption by baseline decile (pp.)",
            'yaxis': "Difference in share (pp.)",
            'records': inequality_decile_records,
            'color': '#0ea5e9',
        },
        {
            'title': "Change in share of total cash benefits by baseline decile (pp.)",
            'yaxis': "Difference in benefit share (pp.)",
            'records': benefits_decile_records,
            'color': '#0ea5e9',
        },
        {
            'title': "Change in share of total direct taxes by baseline decile (pp.)",
            'yaxis': "Difference in tax share (pp.)",
            'records': dir_tax_decile_records,
            'color': '#0ea5e9',
        },
        {
            'title': "Change in share of total indirect taxes by baseline decile (pp.)",
            'yaxis': "Difference in tax share (pp.)",
            'records': indir_tax_decile_records,
            'color': '#0ea5e9',
        },
    ]

    inequality_graph_sections = []
    for chart in inequality_graph_definitions:
        fig, filtered_records = build_graph_records(
            chart['records'],
            chart['yaxis'],
            chart.get('color', '#0ea5e9'),
            chart.get('difference_suffix', " pp"),
        )
        if fig is None:
            continue
        inequality_graph_sections.append(build_graph_section(chart['title'], None, fig))
        inequality_graphs_excel_figures.append({
            'title': chart['title'],
            'subtitle': chart['yaxis'],
            'figure': fig,
        })

    if inequality_graph_sections:
        inequality_graphs_content = build_results_accordion(inequality_graph_sections, "inequality-graphs")
    else:
        inequality_graphs_content = [dbc.Alert("No inequality graph data available for this simulation.", color="warning")]

    gainers_losers_content = [run_placeholder]
    gainers_losers_excel_figures = []
    gainers_losers_results = reform_results.get('gainers_losers') if reform_results else None
    if gainers_losers_results:
        metric_display = [
            ('gain1', 'Gain >1%', '#22c55e'),
            ('gain5', 'Gain >5%', '#15803d'),
            ('lose1', 'Loss >1%', '#f97316'),
            ('lose5', 'Loss >5%', '#b91c1c'),
        ]

        def build_gain_loss_figure(group_data, title):
            labels = [row.get('label', '') for row in group_data]
            fig = go.Figure()
            for key, display_name, color in metric_display:
                values = []
                hover_values = []
                for row in group_data:
                    raw_val = row.get(key, 0.0)
                    try:
                        numeric_val = float(raw_val if raw_val is not None else 0.0)
                    except (TypeError, ValueError):
                        numeric_val = 0.0
                    values.append(numeric_val)
                    hover_values.append(numeric_val)
                hovertext = [
                    f"<b>{label}</b><br>{display_name}: {value:.1f}%"
                    if abs(value - round(value)) > 1e-6 else f"<b>{label}</b><br>{display_name}: {int(round(value))}%"
                    for label, value in zip(labels, hover_values)
                ]
                fig.add_bar(
                    x=labels,
                    y=values,
                    name=display_name,
                    marker_color=color,
                    marker_opacity=0.85,
                    hovertext=hovertext,
                    hovertemplate="%{hovertext}<extra></extra>",
                )
            fig.update_layout(
                barmode='group',
                margin=dict(l=60, r=20, t=20, b=100),
                plot_bgcolor="#ffffff",
                paper_bgcolor="#ffffff",
                showlegend=True,
                legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1),
                font=dict(size=12, color="#1f2a44"),
                hoverlabel=dict(bgcolor="#f8fafc"),
            )
            fig.add_hline(y=0, line_color="#94a3b8", line_width=1)
            fig.update_layout(margin=dict(l=60, r=20, t=20, b=110))
            fig.update_yaxes(title="% of population in group", gridcolor="#e5e7eb", rangemode='tozero')
            fig.update_xaxes(
                tickangle=-30,
                tickfont=dict(size=11),
                automargin=True,
                ticks='outside',
                ticklen=12,
                tickson='labels',
                categoryorder='array',
                categoryarray=labels,
            )
            return fig

        gainers_chart_definitions = [
            ("Gainers and losers by baseline decile", gainers_losers_results.get('deciles')),
            ("Gainers and losers by household structure", gainers_losers_results.get('household')),
            ("Gainers and losers by vulnerability and labour market status", gainers_losers_results.get('vulnerability')),
        ]

        gainers_sections = []
        for title, group_data in gainers_chart_definitions:
            if not group_data:
                continue
            fig = build_gain_loss_figure(group_data, title)
            gainers_sections.append(build_graph_section(title, "(% of population group)", fig))
            gainers_losers_excel_figures.append({
                'title': title,
                'subtitle': "(% of population group)",
                'figure': fig,
            })

        # Numeric tables mirroring the three figures (appended after them so the
        # per-position subtitle lookups stay aligned)
        for title, group_data in gainers_chart_definitions:
            if not group_data:
                continue
            table_df = pd.DataFrame([
                {
                    'Group': row.get('label', ''),
                    'Gainers >1%': row.get('gain1', 0.0),
                    'Gainers >5%': row.get('gain5', 0.0),
                    'Losers >1%': row.get('lose1', 0.0),
                    'Losers >5%': row.get('lose5', 0.0),
                }
                for row in group_data
            ])
            gainers_losers_excel_figures.append((table_df, title, []))

        if gainers_sections:
            gainers_losers_content = build_results_accordion(gainers_sections, "gainers-losers")
        else:
            gainers_losers_content = [dbc.Alert("No gainer/loser data available for this simulation.", color="warning")]

    # --- Benefits Tab Content ---
    benefits_tab_content = [run_placeholder]
    benefits_excel_tables = []
    benefits_baseline = baseline_results.get('benefits', {})
    benefits_reform = reform_results.get('benefits', {})
    benefits_sections = []

    if benefits_baseline and benefits_reform:
        def safe_pct(numerator, denominator):
            if numerator is None or denominator is None:
                return None
            if pd.isna(numerator) or pd.isna(denominator):
                return None
            if denominator == 0:
                return None
            return float(numerator) / float(denominator) * 100

        def safe_ratio_percent(numerator, denominator):
            if numerator is None or denominator is None:
                return None
            if pd.isna(numerator) or pd.isna(denominator):
                return None
            if denominator == 0:
                return None
            return float(numerator) / float(denominator) * 100

        def calc_difference(reform_val, baseline_val):
            if reform_val is None or baseline_val is None:
                return None
            if pd.isna(reform_val) or pd.isna(baseline_val):
                return None
            return reform_val - baseline_val

        povline_yearly = (baseline_results.get('poverty', {}).get('povline') or 0) * 12
        baseline_median_cons = baseline_results.get('BaselineMedianEqConsYearly') or 0
        baseline_median_inc = baseline_results.get('BaselineMedianEqIncYearly') or 0

        sections = []

        def begin_section(title):
            sections.append({'title': title, 'rows': [], 'formats': []})

        def add_data_row(label, baseline_val, reform_val, fmt):
            section = sections[-1]
            diff_val = calc_difference(reform_val, baseline_val) if fmt != 'header' else None
            section['rows'].append({'Metric': label, 'Baseline': baseline_val, 'Reform': reform_val, 'Difference': diff_val})
            section['formats'].append(fmt)

        def add_header_row(label):
            add_data_row(label, None, None, 'header')

        def add_rate_row(label, numerator_key, denominator_key):
            denom_val = benefits_baseline.get(denominator_key)
            baseline_val = safe_pct(benefits_baseline.get(numerator_key), denom_val)
            reform_val = safe_pct(benefits_reform.get(numerator_key), denom_val)
            add_data_row(label, baseline_val, reform_val, 'one_dec')

        def add_share_row(label, numerator_key, total_key):
            baseline_val = safe_pct(benefits_baseline.get(numerator_key), benefits_baseline.get(total_key))
            reform_val = safe_pct(benefits_reform.get(numerator_key), benefits_reform.get(total_key))
            add_data_row(label, baseline_val, reform_val, 'one_dec')

        begin_section("Receipt of benefits by household type, % of households")
        add_rate_row("- All households", "Count_AllHH_GetsAnyBen", "TotalHHCount")
        add_rate_row("- Poor households (baseline poverty status)", "CountBaselinePoorHHGetsAnyDirectBenefit_Reform", "CountPoorHH_BaselineDefinition")
        add_rate_row("- Households with children", "Count_ChildHH_GetsAnyBen", "CountHH_AtLeastOneChild")
        add_rate_row("- Households with an elderly member", "Count_ElderlyHH_GetsAnyBen", "CountHH_AtLeastOneElderly")
        add_rate_row("- Households with no male adults", "Count_NoMaleHH_GetsAnyBen", "CountHH_New_NoMaleAdult")
        add_rate_row("- Households with an informal adult", "Count_InformalAdultHH_GetsAnyBen", "CountHH_New_InformalAdult")

        begin_section("Receipt of cash benefits by household type, % of households")
        add_rate_row("- All households", "Count_AllHH_GetsCashBen", "TotalHHCount")
        add_rate_row("- Poor households (baseline poverty status)", "Count_BaselinePoorHH_GetsCashBen", "CountPoorHH_BaselineDefinition")
        add_rate_row("- Households with children", "Count_ChildHH_GetsCashBen", "CountHH_AtLeastOneChild")
        add_rate_row("- Households with an elderly member", "Count_ElderlyHH_GetsCashBen", "CountHH_AtLeastOneElderly")
        add_rate_row("- Households with no male adults", "Count_NoMaleHH_GetsCashBen", "CountHH_New_NoMaleAdult")
        add_rate_row("- Households with an informal adult", "Count_InformalAdultHH_GetsCashBen", "CountHH_New_InformalAdult")

        begin_section("Receipt of in-kind benefits by household type, % of households")
        add_rate_row("- All households", "Count_AllHH_GetsInKindBen", "TotalHHCount")
        add_rate_row("- Poor households (baseline poverty status)", "Count_BaselinePoorHH_GetsInKindBen", "CountPoorHH_BaselineDefinition")
        add_rate_row("- Households with children", "Count_ChildHH_GetsInKindBen", "CountHH_AtLeastOneChild")
        add_rate_row("- Households with an elderly member", "Count_ElderlyHH_GetsInKindBen", "CountHH_AtLeastOneElderly")
        add_rate_row("- Households with no male adults", "Count_NoMaleHH_GetsInKindBen", "CountHH_New_NoMaleAdult")
        add_rate_row("- Households with an informal adult", "Count_InformalAdultHH_GetsInKindBen", "CountHH_New_InformalAdult")

        begin_section("Targeting of poor households (baseline poverty status), % of benefits")
        add_share_row("- Share of benefits received by poor households", "SumAnyDirectBenefitAmount_BaselinePoorHH", "TotalAnyDirectBenefitAmount")
        add_share_row("- Share of cash benefits received by poor households", "SumBen_BaselinePoorHH", "TotalDirectCashBenefits")
        add_share_row("- Share of in-kind benefits received by poor households", "SumInKindBen_BaselinePoorHH", "TotalInKindBenefitAmount")

        cash_mean_baseline = benefits_baseline.get('Mean_eq_indiv_cash_ben_yearly')
        cash_mean_reform = benefits_reform.get('Mean_eq_indiv_cash_ben_yearly')
        inkind_mean_baseline = benefits_baseline.get('Mean_eq_indiv_inkind_ben_yearly')
        inkind_mean_reform = benefits_reform.get('Mean_eq_indiv_inkind_ben_yearly')

        begin_section("Per-capita adequacy")
        add_header_row("Cash benefits")
        add_data_row("- Mean cash benefit amount per beneficiary, yearly", cash_mean_baseline, cash_mean_reform, 'int')
        add_data_row(
            "- Mean cash benefit amount as a share of poverty line, %",
            safe_ratio_percent(cash_mean_baseline, povline_yearly),
            safe_ratio_percent(cash_mean_reform, povline_yearly),
            'one_dec'
        )
        add_data_row(
            "- Mean cash benefit amount as a share of baseline median consumption, %",
            safe_ratio_percent(cash_mean_baseline, baseline_median_cons),
            safe_ratio_percent(cash_mean_reform, baseline_median_cons),
            'one_dec'
        )
        add_data_row(
            "- Mean cash benefit amount as a share of baseline median disposable income, %",
            safe_ratio_percent(cash_mean_baseline, baseline_median_inc),
            safe_ratio_percent(cash_mean_reform, baseline_median_inc),
            'one_dec'
        )
        add_header_row("In-kind benefits")
        add_data_row("- Mean in-kind benefit amount per beneficiary, yearly", inkind_mean_baseline, inkind_mean_reform, 'int')
        add_data_row(
            "- Mean in-kind benefit amount as a share of poverty line, %",
            safe_ratio_percent(inkind_mean_baseline, povline_yearly),
            safe_ratio_percent(inkind_mean_reform, povline_yearly),
            'one_dec'
        )
        add_data_row(
            "- Mean in-kind benefit amount as a share of baseline median consumption, %",
            safe_ratio_percent(inkind_mean_baseline, baseline_median_cons),
            safe_ratio_percent(inkind_mean_reform, baseline_median_cons),
            'one_dec'
        )
        add_data_row(
            "- Mean in-kind benefit amount as a share of baseline median disposable income, %",
            safe_ratio_percent(inkind_mean_baseline, baseline_median_inc),
            safe_ratio_percent(inkind_mean_reform, baseline_median_inc),
            'one_dec'
        )

        begin_section("Distribution of cash benefits across baseline deciles, %")
        for decile in range(1, 11):
            label = f"- Decile {decile}"
            num_key = f"SumBen_InBaselineDec{decile}"
            baseline_val = safe_pct(benefits_baseline.get(num_key), benefits_baseline.get('TotalDirectCashBenefits'))
            reform_val = safe_pct(benefits_reform.get(num_key), benefits_reform.get('TotalDirectCashBenefits'))
            add_data_row(label, baseline_val, reform_val, 'one_dec')
        baseline_total_share = safe_pct(
            benefits_baseline.get('TotalDirectCashBenefits'),
            benefits_baseline.get('TotalDirectCashBenefits')
        )
        reform_total_share = safe_pct(
            benefits_reform.get('TotalDirectCashBenefits'),
            benefits_reform.get('TotalDirectCashBenefits')
        )
        add_data_row("Total", baseline_total_share, reform_total_share, 'one_dec')

        begin_section("Distribution of in-kind benefits across baseline deciles, %")
        for decile in range(1, 11):
            label = f"- Decile {decile}"
            num_key = f"SumInKindBen_InBaselineDec{decile}"
            baseline_val = safe_pct(benefits_baseline.get(num_key), benefits_baseline.get('TotalInKindBenefitAmount'))
            reform_val = safe_pct(benefits_reform.get(num_key), benefits_reform.get('TotalInKindBenefitAmount'))
            add_data_row(label, baseline_val, reform_val, 'one_dec')
        baseline_total_share_inkind = safe_pct(
            benefits_baseline.get('TotalInKindBenefitAmount'),
            benefits_baseline.get('TotalInKindBenefitAmount')
        )
        reform_total_share_inkind = safe_pct(
            benefits_reform.get('TotalInKindBenefitAmount'),
            benefits_reform.get('TotalInKindBenefitAmount')
        )
        add_data_row("Total", baseline_total_share_inkind, reform_total_share_inkind, 'one_dec')

        benefits_tab_content = []
        benefits_excel_tables = []

        for section in sections:
            section_rows = section['rows']
            section_formats = section['formats']
            df = pd.DataFrame(section_rows)
            if 'Metric' in df.columns:
                df = df.rename(columns={'Metric': ''})
            
            # Rename Difference column to Difference (pp.) for all sections except Per-capita adequacy
            if section['title'] != "Per-capita adequacy":
                df = df.rename(columns={'Difference': 'Difference (pp.)'})
            
            df_excel = df.copy()
            df_display = df.copy().astype('object')

            # Determine which column name to use
            diff_col = 'Difference (pp.)' if section['title'] != "Per-capita adequacy" else 'Difference'

            for idx, fmt in enumerate(section_formats):
                if fmt == 'header':
                    for col in ['Baseline', 'Reform', 'Difference', 'Difference (pp.)']:
                        if col in df_display.columns:
                            df_display.at[idx, col] = None
                            df_excel.at[idx, col] = np.nan
                    continue
                if fmt == 'one_dec':
                    for col in ['Baseline', 'Reform']:
                        if col in df_display.columns:
                            df_display.at[idx, col] = format_one_decimal_value(df_excel.at[idx, col])
                    if diff_col in df_display.columns:
                        diff_val = section_rows[idx]['Difference']
                        df_display.at[idx, diff_col] = format_one_decimal_difference(diff_val)
                elif fmt == 'int':
                    for col in ['Baseline', 'Reform']:
                        if col in df_display.columns:
                            df_display.at[idx, col] = format_int_value(df_excel.at[idx, col])
                    if diff_col in df_display.columns:
                        diff_val = section_rows[idx]['Difference']
                        df_display.at[idx, diff_col] = format_int_difference(diff_val)

            benefits_sections.append(create_styled_table(df_display.to_dict('list'), section['title'], ""))
            benefits_excel_tables.append(
                (df_excel, section['title'], [diff_col], "+0.0;-0.0;0.0", "#,##0.0")
            )

    if benefits_sections:
        benefits_tab_content = build_results_accordion(benefits_sections, "benefits")

    # --- Taxes Tab Content ---
    taxes_tab_content = [run_placeholder]
    taxes_excel_tables = []
    taxes_baseline = baseline_results.get('taxes', {})
    taxes_reform = reform_results.get('taxes', {})
    taxes_sections = []

    if taxes_baseline and taxes_reform:
        def safe_pct(numerator, denominator):
            if numerator is None or denominator is None:
                return None
            if pd.isna(numerator) or pd.isna(denominator):
                return None
            if denominator == 0:
                return None
            return float(numerator) / float(denominator) * 100

        def calc_difference(reform_val, baseline_val):
            if reform_val is None or baseline_val is None:
                return None
            if pd.isna(reform_val) or pd.isna(baseline_val):
                return None
            return reform_val - baseline_val

        def format_percentage_table(rows):
            df = pd.DataFrame(rows).replace({None: np.nan})
            if 'Metric' in df.columns:
                df = df.rename(columns={'Metric': ''})
            df_excel = df.copy()
            df_display = df.copy().astype('object')
            for col in ['Baseline (%)', 'Reform (%)']:
                if col in df_display.columns:
                    df_display[col] = df_display[col].apply(format_one_decimal_value)
            if 'Difference (pp.)' in df_display.columns:
                df_display['Difference (pp.)'] = df_display['Difference (pp.)'].apply(format_one_decimal_difference)
            return df_display, df_excel

        taxes_tab_content = []

        payment_row_specs = [
            ('- All households', 'TotalHHCount', 'AllHH'),
            ('- Poor households (baseline poverty status)', 'CountPoorHH_BaselineDefinition', 'BaselinePoorHH'),
            ('- Households with children', 'CountHH_AtLeastOneChild', 'ChildHH'),
            ('- Households with an elderly member', 'CountHH_AtLeastOneElderly', 'ElderlyHH'),
            ('- Households with no male adults', 'CountHH_New_NoMaleAdult', 'NoMaleHH'),
            ('- Households with an informal adult', 'CountHH_New_InformalAdult', 'InformalAdultHH'),
        ]

        def build_payment_table(tax_suffix, title, excel_note):
            rows = []
            for label, denom_key, num_prefix in payment_row_specs:
                numerator_key = f'Count_{num_prefix}_Pays{tax_suffix}'
                baseline_val = safe_pct(taxes_baseline.get(numerator_key), taxes_baseline.get(denom_key))
                reform_val = safe_pct(taxes_reform.get(numerator_key), taxes_baseline.get(denom_key))
                diff_val = calc_difference(reform_val, baseline_val)
                rows.append({
                    'Metric': label,
                    'Baseline (%)': baseline_val,
                    'Reform (%)': reform_val,
                    'Difference (pp.)': diff_val,
                })
            df_display, df_excel = format_percentage_table(rows)
            taxes_sections.append(create_styled_table(df_display.to_dict('list'), title, ""))
            taxes_excel_tables.append((df_excel, excel_note or title, ["Difference (pp.)"], "+0.0;-0.0;0.0", "#,##0.0"))

        def build_etr_table():
            rows = []
            total_income = taxes_baseline.get('TotalOriginalIncome')
            reform_total_income = taxes_reform.get('TotalOriginalIncome')
            for label in ['- Direct taxes as % of original income',
                          '- Direct and indirect taxes as % of original income',
                          '- Direct taxes, indirect taxes, and SIC as % of original income']:
                if label.startswith('- Direct taxes, indirect taxes, and SIC'):
                    base_num = (
                        (taxes_baseline.get('TotalDirectTaxes') or 0) +
                        (taxes_baseline.get('TotalIndirectTaxes') or 0) +
                        (taxes_baseline.get('TotalSICEE_SE') or 0)
                    )
                    reform_num = (
                        (taxes_reform.get('TotalDirectTaxes') or 0) +
                        (taxes_reform.get('TotalIndirectTaxes') or 0) +
                        (taxes_reform.get('TotalSICEE_SE') or 0)
                    )
                elif label.startswith('- Direct and indirect taxes'):
                    base_num = (
                        (taxes_baseline.get('TotalDirectTaxes') or 0) +
                        (taxes_baseline.get('TotalIndirectTaxes') or 0)
                    )
                    reform_num = (
                        (taxes_reform.get('TotalDirectTaxes') or 0) +
                        (taxes_reform.get('TotalIndirectTaxes') or 0)
                    )
                else:
                    base_num = taxes_baseline.get('TotalDirectTaxes')
                    reform_num = taxes_reform.get('TotalDirectTaxes')
                baseline_val = safe_pct(base_num, total_income)
                reform_val = safe_pct(reform_num, reform_total_income)
                diff_val = calc_difference(reform_val, baseline_val)
                rows.append({
                    'Metric': label,
                    'Baseline (%)': baseline_val,
                    'Reform (%)': reform_val,
                    'Difference (pp.)': diff_val,
                })
            df_display, df_excel = format_percentage_table(rows)
            title = "Effective tax rates"
            taxes_sections.append(create_styled_table(df_display.to_dict('list'), title, "(%)"))
            excel_note = f"{title} (share of original income)"
            taxes_excel_tables.append((df_excel, excel_note, ["Difference (pp.)"], "+0.0;-0.0;0.0", "#,##0.0"))

        def build_distribution_table(sum_prefix, total_key, title, excel_note):
            rows = []
            baseline_total = taxes_baseline.get(total_key)
            reform_total = taxes_reform.get(total_key)
            for decile in range(1, 11):
                numerator_key = f'Sum{sum_prefix}_InBaselineDec{decile}'
                baseline_val = safe_pct(taxes_baseline.get(numerator_key), baseline_total)
                reform_val = safe_pct(taxes_reform.get(numerator_key), reform_total)
                diff_val = calc_difference(reform_val, baseline_val)
                rows.append({
                    'Metric': f'- Decile {decile}',
                    'Baseline (%)': baseline_val,
                    'Reform (%)': reform_val,
                    'Difference (pp.)': diff_val,
                })
            baseline_total_share = safe_pct(baseline_total, baseline_total)
            reform_total_share = safe_pct(reform_total, reform_total)
            rows.append({
                'Metric': 'Total',
                'Baseline (%)': baseline_total_share,
                'Reform (%)': reform_total_share,
                'Difference (pp.)': calc_difference(reform_total_share, baseline_total_share)
            })
            df_display, df_excel = format_percentage_table(rows)
            taxes_sections.append(create_styled_table(df_display.to_dict('list'), title, ""))
            taxes_excel_tables.append((df_excel, excel_note or title, ["Difference (pp.)"], "+0.0;-0.0;0.0", "#,##0.0"))

        build_payment_table('DirTax', "Direct tax payment by household type, % of households",
                            "Direct tax payment by household type, % of households (share of households paying direct tax)")
        build_payment_table('IndirTax', "Indirect tax payment by household type, % of households",
                            "Indirect tax payment by household type, % of households (share of households paying indirect tax)")
        build_payment_table('SSC_EE_SE', "Social contribution payment by household type, % of households",
                            "Social contribution payment by household type, % of households (share of households paying employee/self-employed SIC)")
        build_etr_table()
        build_distribution_table('DirTax', 'TotalDirectTaxes', "Distribution of total direct taxes across baseline deciles, %",
                                 "Distribution of total direct taxes across baseline deciles, % (share of total direct taxes)")
        build_distribution_table('IndirTax', 'TotalIndirectTaxes', "Distribution of total indirect taxes across baseline deciles, %",
                                 "Distribution of total indirect taxes across baseline deciles, % (share of total indirect taxes)")
        build_distribution_table('SICEE_SE', 'TotalSICEE_SE', "Distribution of total social contributions across baseline deciles, %",
                                 "Distribution of total social contributions across baseline deciles, % (share of total SIC)")

    if taxes_sections:
        taxes_tab_content = build_results_accordion(taxes_sections, "taxes")

    # --- Policy Effects Tab Content ---
    policy_effects_content = [run_placeholder]
    policy_effects_excel_tables = []

    required_baseline_keys = [
        'PovertyRate_Bef', 'PovertyRate_Aft',
        'Gini_Bef', 'Gini_Aft',
    ]
    required_reform_keys = [
        'PovertyRate_Aft',
        'Gini_Aft',
    ]

    if all(key in baseline_results for key in required_baseline_keys) and all(
        key in reform_results for key in required_reform_keys
    ):
        def cast_numeric(val):
            if val is None:
                return np.nan
            if isinstance(val, str) and not val.strip():
                return np.nan
            if pd.isna(val):
                return np.nan
            try:
                return float(val)
            except (TypeError, ValueError):
                return np.nan

        def compute_effects(before_val, after_baseline_val, after_reform_val):
            before_num = cast_numeric(before_val)
            after_baseline_num = cast_numeric(after_baseline_val)
            after_reform_num = cast_numeric(after_reform_val)
            baseline_effect = np.nan
            reform_effect = np.nan
            if not np.isnan(before_num) and not np.isnan(after_baseline_num):
                baseline_effect = after_baseline_num - before_num
            if not np.isnan(before_num) and not np.isnan(after_reform_num):
                reform_effect = after_reform_num - before_num

            def effect_percent(effect):
                if np.isnan(effect) or np.isnan(before_num) or before_num == 0:
                    return np.nan
                return (effect / before_num) * 100

            return baseline_effect, reform_effect, effect_percent(baseline_effect), effect_percent(reform_effect)

        policy_rows = []
        row_formats = []

        def add_header(label):
            policy_rows.append({'Metric': label, 'Baseline': np.nan, 'Reform': np.nan, 'Difference': np.nan})
            row_formats.append('header')

        def add_data_row(label, baseline_val, reform_val):
            baseline_num = cast_numeric(baseline_val)
            reform_num = cast_numeric(reform_val)
            if np.isnan(baseline_num):
                baseline_num = np.nan
            if np.isnan(reform_num):
                reform_num = np.nan
            if np.isnan(baseline_num) or np.isnan(reform_num):
                diff_val = np.nan
            else:
                diff_val = reform_num - baseline_num
            policy_rows.append({'Metric': label, 'Baseline': baseline_num, 'Reform': reform_num, 'Difference': diff_val})
            row_formats.append('two_dec')

        # Poverty rate
        poverty_rate_bef = baseline_results.get('PovertyRate_Bef')
        poverty_rate_aft_baseline = baseline_results.get('PovertyRate_Aft')
        poverty_rate_aft_reform = reform_results.get('PovertyRate_Aft')
        baseline_effect_pp, reform_effect_pp, _, _ = compute_effects(
            poverty_rate_bef, poverty_rate_aft_baseline, poverty_rate_aft_reform
        )

        add_header("Poverty rate")
        add_data_row("- Before taxes and benefits, %", poverty_rate_bef, poverty_rate_bef)
        add_data_row("- After taxes and benefits, %", poverty_rate_aft_baseline, poverty_rate_aft_reform)
        add_data_row("- Effects of tax-benefit system on the poverty rate, pp.", baseline_effect_pp, reform_effect_pp)

        # Gini coefficient
        gini_bef = baseline_results.get('Gini_Bef')
        gini_aft_baseline = baseline_results.get('Gini_Aft')
        gini_aft_reform = reform_results.get('Gini_Aft')
        gini_effect_pp_baseline, gini_effect_pp_reform, _, _ = compute_effects(
            gini_bef, gini_aft_baseline, gini_aft_reform
        )

        add_header("Gini coefficient")
        add_data_row("- Before taxes and benefits", gini_bef, gini_bef)
        add_data_row("- After taxes and benefits", gini_aft_baseline, gini_aft_reform)
        add_data_row("- Effects of tax-benefit system on the Gini coefficient, pp.", gini_effect_pp_baseline, gini_effect_pp_reform)

        policy_df = pd.DataFrame(policy_rows)
        if 'Metric' in policy_df.columns:
            policy_df = policy_df.rename(columns={'Metric': ''})
        policy_df_excel = policy_df.copy()
        policy_df_display = policy_df.copy().astype('object')

        for idx, fmt in enumerate(row_formats):
            if fmt == 'header':
                for col in ['Baseline', 'Reform', 'Difference']:
                    policy_df_display.at[idx, col] = None
                continue
            for col in ['Baseline', 'Reform']:
                raw_val = policy_df_excel.at[idx, col]
                policy_df_display.at[idx, col] = format_two_decimal_value(raw_val)
            diff_val = policy_df_excel.at[idx, 'Difference']
            policy_df_display.at[idx, 'Difference'] = format_signed_value(diff_val) if pd.notna(diff_val) else ""

        policy_effects_section = create_styled_table(
            policy_df_display.to_dict('list'),
            "Redistributive effect of the tax-benefit system",
            ""
        )
        policy_effects_content = build_results_accordion([policy_effects_section], "policy-effects")
        policy_effects_excel_tables.append(
            (
                policy_df_excel,
                "Redistributive effect of the tax-benefit system",
                ["Difference"],
                "+0.00;-0.00;0.00",
                "#,##0.00",
            )
        )

    # --- Placeholder tabs ---
    placeholder_content = [dev_placeholder]
    if inequality_graphs_content == [run_placeholder]:
        inequality_graphs_content = placeholder_content
    if benefits_tab_content == [dev_placeholder]:
        benefits_tab_content = placeholder_content
    if policy_effects_content == [run_placeholder]:
        policy_effects_content = placeholder_content
    if gainers_losers_content == [run_placeholder]:
        gainers_losers_content = placeholder_content
    
    # --- Prepare Download Data ---
    download_output = dash.no_update
    
    if generate_excel:
        try:
            generation_dt = datetime.now()
            generation_date = generation_dt.strftime("%Y-%m-%d_%H-%M")
            generation_display = generation_dt.strftime("%Y-%m-%d %H:%M")

            policy_changes_lines = []
            for section in POLICY_PARAM_SECTIONS:
                prefix = section.get('prefix', '')
                for param_id, label in section['items']:
                    baseline_val = BASELINE_PARAMS.get(param_id)
                    reform_val = reform_params.get(param_id, baseline_val)
                    if policy_values_equal(param_id, baseline_val, reform_val):
                        continue
                    baseline_display = format_policy_value(param_id, baseline_val)
                    reform_display = format_policy_value(param_id, reform_val)
                    display_label = f"{prefix}{label}:"
                    policy_changes_lines.append(f"{display_label} {baseline_display} -> {reform_display}")

            baseline_vat_count = len(BASELINE_VAT_STD_RATE_ITEMS)
            reform_vat_count = len(selected_vat_items)
            baseline_vat_set = set(BASELINE_VAT_STD_RATE_ITEMS)
            reform_vat_set = set(selected_vat_items)
            if reform_vat_set != baseline_vat_set:
                policy_changes_lines.append(
                    f"Value-added tax (VAT) – Standard-rated items count: {baseline_vat_count} -> {reform_vat_count}"
                )
                newly_standard_rated = sorted(reform_vat_set - baseline_vat_set)
                newly_exempt = sorted(baseline_vat_set - reform_vat_set)
                if newly_standard_rated:
                    policy_changes_lines.append(
                        "  Added to standard-rated list: " + "; ".join(
                            VAT_ITEM_MAP.get(item, {}).get('label', item) for item in newly_standard_rated
                        )
                    )
                if newly_exempt:
                    policy_changes_lines.append(
                        "  Moved to exemptions: " + "; ".join(
                            VAT_ITEM_MAP.get(item, {}).get('label', item) for item in newly_exempt
                        )
                    )

            policy_changes_lines = [line for line in policy_changes_lines if line]

            info_rows = [
                {"Field": "Reform name", "Value": reform_name or DEFAULT_REFORM_NAME},
                {"Field": "Distribution statistic", "Value": distribution_label},
                {"Field": "Date/time generated", "Value": generation_display},
                {"Field": "Input file", "Value": INPUT_FILE},
                {"Field": "Baseline system", "Value": "2023"},
            ]
            if policy_changes_lines:
                info_rows.append({"Field": "Policy changes", "Value": policy_changes_lines[0]})
                for line in policy_changes_lines[1:]:
                    info_rows.append({"Field": "", "Value": line})
            else:
                info_rows.append({"Field": "Policy changes", "Value": "None"})

            info_df = pd.DataFrame(info_rows)

            output_stream = BytesIO()
            with pd.ExcelWriter(output_stream, engine='openpyxl') as writer:
                info_df.to_excel(writer, sheet_name='Info', index=False, header=False)

                section_map = {
                    'taxbenpol': ('TaxBenPolicy', 'taxbenpol', [
        (abs_df_excel, "Total revenue and expenditure (yearly, millions of national currency)", ["Difference"], "+#,##0;-#,##0;0", "#,##0"),
        (share_df_excel, "Shares of total revenue and expenditure (%)", ["Difference (pp.)"]),
                    ]),
                    'households': ('Households', 'households', [
                        (households_table1_df_excel, "Taxpayer and benefit recipient households (number of households)", ["Difference"], "+#,##0;-#,##0;0", "#,##0"),
                        (households_table2_df_excel, "Household categories (number of households)", ["Difference"], "+#,##0;-#,##0;0", "#,##0"),
                        (households_table3_df_excel, "Household decile distribution (number of households)", ["Difference"], "+#,##0;-#,##0;0", "#,##0"),
                    ]),
                    'individuals': ('Individuals', 'individuals', [
                        (individuals_table1_df_excel, "Taxpayer and benefit recipient individuals (number of individuals)", ["Difference"], "+#,##0;-#,##0;0", "#,##0"),
                        (individuals_table2_df_excel, "Individual categories (number of individuals)", ["Difference"], "+#,##0;-#,##0;0", "#,##0"),
                        (individuals_table3_df_excel, "Individual decile distribution (number of individuals)", ["Difference"], "+#,##0;-#,##0;0", "#,##0"),
                    ]),
                    'poverty': ('Poverty', 'poverty', [
        (pov_rate_df_excel, "Poverty rate (share of poor population, %)", ["Difference (pp.)"]),
        (pov_gap_df_excel, "Poverty gap (average normalised poverty gap, %)", ["Difference (pp.)"]),
                    ]),
                    'inequality': ('Inequality', 'inequality', inequality_excel_tables),
                }
                if poverty_graphs_excel_figures:
                    section_map['poverty-graphs'] = ('Poverty_Graphs', 'poverty-graphs', poverty_graphs_excel_figures)
                if inequality_graphs_excel_figures:
                    section_map['inequality-graphs'] = ('Inequality_Graphs', 'inequality-graphs', inequality_graphs_excel_figures)
                if benefits_excel_tables:
                    section_map['benefits'] = ('Benefits', 'benefits', benefits_excel_tables)
                if taxes_excel_tables:
                    section_map['taxes'] = ('Taxes', 'taxes', taxes_excel_tables)
                if policy_effects_excel_tables:
                    section_map['policy-effects'] = ('Policy_Effects', 'policy-effects', policy_effects_excel_tables)
                if gainers_losers_excel_figures:
                    section_map['gainers-losers'] = ('Gainers_Losers', 'gainers-losers', gainers_losers_excel_figures)

                ordered_sections = [
                    ('TaxBenPolicy', 'taxbenpol'),
                    ('Households', 'households'),
                    ('Individuals', 'individuals'),
                    ('Poverty', 'poverty'),
                    ('Poverty_Graphs', 'poverty-graphs'),
                    ('Inequality', 'inequality'),
                    ('Inequality_Graphs', 'inequality-graphs'),
                    ('Benefits', 'benefits'),
                    ('Taxes', 'taxes'),
                    ('Policy_Effects', 'policy-effects'),
                    ('Gainers_Losers', 'gainers-losers'),
                ]

                table_specs = [section_map[key] for _, key in ordered_sections if key in section_map]

                sheet_meta = {}
                for sheet_name, info_key, tables in table_specs:
                    description_lines = []
                    if info_key:
                        description_lines = extract_description_lines(info_key)
                    sheet_meta[sheet_name] = {
                        'sections': [],
                        'description_lines': description_lines,
                    }
                    start_row = 0
                    for table_entry in tables:
                        if isinstance(table_entry, dict) and table_entry.get('figure') is not None:
                            # Define subtitles for each sheet and table position (same as for tables)
                            sheet_subtitles = {
                                'TaxBenPolicy': ['Yearly, millions of national currency', '% of total revenue/expenditure'],
                                'Households': ['Number of households', 'Number of households', 'Number of households'],
                                'Individuals': ['Number of individuals', 'Number of individuals', 'Number of individuals'],
                                'Poverty': ['Share of poor population (%)', 'Average normalised poverty gap (%)'],
                                'Poverty_Graphs': ['Difference from baseline (pp.)', 'Difference from baseline (pp.)', 'Difference from baseline (pp.)', 'Difference from baseline (pp.)'],
                                'Inequality': ['', 'Yearly level', 'Share of total (%)'],  # First table has no subtitle
                                'Inequality_Graphs': ['Difference in yearly level from baseline', 'Difference in share from baseline (pp.)', 'Difference in share from baseline (pp.)', 'Difference in share from baseline (pp.)', 'Difference in share from baseline (pp.)'],
                                'Benefits': ['Share of households (%)', 'Share of households (%)', 'Share of households (%)', 'Share of benefits (%)', 'Yearly amounts and shares', 'Share of total (%)', 'Share of total (%)'],
                                'Taxes': ['Share of households (%)', 'Share of households (%)', 'Share of households (%)', 'Share of original income (%)', 'Share of total (%)', 'Share of total (%)', 'Share of total (%)'],
                                'Policy_Effects': ['Outcomes before and after taxes and benefits'],
                                'Gainers_Losers': ['Share of population in group (%)', 'Share of population in group (%)', 'Share of population in group (%)', 'Share of population in group (%)']
                            }
                            
                            def extract_clean_title(title_str):
                                """Extract clean main title by removing subtitle patterns"""
                                if not title_str:
                                    return ""
                                
                                # Remove content in parentheses first
                                if ' (' in title_str:
                                    title_str = title_str.split(' (')[0]
                                
                                # Remove common subtitle patterns
                                patterns_to_remove = [
                                    ', yearly',
                                    ', %',
                                    ', % of households',
                                    ', % of benefits'
                                ]
                                
                                for pattern in patterns_to_remove:
                                    if title_str.endswith(pattern):
                                        title_str = title_str[:-len(pattern)]
                                        break
                                
                                return title_str.strip()
                            
                            fig = table_entry['figure']
                            note = table_entry.get('title', '')
                            
                            # Extract clean main title
                            base_title = extract_clean_title(note)
                            
                            # Get subtitle based on sheet name and figure position
                            figure_position = len(sheet_meta[sheet_name]['sections'])
                            subtitles_for_sheet = sheet_subtitles.get(sheet_name, [])
                            subtitle_text = subtitles_for_sheet[figure_position] if figure_position < len(subtitles_for_sheet) else ""
                            
                            ws = writer.sheets.get(sheet_name)
                            if ws is None:
                                ws = writer.book.create_sheet(title=sheet_name)
                                writer.sheets[sheet_name] = ws
                            title_row = start_row + 1
                            subtitle_row = title_row + 1 if subtitle_text else None
                            image_row = (subtitle_row or title_row) + 1

                            title_cell = ws.cell(row=title_row, column=1, value=base_title)
                            title_cell.font = Font(bold=True, italic=True, size=13, color="000000")
                            if subtitle_text:
                                subtitle_cell = ws.cell(row=subtitle_row, column=1, value=subtitle_text)
                                subtitle_cell.font = Font(italic=True, color="1f2937")

                            rows_occupied = 6
                            try:
                                image_bytes = fig.to_image(format="png", scale=2)
                                image_stream = BytesIO(image_bytes)
                                image_stream.seek(0)
                                xl_image = XLImage(image_stream)
                                xl_image.anchor = f"A{image_row}"
                                # set width for consistency with UI; adjust height proportionally
                                target_width = 720
                                if xl_image.width != 0:
                                    scale_factor = target_width / xl_image.width
                                    xl_image.width = target_width
                                    xl_image.height = xl_image.height * scale_factor
                                ws.add_image(xl_image)
                                points_height = xl_image.height * 0.75  # convert px to points
                                rows_occupied = max(6, int(points_height / 15) + 2)
                            except Exception as exc:
                                print(f"Failed to render poverty graph '{note}': {exc}")
                                warning_cell = ws.cell(row=image_row, column=1, value="Graph preview unavailable.")
                                warning_cell.font = Font(italic=True, color="9ca3af")

                            section_last_row = image_row + rows_occupied
                            start_row = section_last_row + 4
                            sheet_meta[sheet_name]['sections'].append({
                                'type': 'figure',
                                'start_row': title_row,
                                'title_row': title_row,
                                'subtitle_row': subtitle_row,
                                'col_count': 1,
                                'row_count': section_last_row - title_row + 1,
                                'column_names': [],
                                'diff_column_set': set(),
                                'value_format': None,
                                'diff_format': None,
                                'column_decimals': {},
                                'last_row': section_last_row,
                            })
                            continue

                        if len(table_entry) == 3:
                            df_excel, note, diff_columns = table_entry
                            diff_format = "+0.00;-0.00;0.00"
                            value_format = "#,##0.00"
                        elif len(table_entry) == 4:
                            df_excel, note, diff_columns, diff_format = table_entry
                            value_format = "#,##0.00"
                        else:
                            df_excel, note, diff_columns, diff_format, value_format = table_entry
                        if not diff_format:
                            diff_format = "+0.00;-0.00;0.00"
                        if not value_format:
                            value_format = "#,##0.00"
                        title_row = start_row + 1
                        subtitle_row = start_row + 2
                        data_start = start_row + 3

                        ws = writer.sheets.get(sheet_name)
                        if ws is None:
                            df_excel.to_excel(writer, sheet_name=sheet_name, index=False, startrow=data_start)
                            ws = writer.sheets[sheet_name]
                        else:
                            df_excel.to_excel(writer, sheet_name=sheet_name, index=False, startrow=data_start)

                        diff_column_set = set(diff_columns)
                        column_names = list(df_excel.columns)
                        column_decimals = {}
                        for col_name in column_names:
                            if not col_name or col_name not in df_excel.columns:
                                continue
                            if not pd.api.types.is_numeric_dtype(df_excel[col_name]):
                                continue
                            column_decimals[col_name] = count_max_decimals(df_excel[col_name])

                        for col_idx, col_name in enumerate(column_names, start=1):
                            if col_name not in column_decimals and col_name not in diff_column_set:
                                continue
                            is_diff_col = col_name in diff_column_set
                            decimals = column_decimals.get(col_name, 0)
                            fmt = derive_excel_number_format(
                                diff_format if is_diff_col else value_format,
                                decimals,
                                signed=is_diff_col
                            )
                            for row_idx in range(data_start + 1, data_start + df_excel.shape[0] + 1):
                                cell = ws.cell(row=row_idx, column=col_idx)
                                if isinstance(cell.value, (int, float)):
                                    cell.alignment = Alignment(horizontal='right', vertical='center')
                                    cell.number_format = fmt

                        # Define subtitles for each sheet and table position
                        sheet_subtitles = {
                            'TaxBenPolicy': ['Yearly, millions of national currency', '% of total revenue/expenditure'],
                            'Households': ['Number of households', 'Number of households', 'Number of households'],
                            'Individuals': ['Number of individuals', 'Number of individuals', 'Number of individuals'],
                            'Poverty': ['Share of poor population (%)', 'Average normalised poverty gap (%)'],
                            'Poverty_Graphs': ['Difference from baseline (pp.)', 'Difference from baseline (pp.)', 'Difference from baseline (pp.)', 'Difference from baseline (pp.)'],
                            'Inequality': ['', 'Yearly level', 'Share of total (%)'],  # First table has no subtitle
                            'Inequality_Graphs': ['Difference in yearly level from baseline', 'Difference in share from baseline (pp.)', 'Difference in share from baseline (pp.)', 'Difference in share from baseline (pp.)', 'Difference in share from baseline (pp.)'],
                            'Benefits': ['Share of households (%)', 'Share of households (%)', 'Share of households (%)', 'Share of benefits (%)', 'Yearly amounts and shares', 'Share of total (%)', 'Share of total (%)'],
                            'Taxes': ['Share of households (%)', 'Share of households (%)', 'Share of households (%)', 'Share of original income (%)', 'Share of total (%)', 'Share of total (%)', 'Share of total (%)'],
                            'Policy_Effects': ['Outcomes before and after taxes and benefits'],
                            'Gainers_Losers': ['Share of population in group (%)', 'Share of population in group (%)', 'Share of population in group (%)', 'Share of population in group (%)']
                        }
                        
                        def extract_clean_title(title_str):
                            """Extract clean main title by removing subtitle patterns"""
                            if not title_str:
                                return ""
                            
                            # Remove content in parentheses first
                            if ' (' in title_str:
                                title_str = title_str.split(' (')[0]
                            
                            # Remove common subtitle patterns
                            patterns_to_remove = [
                                ', yearly',
                                ', %',
                                ', % of households',
                                ', % of benefits'
                            ]
                            
                            for pattern in patterns_to_remove:
                                if title_str.endswith(pattern):
                                    title_str = title_str[:-len(pattern)]
                                    break
                            
                            return title_str.strip()
                        
                        note_str = note or ""
                        # Extract clean main title
                        base_title = extract_clean_title(note_str)
                        
                        # Get subtitle based on sheet name and table position
                        table_position = len([sec for sec in sheet_meta[sheet_name]['sections'] if sec.get('type') != 'figure'])
                        subtitles_for_sheet = sheet_subtitles.get(sheet_name, [])
                        subtitle_text = subtitles_for_sheet[table_position] if table_position < len(subtitles_for_sheet) else ""
                        
                        title_cell = ws.cell(row=title_row, column=1, value=base_title)
                        title_cell.font = Font(bold=True, italic=True, size=13, color="000000")
                        subtitle_cell = ws.cell(row=subtitle_row, column=1, value=subtitle_text)
                        if subtitle_text:
                            subtitle_cell.font = Font(italic=True, color="1f2937")
                        else:
                            subtitle_cell.font = Font(italic=True, color="1f2937")

                        sheet_meta[sheet_name]['sections'].append({
                            'start_row': data_start,
                            'title_row': title_row,
                            'subtitle_row': subtitle_row,
                            'col_count': df_excel.shape[1],
                            'row_count': df_excel.shape[0] + 1,
                            'note': note,
                            'column_names': column_names,
                            'diff_column_set': diff_column_set,
                            'value_format': value_format,
                            'diff_format': diff_format,
                            'column_decimals': column_decimals,
                        })
                        start_row = data_start + df_excel.shape[0] + 4

                # Placeholder tabs
                existing_sheet_names = {spec[0] for spec in table_specs}
                placeholder_sheets = [
                    sheet_name for sheet_name, key in ordered_sections
                    if sheet_name not in existing_sheet_names
                ]
                placeholder_df = pd.DataFrame(["Output for this tab is under development."])
                for sheet in placeholder_sheets:
                    placeholder_df.to_excel(writer, sheet_name=sheet, index=False, header=False)

                wb = writer.book
                info_ws = wb['Info']
                info_ws.column_dimensions['A'].width = 24
                info_ws.column_dimensions['B'].width = 70
                for row in info_ws.iter_rows(min_row=1, max_col=1):
                    for cell in row:
                        cell.font = Font(bold=True)

                for sheet_name, meta in sheet_meta.items():
                    if sheet_name not in wb.sheetnames:
                        continue
                    ws = wb[sheet_name]
                    ws.column_dimensions['A'].width = 44
                    table_sections = [sec for sec in meta['sections'] if sec.get('type') != 'figure']
                    if table_sections:
                        max_cols = max(section['col_count'] for section in table_sections)
                        for col_idx in range(2, max_cols + 1):
                            col_letter = get_column_letter(col_idx)
                            ws.column_dimensions[col_letter].width = 18
                    for section in table_sections:
                        column_names = section.get('column_names', [])
                        diff_column_set = section.get('diff_column_set', set())
                        value_format_template = section.get('value_format')
                        diff_format_template = section.get('diff_format')
                        column_decimals = section.get('column_decimals', {})
                        start_row = section['start_row']
                        col_count = section['col_count']
                        row_count = section['row_count']
                        note = section['note']

                        header_row = start_row + 1
                        for col_idx in range(2, col_count + 1):
                            header_cell = ws.cell(row=header_row, column=col_idx)
                            header_cell.alignment = Alignment(horizontal='right', vertical='center')
                            header_cell.font = Font(bold=True)
                            col_name = column_names[col_idx - 1] if col_idx - 1 < len(column_names) else None
                            is_diff_col = col_name in diff_column_set if col_name else False
                            decimals = column_decimals.get(col_name, 0)
                            fmt = derive_excel_number_format(
                                diff_format_template if is_diff_col else value_format_template,
                                decimals,
                                signed=is_diff_col
                            )
                            for data_row in range(header_row + 1, header_row + row_count):
                                cell = ws.cell(row=data_row, column=col_idx)
                                if isinstance(cell.value, (int, float)):
                                    cell.alignment = Alignment(horizontal='right', vertical='center')
                                    cell.number_format = fmt
                        last_row = header_row + row_count - 1
                        for data_row in range(header_row + 1, header_row + row_count):
                            first_cell = ws.cell(row=data_row, column=1)
                            first_val = (first_cell.value or "").strip()
                            is_section_header = first_val in TABLE_HEADER_LABELS
                            is_strong = first_val in TABLE_STRONG_LABELS
                            if is_section_header or is_strong:
                                for col_idx in range(1, col_count + 1):
                                    ws.cell(row=data_row, column=col_idx).font = Font(bold=True)
                            if is_section_header:
                                for col_idx in range(2, col_count + 1):
                                    ws.cell(row=data_row, column=col_idx).value = None
                                ws.merge_cells(start_row=data_row, start_column=1, end_row=data_row, end_column=col_count)
                                merged_cell = ws.cell(row=data_row, column=1)
                                merged_cell.alignment = Alignment(horizontal='left', vertical='center')
                            elif is_strong:
                                first_cell.alignment = Alignment(horizontal='left', vertical='center')

                        section['last_row'] = last_row

                    description_lines = meta.get('description_lines', [])
                    if description_lines:
                        last_used_row = max(sec.get('last_row', sec['start_row'] + sec['row_count'] - 1) for sec in meta['sections'])
                        desc_row = last_used_row + 2
                        ws.cell(row=desc_row, column=1, value="Description:").font = Font(bold=True, italic=True, size=13)
                        desc_row += 1
                        for style, text in description_lines:
                            cell = ws.cell(row=desc_row, column=1, value=text)
                            if style == 'header':
                                cell.font = Font(bold=True, color="1D4ED8")
                            else:
                                cell.font = Font(italic=False)
                            desc_row += 1

            excel_data = output_stream.getvalue()
            
            download_output = dcc.send_bytes(excel_data, f"DEVMOD_online_output_{generation_date}.xlsx")
            
        except Exception as e:
            print(f"Error generating Excel file: {e}")
            # If Excel fails, don't crash the whole app
            download_output = dash.no_update 
            # Optionally, update the loading message to show an error
            # ...

    results_title_text = RESULTS_TITLE_TEXT

    return (
        tab1_content, 
        tab3_content, 
        tab4_content,
        tab2_content,
        poverty_graphs_content,
        inequality_table_content,
        inequality_graphs_content,
        benefits_tab_content,
        taxes_tab_content,
        policy_effects_content,
        gainers_losers_content,
        "Simulation complete.",
        results_title_text,
        policy_changes_data,
        download_output
    )

# Show the clear-results button only once results exist
@app.callback(
    Output('clear-results-button', 'style'),
    Input('policy-changes-data', 'data'),
)
def update_clear_results_button(policy_data):
    if policy_data and policy_data.get('sections'):
        return {'display': 'inline-flex'}
    return {'display': 'none'}


# The preset hover preview sits under the "run a simulation" note while the
# results area is empty, and over the top of the first table once it is not
@app.callback(
    Output('results-content-wrapper', 'className'),
    Input('policy-changes-data', 'data'),
)
def update_results_content_state(policy_data):
    state = 'results-loaded' if policy_data and policy_data.get('sections') else 'results-empty'
    return f"results-content-wrapper {state}"


@app.callback(
    Output('policy-changes-modal', 'is_open'),
    Output('policy-changes-modal-title', 'children'),
    Output('policy-changes-modal-body', 'children'),
    Input('tab-policy-changes-button', 'n_clicks'),
    Input('close-policy-changes-modal', 'n_clicks'),
    State('policy-changes-modal', 'is_open'),
    State({'type': 'param-input', 'index': ALL}, 'id'),
    State({'type': 'param-input', 'index': ALL}, 'value'),
    State('vat-checklist', 'value'),
)
def toggle_policy_changes_modal(button_click, close_click, is_open,
                                param_ids, param_values, vat_checklist_value):
    ctx = dash.callback_context
    if not ctx.triggered:
        return is_open, dash.no_update, dash.no_update

    trigger = ctx.triggered[0]['prop_id']
    if "close-policy-changes-modal" in trigger:
        return False, dash.no_update, dash.no_update

    if "tab-policy-changes-button" in trigger:
        # Build the comparison live from the current inputs so the modal works
        # before any simulation has been run
        reform_params, _, added_exemptions, removed_exemptions = collect_reform_params(
            param_ids, param_values, vat_checklist_value
        )
        policy_data = build_policy_changes_data(
            reform_params, None, added_exemptions, removed_exemptions
        )
        body_children = build_policy_changes_modal_body(policy_data)
        return True, "Baseline and reform parameters", body_children

    return is_open, dash.no_update, dash.no_update


# Preset reform toggles + parameter reset, handled client-side so the values
# and their highlights change with no server round trip. Presets act as layers:
# toggling one ON writes only its own parameters (other edits survive); toggling
# it OFF returns only its own parameters to baseline. 'Reset all' restores
# everything, including the VAT item list.
app.clientside_callback(
    """
    function(taxClicks, benClicks, resetClicks, presetState, paramIds, metaStore,
             presetDefs, baselineVat) {
        const nu = window.dash_clientside.no_update;
        const ctx = window.dash_clientside.callback_context;
        if (!ctx || !ctx.triggered || !ctx.triggered.length) return nu;
        const trigger = ctx.triggered[0].prop_id.split('.')[0];
        const ids = paramIds || [];
        const meta = metaStore || {};
        const state = Object.assign({tax: false, benefits: false}, presetState || {});

        // Mirrors format_param_value on the server
        const format = function (pid, value) {
            const m = meta[pid] || {};
            const precision = (m.precision === null || m.precision === undefined) ? 2 : m.precision;
            let out = Number(value).toLocaleString('en-US', {
                minimumFractionDigits: precision,
                maximumFractionDigits: precision,
                useGrouping: !!m.thousands,
            });
            if (m.strip_trailing !== false && out.indexOf('.') !== -1) {
                out = out.replace(/0+$/, '').replace(/\.$/, '');
            }
            return out;
        };

        let values = ids.map(function () { return nu; });
        let vat = nu;

        if (trigger === 'preset-reset-button') {
            state.tax = false;
            state.benefits = false;
            values = ids.map(function (cid) {
                const pid = cid && cid.index;
                if (!pid || !(pid in meta)) return nu;
                return format(pid, meta[pid].baseline);
            });
            vat = (baselineVat || []).slice();
        } else if (trigger === 'preset-tax-button' || trigger === 'preset-benefits-button') {
            const key = (trigger === 'preset-tax-button') ? 'tax' : 'benefits';
            const turningOn = !state[key];
            state[key] = turningOn;
            const presetParams = (presetDefs || {})[key] || {};
            values = ids.map(function (cid) {
                const pid = cid && cid.index;
                if (!pid || !(pid in presetParams)) return nu;
                const target = turningOn ? presetParams[pid]
                                         : (meta[pid] ? meta[pid].baseline : presetParams[pid]);
                return format(pid, target);
            });
        } else {
            return nu;
        }

        return [
            state,
            state.tax ? 'preset-btn preset-btn-active' : 'preset-btn',
            state.benefits ? 'preset-btn preset-btn-active' : 'preset-btn',
            values,
            vat,
        ];
    }
    """,
    Output('preset-state-store', 'data'),
    Output('preset-tax-button', 'className'),
    Output('preset-benefits-button', 'className'),
    Output({'type': 'param-input', 'index': ALL}, 'value', allow_duplicate=True),
    Output('vat-checklist', 'value', allow_duplicate=True),
    Input('preset-tax-button', 'n_clicks'),
    Input('preset-benefits-button', 'n_clicks'),
    Input('preset-reset-button', 'n_clicks'),
    State('preset-state-store', 'data'),
    State({'type': 'param-input', 'index': ALL}, 'id'),
    State('param-meta-store', 'data'),
    State('preset-defs-store', 'data'),
    State('vat-baseline-store', 'data'),
    prevent_initial_call=True,
)


# Clear results and reset every parameter without reloading the page
@app.callback(
    [Output(f'tab-{tab_name}', 'children', allow_duplicate=True) for tab_name in
     ['taxbenpol', 'households', 'individuals', 'poverty', 'poverty-graphs',
      'inequality', 'inequality-graphs', 'benefits', 'taxes',
      'policy-effects', 'gainers-losers']],
    Output('loading-output', 'children', allow_duplicate=True),
    Output('results-title', 'children', allow_duplicate=True),
    Output('policy-changes-data', 'data', allow_duplicate=True),
    Output('preset-state-store', 'data', allow_duplicate=True),
    Output('preset-tax-button', 'className', allow_duplicate=True),
    Output('preset-benefits-button', 'className', allow_duplicate=True),
    Output({'type': 'param-input', 'index': ALL}, 'value', allow_duplicate=True),
    Output('vat-checklist', 'value', allow_duplicate=True),
    Input('clear-results-button', 'n_clicks'),
    State({'type': 'param-input', 'index': ALL}, 'id'),
    prevent_initial_call=True,
)
def clear_results_and_parameters(n_clicks, param_ids):
    if not n_clicks:
        raise dash.exceptions.PreventUpdate
    run_placeholder = html.Div(dbc.Alert("Run a simulation to see results.", color="info"), className="p-4")
    values = []
    for pid in param_ids or []:
        key = pid.get('index') if isinstance(pid, dict) else None
        values.append(format_param_value(key, BASELINE_PARAMS.get(key, 0)) if key else dash.no_update)
    return ([run_placeholder] * 11 +
            ["", RESULTS_TITLE_PLACEHOLDER, None,
             {'tax': False, 'benefits': False}, "preset-btn", "preset-btn",
             values, list(BASELINE_VAT_STD_RATE_ITEMS)])


# --- MAIN EXECUTION ---
if __name__ == '__main__':
    app.run_server(debug=True, port=8051)
